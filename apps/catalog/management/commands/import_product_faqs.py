"""
Import product-bound Technical Inquiry FAQ rows from the FAQ workbook.

Usage:
    python manage.py import_product_faqs
    python manage.py import_product_faqs --excel "F:/菱威仓管/protaylor_product_technical_inquiry_faqs.xlsx"
    python manage.py import_product_faqs --dry-run
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.db.models import Count
from django.db.models import Q


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_product_technical_inquiry_faqs.xlsx")
FAQ_SHEET = "ProductTechnicalFaqs"
FAQ_HEADERS = (
    "Category",
    "Subcategory",
    "Product Name",
    "Product URL",
    "Source Product URL",
    "Model Code",
    "Series Label",
    "Primary Query",
    "Resource Topic ID",
    "Resource Title",
    "FAQ Family",
    "FAQ Slot",
    "Question",
    "Answer",
    "Sort Order",
    "Question Word Count",
    "Answer Word Count",
    "Key Signals",
    "QA Status",
    "Notes",
)
EXPECTED_SLOTS = {1, 2, 3}
EXPECTED_SORT_ORDERS = {10, 20, 30}


@dataclass(frozen=True)
class ProductSheetKey:
    category_name: str
    product_name: str


@dataclass(frozen=True)
class ProductFaqRow:
    key: ProductSheetKey
    source_product_url: str
    model_code: str
    faq_slot: int
    question: str
    answer: str
    sort_order: int


@dataclass(frozen=True)
class PreparedProductFaq:
    product: "Any"
    question: str
    answer: str
    sort_order: int


class Command(BaseCommand):
    help = "Import product-level Technical Inquiry FAQs from protaylor_product_technical_inquiry_faqs.xlsx."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to the Technical Inquiry FAQ workbook. Defaults to protaylor_product_technical_inquiry_faqs.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import Product
        from apps.content.models import FAQItem

        excel_path = self._resolve_excel_path(options.get("excel"))
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        raw_rows = self._load_sheet_records(workbook, FAQ_SHEET, FAQ_HEADERS)
        faq_rows_by_product = self._build_faq_rows(raw_rows)

        workbook_products = len(faq_rows_by_product)
        self.stdout.write(f"Workbook products: {workbook_products}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))

        product_names = {key.product_name for key in faq_rows_by_product}
        source_product_urls = {
            row.source_product_url
            for rows in faq_rows_by_product.values()
            for row in rows
            if row.source_product_url
        }
        product_query = Product.objects.all()
        if source_product_urls:
            product_query = product_query.filter(
                Q(source_url__in=source_product_urls) | Q(name__in=product_names)
            )
        else:
            product_query = product_query.filter(name__in=product_names)

        products_by_name: dict[str, list[Product]] = defaultdict(list)
        products_by_source_url: dict[str, list[Product]] = defaultdict(list)
        for product in product_query.select_related("category"):
            products_by_name[product.name].append(product)
            if product.source_url:
                products_by_source_url[product.source_url].append(product)

        product_ct = ContentType.objects.get_for_model(Product, for_concrete_model=False)
        matched_products = replaced_products = skipped_products = created_rows = 0
        unmatched_products: list[str] = []
        ambiguous_products: list[str] = []
        category_warnings: list[str] = []
        model_code_warnings: list[str] = []
        prepared_rows: list[PreparedProductFaq] = []
        imported_product_ids: set[int] = set()
        matched_product_keys: dict[int, ProductSheetKey] = {}

        for product_key in sorted(faq_rows_by_product.keys(), key=lambda item: (item.category_name.lower(), item.product_name.lower())):
            faq_rows = faq_rows_by_product[product_key]
            source_product_url = next((row.source_product_url for row in faq_rows if row.source_product_url), "")
            product_matches = products_by_source_url.get(source_product_url, []) if source_product_url else []
            if len(product_matches) > 1:
                skipped_products += 1
                ambiguous_products.append(
                    f"{product_key.category_name} / {product_key.product_name}: "
                    f"Source Product URL {source_product_url!r} matched multiple products"
                )
                continue
            if not product_matches:
                product_matches = products_by_name.get(product_key.product_name, [])
                if len(product_matches) > 1:
                    narrowed_matches = [
                        product
                        for product in product_matches
                        if product.category.name == product_key.category_name
                    ]
                    if narrowed_matches:
                        product_matches = narrowed_matches

            if len(product_matches) == 0:
                skipped_products += 1
                unmatched_products.append(f"{product_key.category_name} / {product_key.product_name}: product not found")
                continue
            if len(product_matches) > 1:
                skipped_products += 1
                ambiguous_products.append(f"{product_key.category_name} / {product_key.product_name}: duplicate product match")
                continue

            product = product_matches[0]
            previous_key = matched_product_keys.get(product.id)
            if previous_key and previous_key != product_key:
                raise CommandError(
                    "Workbook maps multiple product keys to the same database product: "
                    f"{previous_key.category_name} / {previous_key.product_name} and "
                    f"{product_key.category_name} / {product_key.product_name}"
                )
            matched_product_keys[product.id] = product_key
            matched_products += 1

            if product.category.name != product_key.category_name:
                category_warnings.append(
                    f"{product_key.category_name} / {product_key.product_name}: "
                    f"Excel category={product_key.category_name!r}, DB category={product.category.name!r}"
                )

            workbook_model_code = next((row.model_code for row in faq_rows if row.model_code), "")
            if workbook_model_code and workbook_model_code != (product.model_code or ""):
                model_code_warnings.append(
                    f"{product.category.name} / {product.name}: Excel Model Code={workbook_model_code!r}, DB Model Code={(product.model_code or '')!r}"
                )

            if dry_run:
                replaced_products += 1
                created_rows += len(faq_rows)
                continue

            replaced_products += 1
            imported_product_ids.add(product.id)
            for faq_row in sorted(faq_rows, key=lambda item: item.sort_order):
                prepared_row = PreparedProductFaq(
                    product=product,
                    question=faq_row.question,
                    answer=faq_row.answer,
                    sort_order=faq_row.sort_order,
                )
                prepared_rows.append(prepared_row)
                created_rows += 1

        if not dry_run:
            with transaction.atomic():
                target_product_ids = sorted(imported_product_ids)
                self._delete_all_product_faq_rows(product_ct=product_ct)
                FAQItem.objects.bulk_create(
                    [
                        FAQItem(
                            content_type=product_ct,
                            object_id=item.product.id,
                            question=item.question,
                            answer=item.answer,
                            is_featured=True,
                            sort_order=item.sort_order,
                        )
                        for item in prepared_rows
                    ],
                    batch_size=500,
                )
                self._repair_products_with_invalid_faq_counts(
                    product_ct=product_ct,
                    target_product_ids=target_product_ids,
                    prepared_rows=prepared_rows,
                    faq_model=FAQItem,
                )

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Workbook products: {workbook_products}")
        self.stdout.write(f"  Matched products: {matched_products}")
        self.stdout.write(f"  Replaced products: {replaced_products}")
        self.stdout.write(f"  Created FAQ rows: {created_rows}")
        self.stdout.write(f"  Skipped products: {skipped_products}")

        if unmatched_products:
            self.stdout.write(self.style.WARNING("\nUnmatched products"))
            for item in unmatched_products[:100]:
                self.stdout.write(f"  - {item}")

        if ambiguous_products:
            self.stdout.write(self.style.WARNING("\nAmbiguous products"))
            for item in ambiguous_products[:100]:
                self.stdout.write(f"  - {item}")

        if category_warnings:
            self.stdout.write(self.style.WARNING("\nCategory warnings"))
            for item in category_warnings[:100]:
                self.stdout.write(f"  - {item}")

        if model_code_warnings:
            self.stdout.write(self.style.WARNING("\nModel code warnings"))
            for item in model_code_warnings[:100]:
                self.stdout.write(f"  - {item}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _delete_all_product_faq_rows(self, *, product_ct: ContentType) -> None:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM faq_item WHERE content_type_id = %s", [product_ct.id])

    def _load_sheet_records(self, workbook: "Any", sheet_name: str, required_headers: tuple[str, ...]) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [self._clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        index = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                header: row[index[header]] if index[header] < len(row) else None
                for header in required_headers
            }
            if all(self._clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_number
            records.append(record)
        workbook.close()
        return records

    def _build_faq_rows(self, raw_rows: list[dict[str, Any]]) -> dict[ProductSheetKey, list[ProductFaqRow]]:
        faq_rows_by_product: dict[ProductSheetKey, list[ProductFaqRow]] = defaultdict(list)
        for raw_row in raw_rows:
            row_number = int(raw_row["__row_number__"])
            category = self._clean(raw_row["Category"])
            subcategory = self._clean(raw_row["Subcategory"])
            product_name = self._clean(raw_row["Product Name"])
            question = self._clean(raw_row["Question"])
            answer = self._clean(raw_row["Answer"])
            if not category:
                raise CommandError(f"ProductTechnicalFaqs row {row_number}: Category is required.")
            if not product_name:
                raise CommandError(f"ProductTechnicalFaqs row {row_number}: Product Name is required.")
            if not question:
                raise CommandError(f"ProductTechnicalFaqs row {row_number}: Question is required.")
            if not answer:
                raise CommandError(f"ProductTechnicalFaqs row {row_number}: Answer is required.")

            key = ProductSheetKey(category_name=subcategory or category, product_name=product_name)
            faq_rows_by_product[key].append(
                ProductFaqRow(
                    key=key,
                    source_product_url=self._clean(raw_row["Source Product URL"]),
                    model_code=self._clean(raw_row["Model Code"]),
                    faq_slot=self._parse_int(raw_row["FAQ Slot"], "FAQ Slot", row_number),
                    question=question,
                    answer=answer,
                    sort_order=self._parse_int(raw_row["Sort Order"], "Sort Order", row_number),
                )
            )

        for key, rows in faq_rows_by_product.items():
            if len(rows) != 3:
                raise CommandError(f"{key.category_name} / {key.product_name}: expected exactly 3 FAQ rows, found {len(rows)}.")
            slots = {row.faq_slot for row in rows}
            if slots != EXPECTED_SLOTS:
                raise CommandError(f"{key.category_name} / {key.product_name}: FAQ Slot must be 1, 2, and 3.")
            sort_orders = {row.sort_order for row in rows}
            if sort_orders != EXPECTED_SORT_ORDERS:
                raise CommandError(f"{key.category_name} / {key.product_name}: Sort Order must be 10, 20, and 30.")
        return faq_rows_by_product

    def _repair_products_with_invalid_faq_counts(
        self,
        *,
        product_ct: ContentType,
        target_product_ids: list[int],
        prepared_rows: list[PreparedProductFaq],
        faq_model,
    ) -> None:
        invalid_product_ids = list(
            faq_model.objects.filter(content_type=product_ct, object_id__in=target_product_ids)
            .values("object_id")
            .annotate(total=Count("id"))
            .exclude(total=3)
            .values_list("object_id", flat=True)
        )
        if not invalid_product_ids:
            return

        self.stdout.write(
            self.style.WARNING(
                f"Repairing FAQ count anomalies for {len(invalid_product_ids)} product(s) after bulk import."
            )
        )
        self._delete_all_product_faq_rows(product_ct=product_ct)
        faq_model.objects.bulk_create(
            [
                faq_model(
                    content_type=product_ct,
                    object_id=item.product.id,
                    question=item.question,
                    answer=item.answer,
                    is_featured=True,
                    sort_order=item.sort_order,
                )
                for item in prepared_rows
            ],
            batch_size=500,
        )

        remaining_invalid = list(
            faq_model.objects.filter(content_type=product_ct, object_id__in=target_product_ids)
            .values("object_id")
            .annotate(total=Count("id"))
            .exclude(total=3)
            .values_list("object_id", flat=True)
        )
        if remaining_invalid:
            raise CommandError(
                "FAQ import finished with invalid counts for product IDs: "
                + ", ".join(str(product_id) for product_id in remaining_invalid)
            )

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_int(self, value: Any, field_name: str, row_number: int) -> int:
        cleaned = self._clean(value)
        if not cleaned:
            raise CommandError(f"ProductTechnicalFaqs row {row_number}: {field_name} is required.")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise CommandError(f"ProductTechnicalFaqs row {row_number}: invalid {field_name} value {value!r}.") from exc
