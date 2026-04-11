"""
Django 管理命令：从 reviewed 规格工作簿导入产品规格。

使用方式（在 backend/ 目录下）：
    python manage.py import_product_specs
    python manage.py import_product_specs --excel "F:/菱威仓管/protaylor_product_specs_reviewed_v3.xlsx"
    python manage.py import_product_specs --dry-run

行为：
    - 只导入 ProductSpecGroup / ProductSpecRow
    - 输入工作簿必须包含 SpecGroups / SpecRows 两个 sheet
    - 允许存在 ReviewSummary / RuleNotes 等附加 sheet，命令会忽略它们
    - 产品匹配键使用：分类名 + 产品名
    - 采用“整产品覆盖”策略：命中产品后先删除旧规格，再按 Excel 重建
    - 若某个产品在表里没有任何有效规格行，则跳过该产品，不清空数据库原有规格
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_product_specs_reviewed_v3.xlsx")
SHEET_GROUPS = "SpecGroups"
SHEET_ROWS = "SpecRows"

GROUP_HEADERS = (
    "Category",
    "Subcategory",
    "Product Name",
    "Model Code",
    "Group Title",
    "Group Kind Code",
    "Group Sort Order",
)
ROW_HEADERS = (
    "Category",
    "Subcategory",
    "Product Name",
    "Model Code",
    "Group Title",
    "Group Kind Code",
    "Row Label",
    "Row Value",
    "Row Unit",
    "Is Highlight",
    "Row Sort Order",
)


@dataclass(frozen=True)
class ProductSheetKey:
    category_name: str
    product_name: str


@dataclass(frozen=True)
class GroupSheetKey:
    category_name: str
    product_name: str
    group_title: str
    group_kind_code: str


@dataclass(frozen=True)
class GroupImportRecord:
    key: GroupSheetKey
    model_code: str
    sort_order: int


@dataclass(frozen=True)
class RowImportRecord:
    key: GroupSheetKey
    label: str
    value: str
    unit: str
    is_highlight: bool
    sort_order: int
    source_header: str


class Command(BaseCommand):
    help = "Import ProductSpecGroup/ProductSpecRow from a reviewed workbook."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to a reviewed product specs workbook. Defaults to protaylor_product_specs_reviewed_v3.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import Product, ProductCategory, ProductSpecGroup, ProductSpecGroupKind, ProductSpecRow

        excel_path = self._resolve_excel_path(options.get("excel"))
        dry_run = bool(options["dry_run"])

        # 这次命令不重新推断原始规格，只消费已经整理好的双 sheet 导入表。
        # 这样可以把“字段筛选/归组”的人工决策稳定固化在 Excel 里，
        # 命令只负责忠实落库，不再做第二轮猜测。
        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)

        group_records = self._load_sheet_records(workbook, SHEET_GROUPS, GROUP_HEADERS)
        row_records = self._load_sheet_records(workbook, SHEET_ROWS, ROW_HEADERS)

        group_records_by_product: dict[ProductSheetKey, list[GroupImportRecord]] = defaultdict(list)
        group_record_lookup: dict[GroupSheetKey, GroupImportRecord] = {}

        for raw_record in group_records:
            group_key = self._build_group_key(raw_record)
            if group_key in group_record_lookup:
                raise CommandError(
                    "Duplicate group definition found for "
                    f"{group_key.category_name} / {group_key.product_name} / {group_key.group_title}."
                )

            group_record = GroupImportRecord(
                key=group_key,
                model_code=_clean_text(raw_record["Model Code"]),
                sort_order=self._parse_int(
                    raw_record["Group Sort Order"],
                    field_name="Group Sort Order",
                    sheet_name=SHEET_GROUPS,
                    product_name=group_key.product_name,
                ),
            )
            group_record_lookup[group_key] = group_record
            group_records_by_product[
                ProductSheetKey(group_key.category_name, group_key.product_name)
            ].append(group_record)

        row_records_by_group: dict[GroupSheetKey, list[RowImportRecord]] = defaultdict(list)
        skipped_empty_rows = 0
        for raw_record in row_records:
            group_key = self._build_group_key(raw_record)
            if group_key not in group_record_lookup:
                raise CommandError(
                    "SpecRows contains a row without matching SpecGroups definition for "
                    f"{group_key.category_name} / {group_key.product_name} / {group_key.group_title}."
                )

            label = _clean_text(raw_record["Row Label"])
            value = _clean_text(raw_record["Row Value"])
            if not label or not value:
                skipped_empty_rows += 1
                continue

            row_records_by_group[group_key].append(
                RowImportRecord(
                    key=group_key,
                    label=label,
                    value=value,
                    unit=_clean_text(raw_record["Row Unit"]),
                    is_highlight=self._parse_bool(
                        raw_record["Is Highlight"],
                        field_name="Is Highlight",
                        sheet_name=SHEET_ROWS,
                        product_name=group_key.product_name,
                    ),
                    sort_order=self._parse_int(
                        raw_record["Row Sort Order"],
                        field_name="Row Sort Order",
                        sheet_name=SHEET_ROWS,
                        product_name=group_key.product_name,
                    ),
                    source_header=_clean_text(raw_record.get("Source Attribute Header", "")),
                )
            )

        total_products = len(group_records_by_product)
        matched_products = 0
        replaced_products = 0
        skipped_products = 0
        products_without_valid_rows = 0
        skipped_empty_products = 0
        unmatched_products: list[str] = []
        ambiguous_categories: list[str] = []
        ambiguous_products: list[str] = []
        empty_products: list[str] = []
        model_code_warnings: list[str] = []

        self.stdout.write(f"Products defined in workbook: {total_products}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no database writes."))

        group_kind_map = {
            "quick_facts": ProductSpecGroupKind.QUICK_FACTS,
            "technical": ProductSpecGroupKind.TECHNICAL,
        }

        # 这里先把分类和产品一次性批量拉进内存，避免 600+ 产品导入时
        # 变成“每个产品都去远程 PostgreSQL 查 2-4 次”的 N+1 风格慢查询。
        category_names = {product_key.category_name for product_key in group_records_by_product}
        categories_by_name: dict[str, list[ProductCategory]] = defaultdict(list)
        for category in ProductCategory.objects.filter(name__in=category_names):
            categories_by_name[category.name].append(category)

        unique_category_ids = [
            categories[0].id for categories in categories_by_name.values() if len(categories) == 1
        ]
        product_names = {product_key.product_name for product_key in group_records_by_product}
        products_by_key: dict[tuple[int, str], list[Product]] = defaultdict(list)
        for product in Product.objects.filter(
            category_id__in=unique_category_ids,
            name__in=product_names,
        ).select_related("category"):
            products_by_key[(product.category_id, product.name)].append(product)

        for product_key in sorted(
            group_records_by_product.keys(),
            key=lambda item: (item.category_name.lower(), item.product_name.lower()),
        ):
            category_matches = categories_by_name.get(product_key.category_name, [])
            if len(category_matches) == 0:
                skipped_products += 1
                unmatched_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：未找到分类"
                )
                continue
            if len(category_matches) > 1:
                skipped_products += 1
                ambiguous_categories.append(
                    f"{product_key.category_name} / {product_key.product_name}：分类名重复，无法唯一匹配"
                )
                continue
            category = category_matches[0]

            product_matches = products_by_key.get((category.id, product_key.product_name), [])
            if len(product_matches) == 0:
                skipped_products += 1
                unmatched_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：未找到产品"
                )
                continue
            if len(product_matches) > 1:
                skipped_products += 1
                ambiguous_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：产品名重复，无法唯一匹配"
                )
                continue

            product = product_matches[0]
            matched_products += 1

            product_group_records = sorted(
                group_records_by_product[product_key],
                key=lambda record: (record.sort_order, record.key.group_title.lower()),
            )
            valid_group_payloads: list[tuple[GroupImportRecord, list[RowImportRecord]]] = []
            for group_record in product_group_records:
                rows = sorted(
                    row_records_by_group.get(group_record.key, []),
                    key=lambda row: (row.sort_order, row.label.lower()),
                )
                if rows:
                    valid_group_payloads.append((group_record, rows))

            if not valid_group_payloads:
                products_without_valid_rows += 1
                skipped_empty_products += 1
                empty_products.append(
                    f"{product.category.name} / {product.name}：工作簿里没有有效规格行，保留数据库现有规格"
                )
                continue

            # 这里故意不用 Product URL / Product Slug 做主匹配：
            # 1. 现有 import_products.py 没把 source_url 回填进数据库；
            # 2. 导入表里的 Product Slug 与历史 slugify 规则并不完全一致。
            # 当前最稳定的定位方式，是“分类名 + 产品名”在现有库中的唯一组合。
            if self._has_model_code_mismatch(product_group_records, product.model_code):
                workbook_model_code = next(
                    (record.model_code for record in product_group_records if record.model_code),
                    "",
                )
                model_code_warnings.append(
                    f"{product.category.name} / {product.name}：Excel Model Code={workbook_model_code!r}，"
                    f"数据库 Model Code={product.model_code!r}"
                )

            if dry_run:
                continue

            # 这里采用“整产品覆盖”而不是增量 upsert。
            # 原因是这份 Excel 已经是人工整理后的最终导入源，
            # 重跑导入时我们更关心“数据库是否与当前表格完全一致”，
            # 而不是保留历史残留规格，导致前台出现脏数据。
            with transaction.atomic():
                product.spec_groups.all().delete()

                for group_record, rows in valid_group_payloads:
                    group = ProductSpecGroup.objects.create(
                        product=product,
                        title=group_record.key.group_title,
                        group_kind=group_kind_map[group_record.key.group_kind_code],
                        sort_order=group_record.sort_order,
                    )
                    for row_record in rows:
                        ProductSpecRow.objects.create(
                            group=group,
                            label=row_record.label,
                            value=row_record.value,
                            unit=row_record.unit,
                            is_highlight=row_record.is_highlight,
                            sort_order=row_record.sort_order,
                        )

            replaced_products += 1

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Workbook products: {total_products}")
        self.stdout.write(f"  Matched products: {matched_products}")
        self.stdout.write(f"  Replaced products: {replaced_products}")
        self.stdout.write(f"  Skipped products: {skipped_products}")
        self.stdout.write(f"  Products without valid rows: {products_without_valid_rows}")
        self.stdout.write(f"  Skipped empty products: {skipped_empty_products}")
        self.stdout.write(f"  Skipped empty spec rows: {skipped_empty_rows}")

        if unmatched_products:
            self.stdout.write(self.style.WARNING("\nUnmatched products"))
            for item in unmatched_products:
                self.stdout.write(f"  - {item}")

        if ambiguous_categories:
            self.stdout.write(self.style.WARNING("\nAmbiguous categories"))
            for item in ambiguous_categories:
                self.stdout.write(f"  - {item}")

        if ambiguous_products:
            self.stdout.write(self.style.WARNING("\nAmbiguous products"))
            for item in ambiguous_products:
                self.stdout.write(f"  - {item}")

        if empty_products:
            self.stdout.write(self.style.WARNING("\nProducts kept unchanged because workbook has no valid spec rows"))
            for item in empty_products:
                self.stdout.write(f"  - {item}")

        if model_code_warnings:
            self.stdout.write(self.style.WARNING("\nModel code mismatch warnings"))
            for item in model_code_warnings:
                self.stdout.write(f"  - {item}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        if excel_arg:
            excel_path = Path(excel_arg)
            if not excel_path.exists():
                raise CommandError(f"File not found: {excel_path}")
            return excel_path

        # 正式导库命令只允许一个明确的默认数据源。
        # 不再根据磁盘上“碰巧还留着哪些历史文件”做自动回退，
        # 否则人无法确认这次到底导入了哪一份表。
        if DEFAULT_EXCEL_PATH.exists():
            return DEFAULT_EXCEL_PATH

        raise CommandError(
            "Default workbook not found: "
            f"{DEFAULT_EXCEL_PATH}. Please pass --excel explicitly if you want a different file."
        )

    def _load_sheet_records(
        self,
        workbook,
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"Missing required sheet: {sheet_name}")

        worksheet = workbook[sheet_name]
        row_iter = worksheet.iter_rows(min_row=1, values_only=True)
        try:
            headers = next(row_iter)
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name} is empty.") from exc

        header_map = {_clean_text(value): index for index, value in enumerate(headers)}
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name} is missing required columns: {', '.join(missing_headers)}"
            )

        records: list[dict[str, Any]] = []
        for row in row_iter:
            if row is None:
                continue
            if all(_clean_text(value) == "" for value in row):
                continue
            record = {
                header: row[header_map[header]] if header_map[header] < len(row) else None
                for header in header_map
            }
            records.append(record)
        return records

    def _build_group_key(self, record: dict[str, Any]) -> GroupSheetKey:
        group_kind_code = _clean_text(record["Group Kind Code"]).lower()
        if group_kind_code not in {"quick_facts", "technical"}:
            raise CommandError(
                "Only quick_facts and technical are supported in this import. "
                f"Got {group_kind_code!r} for product {_clean_text(record['Product Name'])!r}."
            )

        return GroupSheetKey(
            category_name=_target_category_name(
                _clean_text(record["Category"]),
                _clean_text(record["Subcategory"]),
            ),
            product_name=_clean_text(record["Product Name"]),
            group_title=_clean_text(record["Group Title"]),
            group_kind_code=group_kind_code,
        )

    def _parse_int(
        self,
        value: Any,
        *,
        field_name: str,
        sheet_name: str,
        product_name: str,
    ) -> int:
        if value is None or _clean_text(value) == "":
            raise CommandError(f"{sheet_name} -> {product_name} -> {field_name} cannot be empty.")
        if isinstance(value, bool):
            raise CommandError(f"{sheet_name} -> {product_name} -> {field_name} must be an integer.")
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            raise CommandError(f"{sheet_name} -> {product_name} -> {field_name} must be an integer.")

        raw_text = _clean_text(value)
        try:
            return int(raw_text)
        except ValueError as exc:
            raise CommandError(
                f"{sheet_name} -> {product_name} -> {field_name} must be an integer, got {raw_text!r}."
            ) from exc

    def _parse_bool(
        self,
        value: Any,
        *,
        field_name: str,
        sheet_name: str,
        product_name: str,
    ) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, int) and value in {0, 1}:
            return bool(value)
        if isinstance(value, float) and value in {0.0, 1.0}:
            return bool(int(value))

        raw_text = _clean_text(value).lower()
        if raw_text in {"true", "1", "yes", "y"}:
            return True
        if raw_text in {"false", "0", "no", "n", ""}:
            return False
        raise CommandError(
            f"{sheet_name} -> {product_name} -> {field_name} must be a boolean, got {raw_text!r}."
        )

    def _has_model_code_mismatch(
        self,
        group_records: list[GroupImportRecord],
        db_model_code: str,
    ) -> bool:
        workbook_model_code = next((record.model_code for record in group_records if record.model_code), "")
        return bool(workbook_model_code and db_model_code and workbook_model_code != db_model_code)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _target_category_name(category: str, subcategory: str) -> str:
    if subcategory and subcategory != category:
        return subcategory
    return category
