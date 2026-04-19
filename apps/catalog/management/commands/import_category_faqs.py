"""
Import category-level Sourcing FAQ seed rows from Excel.

Usage:
    python manage.py import_category_faqs
    python manage.py import_category_faqs --excel "F:/菱威仓管/protaylor_category.xlsx"
    python manage.py import_category_faqs --sheet "Category Sourcing FAQ Seed"
    python manage.py import_category_faqs --dry-run
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_category.xlsx")
DEFAULT_SHEET_NAME = "Category Sourcing FAQ Seed"
SEED_HEADERS = (
    "source_category",
    "source_subcategory",
    "target_scope",
    "target_category_slug",
    "target_category_name",
    "placement_code",
    "sourcing_faq_title",
    "sort_order",
    "question",
    "answer",
    "is_active",
    "source_fields_used",
    "primary_source_url",
    "secondary_source_url",
    "evidence_note",
    "confidence",
)
ALL_MARKERS = {"── all ──", "-- all --", ""}
EXPECTED_TARGET_SCOPES = {"parent_all", "subcategory"}
EXPECTED_SORT_ORDERS = {1, 2, 3}


@dataclass(frozen=True)
class SeedGroupKey:
    source_category: str
    source_subcategory: str
    target_scope: str
    target_category_slug: str
    target_category_name: str
    placement_code: str


@dataclass(frozen=True)
class SeedFaqRow:
    sort_order: int
    question: str
    answer: str
    is_active: bool


@dataclass(frozen=True)
class PreparedSeedGroup:
    category: "Any"
    placement: int
    sourcing_faq_title: str
    items: list[SeedFaqRow]


class Command(BaseCommand):
    help = "Import ProductCategoryFaqItem rows from the Category Sourcing FAQ Seed sheet."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to protaylor_category.xlsx. Defaults to F:/菱威仓管/protaylor_category.xlsx.",
        )
        parser.add_argument(
            "--sheet",
            default=DEFAULT_SHEET_NAME,
            help=f"Worksheet name to import. Defaults to {DEFAULT_SHEET_NAME!r}.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import ProductCategory, ProductCategoryFaqItem

        excel_path = self._resolve_excel_path(options.get("excel"))
        sheet_name = self._clean(options.get("sheet")) or DEFAULT_SHEET_NAME
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        raw_rows = self._load_sheet_records(workbook, sheet_name, SEED_HEADERS)
        groups = self._build_seed_groups(raw_rows)

        self.stdout.write(f"Workbook groups: {len(groups)}")
        self.stdout.write(f"Workbook rows: {sum(len(items) for _, _, items in groups)}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))

        categories_by_slug = {
            category.slug: category
            for category in ProductCategory.objects.select_related("parent")
        }

        prepared_groups: list[PreparedSeedGroup] = []
        category_titles: dict[int, str] = {}
        unmatched_groups: list[str] = []

        for group_key, sourcing_faq_title, faq_rows in groups:
            try:
                category = self._resolve_target_category(
                    group_key=group_key,
                    categories_by_slug=categories_by_slug,
                )
            except CommandError as exc:
                message = str(exc)
                if message.startswith("UNMATCHED:"):
                    unmatched_groups.append(message.removeprefix("UNMATCHED:").strip())
                    continue
                raise

            if category.name != group_key.target_category_name:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: "
                    f"target_category_name {group_key.target_category_name!r} does not match "
                    f"database category name {category.name!r} for slug {group_key.target_category_slug!r}."
                )

            current_title = category_titles.get(category.id)
            if current_title is not None and current_title != sourcing_faq_title:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: "
                    f"conflicting sourcing_faq_title values for category {category.slug!r}."
                )
            category_titles[category.id] = sourcing_faq_title

            prepared_groups.append(
                PreparedSeedGroup(
                    category=category,
                    placement=self._placement_value_for_code(group_key.placement_code),
                    sourcing_faq_title=sourcing_faq_title,
                    items=faq_rows,
                )
            )

        if unmatched_groups:
            raise CommandError(
                "Import aborted because some seed groups did not match a ProductCategory:\n- "
                + "\n- ".join(unmatched_groups)
            )

        if not dry_run:
            with transaction.atomic():
                for prepared_group in prepared_groups:
                    prepared_group.category.sourcing_faq_title = prepared_group.sourcing_faq_title
                    prepared_group.category.save(update_fields=("sourcing_faq_title",))
                    ProductCategoryFaqItem.objects.filter(
                        category=prepared_group.category,
                        placement=prepared_group.placement,
                    ).delete()
                    ProductCategoryFaqItem.objects.bulk_create(
                        [
                            ProductCategoryFaqItem(
                                category=prepared_group.category,
                                placement=prepared_group.placement,
                                question=item.question,
                                answer=item.answer,
                                is_active=item.is_active,
                                sort_order=item.sort_order,
                            )
                            for item in prepared_group.items
                        ]
                    )

        imported_rows = sum(len(group.items) for group in prepared_groups)
        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Imported groups: {len(prepared_groups)}")
        self.stdout.write(f"  Imported rows: {imported_rows}")
        self.stdout.write(f"  Dry run: {dry_run}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: "Any",
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [self._clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        index = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                header: row[index[header]] if index[header] < len(row) else None
                for header in required_headers
            }
            if all(self._clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_number
            records.append(record)
        workbook.close()
        return records

    def _build_seed_groups(
        self,
        raw_rows: list[dict[str, Any]],
    ) -> list[tuple[SeedGroupKey, str, list[SeedFaqRow]]]:
        grouped_rows: dict[SeedGroupKey, list[dict[str, Any]]] = defaultdict(list)
        for raw_row in raw_rows:
            row_number = int(raw_row["__row_number__"])
            source_category = self._clean(raw_row["source_category"])
            source_subcategory = self._clean(raw_row["source_subcategory"])
            target_scope = self._clean(raw_row["target_scope"])
            target_category_slug = self._clean(raw_row["target_category_slug"])
            target_category_name = self._clean(raw_row["target_category_name"])
            placement_code = self._clean(raw_row["placement_code"])

            if not source_category:
                raise CommandError(f"Category Sourcing FAQ Seed row {row_number}: source_category is required.")
            if target_scope not in EXPECTED_TARGET_SCOPES:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: target_scope must be parent_all or subcategory."
                )
            if not target_category_slug:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: target_category_slug is required."
                )
            if not target_category_name:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: target_category_name is required."
                )
            if not placement_code:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: placement_code is required."
                )

            normalized_subcategory = source_subcategory.lower()
            if target_scope == "parent_all" and normalized_subcategory not in ALL_MARKERS:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: parent_all rows must use an All-style source_subcategory."
                )
            if target_scope == "subcategory" and normalized_subcategory in ALL_MARKERS:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: subcategory rows must name a concrete source_subcategory."
                )

            sort_order = self._parse_int(raw_row["sort_order"], "sort_order", row_number)
            question = self._clean(raw_row["question"])
            answer = self._clean(raw_row["answer"])
            if not question:
                raise CommandError(f"Category Sourcing FAQ Seed row {row_number}: question is required.")
            if not answer:
                raise CommandError(f"Category Sourcing FAQ Seed row {row_number}: answer is required.")
            if len(question) > 180:
                raise CommandError(
                    f"Category Sourcing FAQ Seed row {row_number}: question exceeds 180 characters."
                )

            group_key = SeedGroupKey(
                source_category=source_category,
                source_subcategory=source_subcategory,
                target_scope=target_scope,
                target_category_slug=target_category_slug,
                target_category_name=target_category_name,
                placement_code=placement_code,
            )
            grouped_rows[group_key].append(
                raw_row
                | {
                    "sort_order": sort_order,
                    "question": question,
                    "answer": answer,
                    "is_active": self._parse_bool(raw_row["is_active"], "is_active", row_number),
                }
            )

        grouped_payload: list[tuple[SeedGroupKey, str, list[SeedFaqRow]]] = []
        for group_key, group_rows in sorted(
            grouped_rows.items(),
            key=lambda item: (
                item[0].source_category.lower(),
                item[0].target_scope,
                item[0].target_category_slug.lower(),
            ),
        ):
            sourcing_titles = {self._clean(row["sourcing_faq_title"]) for row in group_rows}
            if len(sourcing_titles) != 1:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: inconsistent sourcing_faq_title values."
                )
            sourcing_title = sourcing_titles.pop()
            if not sourcing_title:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: sourcing_faq_title is required."
                )

            if len(group_rows) != 3:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: expected exactly 3 seed rows, found {len(group_rows)}."
                )

            group_sort_orders = {row["sort_order"] for row in group_rows}
            if group_sort_orders != EXPECTED_SORT_ORDERS:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: sort_order must use 1, 2, and 3."
                )

            items = [
                SeedFaqRow(
                    sort_order=row["sort_order"],
                    question=row["question"],
                    answer=row["answer"],
                    is_active=row["is_active"],
                )
                for row in sorted(group_rows, key=lambda row: row["sort_order"])
            ]
            grouped_payload.append((group_key, sourcing_title, items))
        return grouped_payload

    def _resolve_target_category(
        self,
        *,
        group_key: SeedGroupKey,
        categories_by_slug: dict[str, "Any"],
    ):
        category = categories_by_slug.get(group_key.target_category_slug)
        if category is None:
            raise CommandError(
                "UNMATCHED: "
                f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_slug} ({group_key.target_scope}): target_category_slug not found"
            )

        if group_key.target_scope == "parent_all":
            if category.parent_id is not None:
                raise CommandError(
                    "UNMATCHED: "
                    f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_slug} ({group_key.target_scope}): slug resolves to a child category"
                )
            return category

        if category.parent_id is None:
            raise CommandError(
                "UNMATCHED: "
                f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_slug} ({group_key.target_scope}): slug resolves to a top-level category"
            )
        if category.parent and category.parent.name != group_key.source_category:
            raise CommandError(
                "UNMATCHED: "
                f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_slug} ({group_key.target_scope}): parent category mismatch"
            )
        return category

    def _placement_value_for_code(self, placement_code: str) -> int:
        from apps.catalog.models import ProductCategoryFaqItem

        placement_lookup = {
            code: value
            for value, code in ProductCategoryFaqItem.Placement.codes().items()
        }
        if placement_code not in placement_lookup:
            raise CommandError(f"Unsupported placement_code {placement_code!r}.")
        return placement_lookup[placement_code]

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_int(self, value: Any, field_name: str, row_number: int) -> int:
        cleaned = self._clean(value)
        if not cleaned:
            raise CommandError(f"Category Sourcing FAQ Seed row {row_number}: {field_name} is required.")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise CommandError(
                f"Category Sourcing FAQ Seed row {row_number}: invalid {field_name} value {value!r}."
            ) from exc

    def _parse_bool(self, value: Any, field_name: str, row_number: int) -> bool:
        cleaned = self._clean(value).lower()
        if cleaned in {"true", "1", "yes", "y"}:
            return True
        if cleaned in {"false", "0", "no", "n"}:
            return False
        raise CommandError(
            f"Category Sourcing FAQ Seed row {row_number}: {field_name} must be a boolean-like value."
        )
