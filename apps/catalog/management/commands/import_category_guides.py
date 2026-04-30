"""
Import category Guide page staging rows from Excel.

Usage:
    python manage.py import_category_guides
    python manage.py import_category_guides --excel "F:/菱威仓管/protaylor_guide_page_staging.xlsx"
    python manage.py import_category_guides --dry-run
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_guide_page_staging.xlsx")

SHEET_PLANNING = "Guide Page Planning"
SHEET_REVIEW = "Guide Review"
SHEET_HERO = "Guide Hero Draft"
SHEET_DEFINITION = "Guide Definition Draft"
SHEET_DEFINITION_CARDS = "Guide Definition Cards"
SHEET_CONTEXTS = "Guide Contexts Draft"
SHEET_MATRIX = "Guide Matrix Draft"
SHEET_PATHS = "Guide Paths Draft"
SHEET_STANDARDS = "Guide Standards Draft"
SHEET_FAQ = "Guide FAQ Draft"
SHEET_RESOURCES = "Guide Resources Draft"
SHEET_FINAL_CTA = "Guide Final CTA Draft"

FINALIZED_ROW_STATUSES = {"finalized_for_staging", "approved_for_staging", "approved"}
DISPLAY_ONLY_MODE_VALUES = {"show", "visible", "enabled"}

PLANNING_HEADERS = (
    "category_name",
    "category_slug",
    "page_url",
    "page_type",
    "guide_scope",
    "research_gate_status",
    "current_stage",
    "next_step_allowed",
    "hero_mode",
    "definition_mode",
    "contexts_mode",
    "matrix_mode",
    "paths_mode",
    "standards_mode",
    "faq_mode",
    "resources_mode",
    "final_cta_mode",
    "owner",
    "reviewer",
    "source_doc",
    "notes",
)
REVIEW_HEADERS = (
    "category_name",
    "category_slug",
    "review_round",
    "review_status",
    "current_verdict",
    "approved_for_excel_staging",
    "approved_for_import",
    "reviewer",
    "review_doc",
    "hero_risk",
    "definition_risk",
    "contexts_risk",
    "matrix_risk",
    "paths_risk",
    "standards_risk",
    "faq_risk",
    "resources_cta_risk",
    "key_resolution_1",
    "key_resolution_2",
    "key_resolution_3",
    "notes",
)
HERO_HEADERS = (
    "category_name",
    "category_slug",
    "eyebrow",
    "title",
    "description",
    "hero_image_url",
    "hero_image_alt",
    "hero_note_quote",
    "hero_note_attribution",
    "primary_cta_label",
    "primary_cta_href",
    "secondary_cta_label",
    "secondary_cta_href",
    "status",
    "source_doc",
    "notes",
)
DEFINITION_HEADERS = (
    "category_name",
    "category_slug",
    "definition_title",
    "paragraph_1",
    "paragraph_2",
    "paragraph_3",
    "status",
    "source_doc",
    "notes",
)
DEFINITION_CARD_HEADERS = (
    "category_name",
    "category_slug",
    "card_key",
    "card_role",
    "sort_order",
    "card_title",
    "card_copy",
    "icon",
    "status",
    "source_doc",
    "evidence_note",
)
CONTEXT_HEADERS = (
    "category_name",
    "category_slug",
    "contexts_title",
    "context_key",
    "sort_order",
    "context_title",
    "context_copy",
    "image_url",
    "image_alt",
    "status",
    "source_doc",
    "evidence_note",
)
MATRIX_HEADERS = (
    "category_name",
    "category_slug",
    "matrix_title",
    "matrix_eyebrow",
    "factor_key",
    "sort_order",
    "factor_title",
    "factor_copy",
    "icon",
    "evidence_strength",
    "synthesis_flag",
    "status",
    "source_doc",
    "evidence_note",
)
PATH_HEADERS = (
    "category_name",
    "category_slug",
    "paths_title",
    "paths_eyebrow",
    "path_mode",
    "path_key",
    "sort_order",
    "step",
    "path_title",
    "path_copy",
    "bullet_1",
    "bullet_2",
    "bullet_3",
    "target_href",
    "target_category_slug",
    "status",
    "source_doc",
    "evidence_note",
)
STANDARDS_HEADERS = (
    "category_name",
    "category_slug",
    "standards_mode",
    "standards_title",
    "standards_copy",
    "stat_1_value",
    "stat_1_label",
    "stat_2_value",
    "stat_2_label",
    "stat_3_value",
    "stat_3_label",
    "stat_4_value",
    "stat_4_label",
    "status",
    "source_doc",
    "evidence_note",
)
FAQ_HEADERS = (
    "category_name",
    "category_slug",
    "faq_title",
    "faq_source_mode",
    "faq_key",
    "sort_order",
    "question",
    "answer",
    "status",
    "source_doc",
    "evidence_note",
)
RESOURCE_HEADERS = (
    "category_name",
    "category_slug",
    "resources_title",
    "resources_mode",
    "resource_key",
    "sort_order",
    "resource_type",
    "label",
    "title",
    "href",
    "linked_slug",
    "status",
    "source_doc",
    "notes",
)
FINAL_CTA_HEADERS = (
    "category_name",
    "category_slug",
    "cta_mode",
    "cta_title",
    "cta_copy",
    "primary_label",
    "primary_href",
    "primary_target_slug",
    "secondary_label",
    "secondary_href",
    "secondary_target_slug",
    "status",
    "source_doc",
    "notes",
)


@dataclass(frozen=True)
class PlanningSeed:
    category_name: str
    category_slug: str
    paths_mode: str
    standards_mode: str
    resources_mode: str
    cta_mode: str


@dataclass(frozen=True)
class GuideItemSeed:
    section: int
    item_key: str
    sort_order: int
    eyebrow: str
    title: str
    body: str = ""
    supporting_points: str = ""
    icon: str = ""
    asset_url: str = ""
    asset_alt: str = ""
    href: str = ""
    cta_label: str = ""


@dataclass(frozen=True)
class GuideFaqSeed:
    sort_order: int
    question: str
    answer: str


@dataclass
class GuideSeed:
    category_name: str
    category_slug: str
    hero_image_url: str = ""
    guide_defaults: dict[str, Any] = field(default_factory=dict)
    item_seeds: list[GuideItemSeed] = field(default_factory=list)
    faq_seeds: list[GuideFaqSeed] = field(default_factory=list)


@dataclass(frozen=True)
class PreparedGuideSeed:
    category: Any
    seed: GuideSeed


class Command(BaseCommand):
    help = "Import ProductCategoryGuide/ProductCategoryGuideItem rows from guide page staging workbook."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to protaylor_guide_page_staging.xlsx. Defaults to F:/菱威仓管/protaylor_guide_page_staging.xlsx.",
        )
        parser.add_argument(
            "--skip-review-gate",
            action="store_true",
            help="Do not require Guide Review.approved_for_import=yes for each planned category.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import ProductCategory

        excel_path = self._resolve_excel_path(options.get("excel"))
        dry_run = bool(options["dry_run"])
        require_review_gate = not bool(options["skip_review_gate"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            records = {
                SHEET_PLANNING: self._load_sheet_records(workbook, SHEET_PLANNING, PLANNING_HEADERS),
                SHEET_REVIEW: self._load_sheet_records(workbook, SHEET_REVIEW, REVIEW_HEADERS),
                SHEET_HERO: self._load_sheet_records(workbook, SHEET_HERO, HERO_HEADERS),
                SHEET_DEFINITION: self._load_sheet_records(workbook, SHEET_DEFINITION, DEFINITION_HEADERS),
                SHEET_DEFINITION_CARDS: self._load_sheet_records(
                    workbook,
                    SHEET_DEFINITION_CARDS,
                    DEFINITION_CARD_HEADERS,
                ),
                SHEET_CONTEXTS: self._load_sheet_records(workbook, SHEET_CONTEXTS, CONTEXT_HEADERS),
                SHEET_MATRIX: self._load_sheet_records(workbook, SHEET_MATRIX, MATRIX_HEADERS),
                SHEET_PATHS: self._load_sheet_records(workbook, SHEET_PATHS, PATH_HEADERS),
                SHEET_STANDARDS: self._load_sheet_records(workbook, SHEET_STANDARDS, STANDARDS_HEADERS),
                SHEET_FAQ: self._load_sheet_records(workbook, SHEET_FAQ, FAQ_HEADERS),
                SHEET_RESOURCES: self._load_sheet_records(workbook, SHEET_RESOURCES, RESOURCE_HEADERS),
                SHEET_FINAL_CTA: self._load_sheet_records(workbook, SHEET_FINAL_CTA, FINAL_CTA_HEADERS),
            }
        finally:
            workbook.close()

        planning_by_slug = self._build_planning(records[SHEET_PLANNING])
        if require_review_gate:
            self._validate_review_gate(records[SHEET_REVIEW], planning_by_slug)

        guide_seeds = self._build_guide_seeds(records, planning_by_slug)

        self.stdout.write(f"Workbook guide groups: {len(guide_seeds)}")
        self.stdout.write(f"Workbook guide items: {sum(len(seed.item_seeds) for seed in guide_seeds.values())}")
        self.stdout.write(f"Workbook guide FAQs: {sum(len(seed.faq_seeds) for seed in guide_seeds.values())}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))

        categories_by_slug = {
            category.slug: category
            for category in ProductCategory.objects.select_related("parent")
        }
        prepared = self._prepare_guides(guide_seeds, categories_by_slug)

        if not dry_run:
            self._write_guides(prepared)

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Imported guide groups: {len(prepared)}")
        self.stdout.write(f"  Imported guide items: {sum(len(item.seed.item_seeds) for item in prepared)}")
        self.stdout.write(f"  Imported guide FAQs: {sum(len(item.seed.faq_seeds) for item in prepared)}")
        self.stdout.write(f"  Review gate required: {require_review_gate}")
        self.stdout.write(f"  Dry run: {dry_run}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: Any,
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [self._clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
            if all(self._clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_number
            records.append(record)
        return records

    def _build_planning(self, records: list[dict[str, Any]]) -> dict[str, PlanningSeed]:
        planning_by_slug: dict[str, PlanningSeed] = {}
        for record in records:
            row_number = int(record["__row_number__"])
            category_name = self._required(record, "category_name", SHEET_PLANNING, row_number)
            category_slug = self._required(record, "category_slug", SHEET_PLANNING, row_number)
            current_stage = self._required(record, "current_stage", SHEET_PLANNING, row_number)
            next_step_allowed = self._required(record, "next_step_allowed", SHEET_PLANNING, row_number)

            if current_stage != "import_ready":
                raise CommandError(
                    f"{SHEET_PLANNING} row {row_number}: current_stage must be import_ready."
                )
            if next_step_allowed != "backend_model_importer_api_work":
                raise CommandError(
                    f"{SHEET_PLANNING} row {row_number}: next_step_allowed must be backend_model_importer_api_work."
                )
            if category_slug in planning_by_slug:
                raise CommandError(f"{SHEET_PLANNING}: duplicate category_slug {category_slug!r}.")

            planning_by_slug[category_slug] = PlanningSeed(
                category_name=category_name,
                category_slug=category_slug,
                paths_mode=self._required(record, "paths_mode", SHEET_PLANNING, row_number),
                standards_mode=self._required(record, "standards_mode", SHEET_PLANNING, row_number),
                resources_mode=self._required(record, "resources_mode", SHEET_PLANNING, row_number),
                cta_mode=self._required(record, "final_cta_mode", SHEET_PLANNING, row_number),
            )

        if not planning_by_slug:
            raise CommandError(f"{SHEET_PLANNING}: no import-ready categories found.")
        return planning_by_slug

    def _validate_review_gate(
        self,
        records: list[dict[str, Any]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        review_by_slug: dict[str, dict[str, Any]] = {}
        for record in records:
            row_number = int(record["__row_number__"])
            category_slug = self._required(record, "category_slug", SHEET_REVIEW, row_number)
            if category_slug in review_by_slug:
                raise CommandError(f"{SHEET_REVIEW}: duplicate category_slug {category_slug!r}.")
            review_by_slug[category_slug] = record

        missing = sorted(set(planning_by_slug) - set(review_by_slug))
        if missing:
            raise CommandError(f"{SHEET_REVIEW}: missing Final QA rows for: {', '.join(missing)}")

        for category_slug in sorted(planning_by_slug):
            record = review_by_slug[category_slug]
            row_number = int(record["__row_number__"])
            if self._clean(record["review_round"]) != "final_qa":
                raise CommandError(f"{SHEET_REVIEW} row {row_number}: review_round must be final_qa.")
            if self._clean(record["review_status"]) != "completed":
                raise CommandError(f"{SHEET_REVIEW} row {row_number}: review_status must be completed.")
            if self._clean(record["current_verdict"]) != "approved_for_import":
                raise CommandError(f"{SHEET_REVIEW} row {row_number}: current_verdict must be approved_for_import.")
            if self._clean(record["approved_for_import"]).lower() != "yes":
                raise CommandError(f"{SHEET_REVIEW} row {row_number}: approved_for_import must be yes.")
            self._required(record, "review_doc", SHEET_REVIEW, row_number)

    def _build_guide_seeds(
        self,
        records: dict[str, list[dict[str, Any]]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> dict[str, GuideSeed]:
        self._validate_sheet_slugs(records, planning_by_slug)

        seeds: dict[str, GuideSeed] = {
            slug: GuideSeed(category_name=plan.category_name, category_slug=slug)
            for slug, plan in planning_by_slug.items()
        }

        self._apply_hero_records(seeds, records[SHEET_HERO])
        self._apply_definition_records(seeds, records[SHEET_DEFINITION])
        self._apply_definition_card_records(seeds, records[SHEET_DEFINITION_CARDS])
        self._apply_context_records(seeds, records[SHEET_CONTEXTS])
        self._apply_matrix_records(seeds, records[SHEET_MATRIX])
        self._apply_path_records(seeds, records[SHEET_PATHS], planning_by_slug)
        self._apply_standard_records(seeds, records[SHEET_STANDARDS], planning_by_slug)
        self._apply_faq_records(seeds, records[SHEET_FAQ])
        self._apply_resource_records(seeds, records[SHEET_RESOURCES], planning_by_slug)
        self._apply_final_cta_records(seeds, records[SHEET_FINAL_CTA], planning_by_slug)

        self._validate_required_modules(seeds)
        return seeds

    def _validate_sheet_slugs(
        self,
        records: dict[str, list[dict[str, Any]]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        planning_slugs = set(planning_by_slug)
        for sheet_name, sheet_records in records.items():
            if sheet_name in {SHEET_PLANNING, SHEET_REVIEW}:
                continue
            for record in sheet_records:
                row_number = int(record["__row_number__"])
                category_slug = self._required(record, "category_slug", sheet_name, row_number)
                if category_slug not in planning_slugs:
                    raise CommandError(
                        f"{sheet_name} row {row_number}: category_slug {category_slug!r} is not listed in "
                        f"{SHEET_PLANNING}."
                    )
                self._assert_finalized(record, sheet_name, row_number)

    def _apply_hero_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._single_record_by_slug(records, SHEET_HERO)
        for slug, seed in seeds.items():
            record = self._require_group_record(grouped, slug, SHEET_HERO)
            row_number = int(record["__row_number__"])
            self._validate_category_identity(seed, record, SHEET_HERO, row_number)
            seed.hero_image_url = self._clean(record["hero_image_url"])
            seed.guide_defaults.update(
                {
                    "hero_eyebrow": self._required(record, "eyebrow", SHEET_HERO, row_number),
                    "hero_title": self._required(record, "title", SHEET_HERO, row_number),
                    "answer_summary": self._required(record, "description", SHEET_HERO, row_number),
                    "hero_primary_cta_label": self._required(record, "primary_cta_label", SHEET_HERO, row_number),
                    "hero_primary_cta_href": self._required(record, "primary_cta_href", SHEET_HERO, row_number),
                    "hero_secondary_cta_label": self._required(
                        record,
                        "secondary_cta_label",
                        SHEET_HERO,
                        row_number,
                    ),
                    "hero_secondary_cta_href": self._required(
                        record,
                        "secondary_cta_href",
                        SHEET_HERO,
                        row_number,
                    ),
                    "hero_image_alt": self._clean(record["hero_image_alt"]),
                    "hero_note_title": "",
                    "hero_note_copy": "",
                    "hero_note_quote": self._clean(record["hero_note_quote"]),
                    "hero_note_attribution": self._clean(record["hero_note_attribution"]),
                    "is_active": True,
                }
            )

    def _apply_definition_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._single_record_by_slug(records, SHEET_DEFINITION)
        for slug, seed in seeds.items():
            record = self._require_group_record(grouped, slug, SHEET_DEFINITION)
            row_number = int(record["__row_number__"])
            self._validate_category_identity(seed, record, SHEET_DEFINITION, row_number)
            paragraphs = [
                self._clean(record["paragraph_1"]),
                self._clean(record["paragraph_2"]),
                self._clean(record["paragraph_3"]),
            ]
            definition_copy = "\n\n".join(paragraph for paragraph in paragraphs if paragraph)
            if not definition_copy:
                raise CommandError(f"{SHEET_DEFINITION} row {row_number}: at least one paragraph is required.")
            seed.guide_defaults.update(
                {
                    "definition_title": self._required(
                        record,
                        "definition_title",
                        SHEET_DEFINITION,
                        row_number,
                    ),
                    "definition_copy": definition_copy,
                }
            )

    def _apply_definition_card_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_DEFINITION_CARDS)
            self._validate_unique_rows(rows, SHEET_DEFINITION_CARDS, "card_key", "sort_order")
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_DEFINITION_CARDS, row_number)
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("definition_card"),
                        item_key=self._required(record, "card_key", SHEET_DEFINITION_CARDS, row_number),
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        eyebrow=self._clean(record["card_role"]),
                        title=self._required(record, "card_title", SHEET_DEFINITION_CARDS, row_number),
                        body=self._required(record, "card_copy", SHEET_DEFINITION_CARDS, row_number),
                        icon=self._clean(record["icon"]),
                    )
                )

    def _apply_context_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_CONTEXTS)
            self._validate_unique_rows(rows, SHEET_CONTEXTS, "context_key", "sort_order")
            titles = {self._required(record, "contexts_title", SHEET_CONTEXTS, int(record["__row_number__"])) for record in rows}
            if len(titles) != 1:
                raise CommandError(f"{SHEET_CONTEXTS}: inconsistent contexts_title values for {slug!r}.")
            seed.guide_defaults["contexts_title"] = titles.pop()
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_CONTEXTS, row_number)
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("operational_context"),
                        item_key=self._required(record, "context_key", SHEET_CONTEXTS, row_number),
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        eyebrow="",
                        title=self._required(record, "context_title", SHEET_CONTEXTS, row_number),
                        body=self._required(record, "context_copy", SHEET_CONTEXTS, row_number),
                        asset_url=self._clean(record["image_url"]),
                        asset_alt=self._clean(record["image_alt"]),
                    )
                )

    def _apply_matrix_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_MATRIX)
            self._validate_unique_rows(rows, SHEET_MATRIX, "factor_key", "sort_order")
            matrix_titles = {
                self._required(record, "matrix_title", SHEET_MATRIX, int(record["__row_number__"]))
                for record in rows
            }
            matrix_eyebrows = {
                self._required(record, "matrix_eyebrow", SHEET_MATRIX, int(record["__row_number__"]))
                for record in rows
            }
            if len(matrix_titles) != 1:
                raise CommandError(f"{SHEET_MATRIX}: inconsistent matrix_title values for {slug!r}.")
            if len(matrix_eyebrows) != 1:
                raise CommandError(f"{SHEET_MATRIX}: inconsistent matrix_eyebrow values for {slug!r}.")
            seed.guide_defaults["matrix_title"] = matrix_titles.pop()
            seed.guide_defaults["matrix_eyebrow"] = matrix_eyebrows.pop()
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_MATRIX, row_number)
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("decision_factor"),
                        item_key=self._required(record, "factor_key", SHEET_MATRIX, row_number),
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        eyebrow="",
                        title=self._required(record, "factor_title", SHEET_MATRIX, row_number),
                        body=self._required(record, "factor_copy", SHEET_MATRIX, row_number),
                        icon=self._clean(record["icon"]),
                    )
                )

    def _apply_path_records(
        self,
        seeds: dict[str, GuideSeed],
        records: list[dict[str, Any]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_PATHS)
            self._validate_unique_rows(rows, SHEET_PATHS, "path_key", "sort_order")
            titles = {self._required(record, "paths_title", SHEET_PATHS, int(record["__row_number__"])) for record in rows}
            eyebrows = {self._required(record, "paths_eyebrow", SHEET_PATHS, int(record["__row_number__"])) for record in rows}
            modes = {self._required(record, "path_mode", SHEET_PATHS, int(record["__row_number__"])) for record in rows}
            if len(titles) != 1:
                raise CommandError(f"{SHEET_PATHS}: inconsistent paths_title values for {slug!r}.")
            if len(eyebrows) != 1:
                raise CommandError(f"{SHEET_PATHS}: inconsistent paths_eyebrow values for {slug!r}.")
            if len(modes) != 1:
                raise CommandError(f"{SHEET_PATHS}: inconsistent path_mode values for {slug!r}.")
            path_mode = modes.pop()
            if path_mode != planning_by_slug[slug].paths_mode:
                raise CommandError(
                    f"{SHEET_PATHS}: path_mode {path_mode!r} does not match planning paths_mode "
                    f"{planning_by_slug[slug].paths_mode!r} for {slug!r}."
                )
            seed.guide_defaults["paths_title"] = titles.pop()
            seed.guide_defaults["paths_eyebrow"] = eyebrows.pop()
            seed.guide_defaults["paths_mode"] = self._path_mode_value(path_mode)
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_PATHS, row_number)
                supporting_points = "\n".join(
                    point
                    for point in (
                        self._clean(record["bullet_1"]),
                        self._clean(record["bullet_2"]),
                        self._clean(record["bullet_3"]),
                    )
                    if point
                )
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("path"),
                        item_key=self._required(record, "path_key", SHEET_PATHS, row_number),
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        eyebrow=self._required(record, "step", SHEET_PATHS, row_number),
                        title=self._required(record, "path_title", SHEET_PATHS, row_number),
                        body=self._required(record, "path_copy", SHEET_PATHS, row_number),
                        supporting_points=supporting_points,
                        href=self._required(record, "target_href", SHEET_PATHS, row_number),
                    )
                )

    def _apply_standard_records(
        self,
        seeds: dict[str, GuideSeed],
        records: list[dict[str, Any]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        grouped = self._single_record_by_slug(records, SHEET_STANDARDS)
        for slug, seed in seeds.items():
            record = self._require_group_record(grouped, slug, SHEET_STANDARDS)
            row_number = int(record["__row_number__"])
            self._validate_category_identity(seed, record, SHEET_STANDARDS, row_number)
            standards_mode = self._required(record, "standards_mode", SHEET_STANDARDS, row_number)
            if standards_mode != planning_by_slug[slug].standards_mode:
                raise CommandError(
                    f"{SHEET_STANDARDS} row {row_number}: standards_mode {standards_mode!r} does not match "
                    f"planning standards_mode {planning_by_slug[slug].standards_mode!r}."
                )
            seed.guide_defaults.update(
                {
                    "trust_title": self._required(record, "standards_title", SHEET_STANDARDS, row_number),
                    "trust_copy": self._required(record, "standards_copy", SHEET_STANDARDS, row_number),
                    "trust_mode": self._trust_mode_value(standards_mode),
                }
            )
            stat_count = 0
            for index in range(1, 5):
                value = self._clean(record[f"stat_{index}_value"])
                label = self._clean(record[f"stat_{index}_label"])
                if not value and not label:
                    continue
                if not value or not label:
                    raise CommandError(
                        f"{SHEET_STANDARDS} row {row_number}: stat_{index}_value and stat_{index}_label "
                        "must both be filled or both be blank."
                    )
                stat_count += 1
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("trust_metric"),
                        item_key=f"stat_{index}",
                        sort_order=index,
                        eyebrow=value,
                        title=label,
                    )
                )
            if stat_count == 0:
                raise CommandError(f"{SHEET_STANDARDS} row {row_number}: at least one stat is required.")

    def _apply_faq_records(self, seeds: dict[str, GuideSeed], records: list[dict[str, Any]]) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_FAQ)
            self._validate_unique_rows(rows, SHEET_FAQ, "faq_key", "sort_order")
            titles = {self._required(record, "faq_title", SHEET_FAQ, int(record["__row_number__"])) for record in rows}
            if len(titles) != 1:
                raise CommandError(f"{SHEET_FAQ}: inconsistent faq_title values for {slug!r}.")
            seed.guide_defaults["faq_title"] = titles.pop()
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_FAQ, row_number)
                seed.faq_seeds.append(
                    GuideFaqSeed(
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        question=self._required(record, "question", SHEET_FAQ, row_number),
                        answer=self._required(record, "answer", SHEET_FAQ, row_number),
                    )
                )

    def _apply_resource_records(
        self,
        seeds: dict[str, GuideSeed],
        records: list[dict[str, Any]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        grouped = self._records_by_slug(records)
        for slug, seed in seeds.items():
            rows = self._require_group_records(grouped, slug, SHEET_RESOURCES)
            self._validate_unique_rows(rows, SHEET_RESOURCES, "resource_key", "sort_order")
            titles = {self._required(record, "resources_title", SHEET_RESOURCES, int(record["__row_number__"])) for record in rows}
            modes = {self._required(record, "resources_mode", SHEET_RESOURCES, int(record["__row_number__"])) for record in rows}
            if len(titles) != 1:
                raise CommandError(f"{SHEET_RESOURCES}: inconsistent resources_title values for {slug!r}.")
            if len(modes) != 1:
                raise CommandError(f"{SHEET_RESOURCES}: inconsistent resources_mode values for {slug!r}.")
            resources_mode = modes.pop()
            planning_resources_mode = planning_by_slug[slug].resources_mode
            if (
                planning_resources_mode not in DISPLAY_ONLY_MODE_VALUES
                and resources_mode != planning_resources_mode
            ):
                raise CommandError(
                    f"{SHEET_RESOURCES}: resources_mode {resources_mode!r} does not match planning resources_mode "
                    f"{planning_resources_mode!r} for {slug!r}."
                )
            seed.guide_defaults["resources_title"] = titles.pop()
            seed.guide_defaults["resources_mode"] = self._resources_mode_value(resources_mode)
            for record in sorted(rows, key=lambda item: self._parse_int(item["sort_order"], "sort_order", int(item["__row_number__"]))):
                row_number = int(record["__row_number__"])
                self._validate_category_identity(seed, record, SHEET_RESOURCES, row_number)
                seed.item_seeds.append(
                    GuideItemSeed(
                        section=self._section_value("related_resource"),
                        item_key=self._required(record, "resource_key", SHEET_RESOURCES, row_number),
                        sort_order=self._parse_int(record["sort_order"], "sort_order", row_number),
                        eyebrow=self._required(record, "label", SHEET_RESOURCES, row_number),
                        title=self._required(record, "title", SHEET_RESOURCES, row_number),
                        href=self._required(record, "href", SHEET_RESOURCES, row_number),
                    )
                )

    def _apply_final_cta_records(
        self,
        seeds: dict[str, GuideSeed],
        records: list[dict[str, Any]],
        planning_by_slug: dict[str, PlanningSeed],
    ) -> None:
        grouped = self._single_record_by_slug(records, SHEET_FINAL_CTA)
        for slug, seed in seeds.items():
            record = self._require_group_record(grouped, slug, SHEET_FINAL_CTA)
            row_number = int(record["__row_number__"])
            self._validate_category_identity(seed, record, SHEET_FINAL_CTA, row_number)
            cta_mode = self._required(record, "cta_mode", SHEET_FINAL_CTA, row_number)
            planning_cta_mode = planning_by_slug[slug].cta_mode
            if planning_cta_mode not in DISPLAY_ONLY_MODE_VALUES and cta_mode != planning_cta_mode:
                raise CommandError(
                    f"{SHEET_FINAL_CTA} row {row_number}: cta_mode {cta_mode!r} does not match planning "
                    f"final_cta_mode {planning_cta_mode!r}."
                )
            seed.guide_defaults.update(
                {
                    "cta_mode": self._cta_mode_value(cta_mode),
                    "cta_title": self._required(record, "cta_title", SHEET_FINAL_CTA, row_number),
                    "cta_copy": self._required(record, "cta_copy", SHEET_FINAL_CTA, row_number),
                    "cta_primary_label": self._required(record, "primary_label", SHEET_FINAL_CTA, row_number),
                    "cta_primary_href": self._required(record, "primary_href", SHEET_FINAL_CTA, row_number),
                    "cta_secondary_label": self._required(record, "secondary_label", SHEET_FINAL_CTA, row_number),
                    "cta_secondary_href": self._required(record, "secondary_href", SHEET_FINAL_CTA, row_number),
                }
            )

    def _validate_required_modules(self, seeds: dict[str, GuideSeed]) -> None:
        required_defaults = (
            "hero_title",
            "definition_title",
            "contexts_title",
            "matrix_title",
            "paths_title",
            "trust_title",
            "faq_title",
            "resources_title",
            "cta_title",
        )
        for slug, seed in seeds.items():
            missing = [field_name for field_name in required_defaults if field_name not in seed.guide_defaults]
            if missing:
                raise CommandError(f"{slug}: missing guide default fields: {', '.join(missing)}")
            if not seed.item_seeds:
                raise CommandError(f"{slug}: no guide item rows were prepared.")
            if not seed.faq_seeds:
                raise CommandError(f"{slug}: no guide FAQ rows were prepared.")

    def _prepare_guides(
        self,
        guide_seeds: dict[str, GuideSeed],
        categories_by_slug: dict[str, Any],
    ) -> list[PreparedGuideSeed]:
        from apps.catalog.models import ProductCategoryGuide, ProductCategoryGuideItem, ProductCategoryFaqItem

        prepared: list[PreparedGuideSeed] = []
        unmatched: list[str] = []
        for slug, seed in sorted(guide_seeds.items()):
            category = categories_by_slug.get(slug)
            if category is None:
                unmatched.append(f"{slug}: target category slug not found")
                continue
            if category.parent_id is not None:
                unmatched.append(f"{slug}: target category is not a top-level category")
                continue
            if category.name != seed.category_name:
                unmatched.append(
                    f"{slug}: workbook category_name {seed.category_name!r} does not match database "
                    f"category name {category.name!r}"
                )
                continue

            guide_model = ProductCategoryGuide(category=category, **seed.guide_defaults)
            guide_model.full_clean(validate_unique=False, validate_constraints=False)

            self._validate_guide_item_seeds(ProductCategoryGuideItem, guide_model, seed.item_seeds)
            self._validate_faq_seeds(ProductCategoryFaqItem, category, seed.faq_seeds)
            prepared.append(PreparedGuideSeed(category=category, seed=seed))

        if unmatched:
            raise CommandError(
                "Import aborted because some guide seed groups did not match ProductCategory rows:\n- "
                + "\n- ".join(unmatched)
            )
        return prepared

    def _write_guides(self, prepared: list[PreparedGuideSeed]) -> None:
        from apps.catalog.models import ProductCategoryFaqItem, ProductCategoryGuide, ProductCategoryGuideItem

        managed_sections = [
            self._section_value("definition_card"),
            self._section_value("operational_context"),
            self._section_value("decision_factor"),
            self._section_value("path"),
            self._section_value("trust_metric"),
            self._section_value("related_resource"),
        ]

        with transaction.atomic():
            media_cache: dict[str, Any] = {}
            for prepared_seed in prepared:
                seed = prepared_seed.seed
                guide_defaults = dict(seed.guide_defaults)
                guide_defaults["hero_image"] = self._resolve_media_asset(
                    media_cache=media_cache,
                    file_url=seed.hero_image_url,
                    alt_text=guide_defaults.get("hero_image_alt", ""),
                    title_hint=f"{seed.category_name} guide hero",
                )

                guide, _ = ProductCategoryGuide.objects.update_or_create(
                    category=prepared_seed.category,
                    defaults=guide_defaults,
                )

                ProductCategoryGuideItem.objects.filter(
                    guide=guide,
                    section__in=managed_sections,
                ).delete()

                guide_items: list[ProductCategoryGuideItem] = []
                for item_seed in seed.item_seeds:
                    item_model = ProductCategoryGuideItem(
                        guide=guide,
                        section=item_seed.section,
                        item_key=item_seed.item_key,
                        sort_order=item_seed.sort_order,
                        eyebrow=item_seed.eyebrow,
                        title=item_seed.title,
                        body=item_seed.body,
                        supporting_points=item_seed.supporting_points,
                        icon=item_seed.icon,
                        asset=self._resolve_media_asset(
                            media_cache=media_cache,
                            file_url=item_seed.asset_url,
                            alt_text=item_seed.asset_alt,
                            title_hint=f"{seed.category_name} guide {item_seed.title}",
                        ),
                        asset_alt=item_seed.asset_alt,
                        href=item_seed.href,
                        cta_label=item_seed.cta_label,
                        is_active=True,
                    )
                    item_model.full_clean(validate_unique=False, validate_constraints=False)
                    guide_items.append(item_model)
                ProductCategoryGuideItem.objects.bulk_create(guide_items)

                ProductCategoryFaqItem.objects.filter(
                    category=prepared_seed.category,
                    placement=ProductCategoryFaqItem.Placement.GUIDE_FAQ,
                ).delete()
                faq_items = [
                    ProductCategoryFaqItem(
                        category=prepared_seed.category,
                        placement=ProductCategoryFaqItem.Placement.GUIDE_FAQ,
                        sort_order=faq_seed.sort_order,
                        question=faq_seed.question,
                        answer=faq_seed.answer,
                        is_active=True,
                    )
                    for faq_seed in seed.faq_seeds
                ]
                for faq_item in faq_items:
                    faq_item.full_clean(validate_unique=False, validate_constraints=False)
                ProductCategoryFaqItem.objects.bulk_create(faq_items)

    def _validate_guide_item_seeds(
        self,
        guide_item_model: Any,
        guide_model: Any,
        item_seeds: list[GuideItemSeed],
    ) -> None:
        seen: set[tuple[int, int]] = set()
        seen_keys: set[tuple[int, str]] = set()
        for item_seed in item_seeds:
            sort_key = (item_seed.section, item_seed.sort_order)
            if sort_key in seen:
                raise CommandError(
                    f"{guide_model.category.slug}: duplicate guide item sort_order {item_seed.sort_order} "
                    f"for section {item_seed.section}."
                )
            seen.add(sort_key)

            key = (item_seed.section, item_seed.item_key)
            if key in seen_keys:
                raise CommandError(
                    f"{guide_model.category.slug}: duplicate guide item item_key {item_seed.item_key!r} "
                    f"for section {item_seed.section}."
                )
            seen_keys.add(key)

            item_model = guide_item_model(
                guide=guide_model,
                section=item_seed.section,
                item_key=item_seed.item_key,
                sort_order=item_seed.sort_order,
                eyebrow=item_seed.eyebrow,
                title=item_seed.title,
                body=item_seed.body,
                supporting_points=item_seed.supporting_points,
                icon=item_seed.icon,
                asset_alt=item_seed.asset_alt,
                href=item_seed.href,
                cta_label=item_seed.cta_label,
                is_active=True,
            )
            item_model.full_clean(
                exclude=("guide", "asset", "target_category", "target_resource"),
                validate_unique=False,
                validate_constraints=False,
            )

    def _validate_faq_seeds(self, faq_model: Any, category: Any, faq_seeds: list[GuideFaqSeed]) -> None:
        seen_sort_orders: set[int] = set()
        for faq_seed in faq_seeds:
            if faq_seed.sort_order in seen_sort_orders:
                raise CommandError(f"{category.slug}: duplicate Guide FAQ sort_order {faq_seed.sort_order}.")
            seen_sort_orders.add(faq_seed.sort_order)
            faq_item = faq_model(
                category=category,
                placement=faq_model.Placement.GUIDE_FAQ,
                sort_order=faq_seed.sort_order,
                question=faq_seed.question,
                answer=faq_seed.answer,
                is_active=True,
            )
            faq_item.full_clean(validate_unique=False, validate_constraints=False)

    def _resolve_media_asset(
        self,
        *,
        media_cache: dict[str, Any],
        file_url: str,
        alt_text: str,
        title_hint: str,
    ):
        if not file_url:
            return None
        from apps.core.models import MediaAsset

        if file_url in media_cache:
            return media_cache[file_url]

        title = (title_hint or alt_text or file_url.rsplit("/", 1)[-1])[:255]
        alt_text = alt_text[:255]
        asset = MediaAsset.objects.filter(file_url=file_url).order_by("id").first()
        if asset is None:
            asset = MediaAsset(
                title=title,
                asset_kind=MediaAsset.AssetKind.IMAGE,
                file_url=file_url,
                alt_text=alt_text,
            )
            asset.full_clean(validate_unique=False, validate_constraints=False)
            asset.save()
        elif alt_text and not asset.alt_text:
            asset.alt_text = alt_text
            asset.full_clean(validate_unique=False, validate_constraints=False)
            asset.save(update_fields=("alt_text", "updated_at"))

        media_cache[file_url] = asset
        return asset

    def _validate_unique_rows(
        self,
        rows: list[dict[str, Any]],
        sheet_name: str,
        key_field: str,
        sort_field: str,
    ) -> None:
        keys: set[str] = set()
        sort_orders: set[int] = set()
        for record in rows:
            row_number = int(record["__row_number__"])
            key = self._required(record, key_field, sheet_name, row_number)
            sort_order = self._parse_int(record[sort_field], sort_field, row_number)
            if key in keys:
                raise CommandError(f"{sheet_name}: duplicate {key_field} {key!r}.")
            if sort_order in sort_orders:
                raise CommandError(f"{sheet_name}: duplicate {sort_field} {sort_order} for {record['category_slug']!r}.")
            keys.add(key)
            sort_orders.add(sort_order)

    def _single_record_by_slug(self, records: list[dict[str, Any]], sheet_name: str) -> dict[str, dict[str, Any]]:
        grouped = self._records_by_slug(records)
        single_records: dict[str, dict[str, Any]] = {}
        for slug, rows in grouped.items():
            if len(rows) != 1:
                raise CommandError(f"{sheet_name}: expected exactly one row for {slug!r}, found {len(rows)}.")
            single_records[slug] = rows[0]
        return single_records

    def _records_by_slug(self, records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            grouped[self._clean(record["category_slug"])].append(record)
        return grouped

    def _require_group_record(
        self,
        grouped: dict[str, dict[str, Any]],
        category_slug: str,
        sheet_name: str,
    ) -> dict[str, Any]:
        record = grouped.get(category_slug)
        if record is None:
            raise CommandError(f"{sheet_name}: missing row for {category_slug!r}.")
        return record

    def _require_group_records(
        self,
        grouped: dict[str, list[dict[str, Any]]],
        category_slug: str,
        sheet_name: str,
    ) -> list[dict[str, Any]]:
        rows = grouped.get(category_slug, [])
        if not rows:
            raise CommandError(f"{sheet_name}: missing rows for {category_slug!r}.")
        return rows

    def _validate_category_identity(
        self,
        seed: GuideSeed,
        record: dict[str, Any],
        sheet_name: str,
        row_number: int,
    ) -> None:
        category_name = self._required(record, "category_name", sheet_name, row_number)
        category_slug = self._required(record, "category_slug", sheet_name, row_number)
        if category_name != seed.category_name:
            raise CommandError(
                f"{sheet_name} row {row_number}: category_name {category_name!r} does not match planning "
                f"category_name {seed.category_name!r} for {seed.category_slug!r}."
            )
        if category_slug != seed.category_slug:
            raise CommandError(
                f"{sheet_name} row {row_number}: category_slug {category_slug!r} does not match seed "
                f"category_slug {seed.category_slug!r}."
            )

    def _assert_finalized(self, record: dict[str, Any], sheet_name: str, row_number: int) -> None:
        status = self._clean(record["status"]).lower()
        if status not in FINALIZED_ROW_STATUSES:
            raise CommandError(
                f"{sheet_name} row {row_number}: status must be one of "
                f"{', '.join(sorted(FINALIZED_ROW_STATUSES))}."
            )

    def _required(self, record: dict[str, Any], field_name: str, sheet_name: str, row_number: int) -> str:
        value = self._clean(record.get(field_name))
        if not value:
            raise CommandError(f"{sheet_name} row {row_number}: {field_name} is required.")
        return value

    def _parse_int(self, value: Any, field_name: str, row_number: int) -> int:
        cleaned = self._clean(value)
        if not cleaned:
            raise CommandError(f"Guide import row {row_number}: {field_name} is required.")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise CommandError(f"Guide import row {row_number}: invalid {field_name} value {value!r}.") from exc

    def _path_mode_value(self, code: str) -> int:
        from apps.catalog.models import ProductCategoryGuide

        return self._enum_value_for_code(ProductCategoryGuide.PathMode, code, "paths_mode")

    def _trust_mode_value(self, code: str) -> int:
        from apps.catalog.models import ProductCategoryGuide

        return self._enum_value_for_code(ProductCategoryGuide.TrustMode, code, "standards_mode")

    def _resources_mode_value(self, code: str) -> int:
        from apps.catalog.models import ProductCategoryGuide

        return self._enum_value_for_code(ProductCategoryGuide.ResourcesMode, code, "resources_mode")

    def _cta_mode_value(self, code: str) -> int:
        from apps.catalog.models import ProductCategoryGuide

        return self._enum_value_for_code(ProductCategoryGuide.CtaMode, code, "cta_mode")

    def _section_value(self, code: str) -> int:
        from apps.catalog.models import ProductCategoryGuideItem

        return self._enum_value_for_code(ProductCategoryGuideItem.Section, code, "section")

    @staticmethod
    def _enum_value_for_code(enum_cls: Any, code: str, field_name: str) -> int:
        lookup = {enum_code: value for value, enum_code in enum_cls.codes().items()}
        if code not in lookup:
            allowed = ", ".join(sorted(lookup))
            raise CommandError(f"{field_name}: unsupported code {code!r}; allowed values: {allowed}.")
        return lookup[code]

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()
