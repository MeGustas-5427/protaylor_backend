"""
Django 管理命令：从产品应用场景工作簿导入 ProductUseCase。

使用方式（在 backend/ 目录下）：
    python manage.py import_product_use_cases
    python manage.py import_product_use_cases --excel "F:/菱威仓管/protaylor_product_use_cases.xlsx"
    python manage.py import_product_use_cases --dry-run

行为：
    - 只读取 ProductUseCases 这一个 sheet
    - 默认导入源固定为 protaylor_product_use_cases.xlsx
    - 产品优先按 Product URL 匹配，未命中时回退到产品名
    - Category / Model Code 只做校验告警，不再承担主匹配键
    - 采用“整产品覆盖”策略：命中产品后先删除旧 use cases，再按 Excel 重建
    - 工作簿结构错误或字段非法时直接失败
    - 未匹配产品只警告并跳过，不中断整批导入
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_product_use_cases.xlsx")
SHEET_USE_CASES = "ProductUseCases"

USE_CASE_HEADERS = (
    "Category",
    "Subcategory",
    "Product Name",
    "Product URL",
    "Model Code",
    "Scenario Slot",
    "Icon",
    "Title",
    "Summary",
    "Sort Order",
    "Source Buyer Fit",
    "Source Lead Text",
    "Source Application Summary",
)

MAX_ICON_LENGTH = 60
MAX_TITLE_LENGTH = 160


@dataclass(frozen=True)
class ProductSheetKey:
    category_name: str
    product_name: str


@dataclass(frozen=True)
class UseCaseImportRecord:
    key: ProductSheetKey
    product_url: str
    model_code: str
    scenario_slot: int
    icon: str
    title: str
    summary: str
    sort_order: int
    source_buyer_fit: str
    source_lead_text: str
    source_application_summary: str


class Command(BaseCommand):
    help = "Import ProductUseCase from protaylor_product_use_cases.xlsx."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to the product use case workbook. Defaults to protaylor_product_use_cases.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import Product, ProductUseCase

        excel_path = self._resolve_excel_path(options.get("excel"))
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        raw_records = self._load_sheet_records(workbook, SHEET_USE_CASES, USE_CASE_HEADERS)
        use_case_records_by_product = self._build_use_case_records(raw_records)

        total_products = len(use_case_records_by_product)
        self.stdout.write(f"Workbook products: {total_products}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no database writes."))

        product_names = {product_key.product_name for product_key in use_case_records_by_product}
        product_urls = {
            record.product_url
            for records in use_case_records_by_product.values()
            for record in records
            if record.product_url
        }
        product_query = Product.objects.all()
        if product_urls:
            product_query = product_query.filter(
                Q(source_url__in=product_urls) | Q(name__in=product_names)
            )
        else:
            product_query = product_query.filter(name__in=product_names)

        products_by_name: dict[str, list[Product]] = defaultdict(list)
        products_by_source_url: dict[str, list[Product]] = defaultdict(list)
        for product in product_query.select_related("category"):
            products_by_name[product.name].append(product)
            if product.source_url:
                products_by_source_url[product.source_url].append(product)

        matched_products = 0
        replaced_products = 0
        skipped_products = 0
        unmatched_products: list[str] = []
        ambiguous_products: list[str] = []
        category_warnings: list[str] = []
        model_code_warnings: list[str] = []

        for product_key in sorted(
            use_case_records_by_product.keys(),
            key=lambda item: (item.category_name.lower(), item.product_name.lower()),
        ):
            use_case_records = use_case_records_by_product[product_key]
            workbook_product_url = next(
                (record.product_url for record in use_case_records if record.product_url),
                "",
            )
            if workbook_product_url:
                product_matches = products_by_source_url.get(workbook_product_url, [])
                if len(product_matches) > 1:
                    skipped_products += 1
                    ambiguous_products.append(
                        f"{product_key.category_name} / {product_key.product_name}："
                        f"Product URL {workbook_product_url!r} 匹配到多个产品"
                    )
                    continue
            else:
                product_matches = []

            if not product_matches:
                product_matches = products_by_name.get(product_key.product_name, [])
                if len(product_matches) > 1:
                    narrowed_matches = [
                        product
                        for product in product_matches
                        if product.category.name == product_key.category_name
                    ]
                    if narrowed_matches:
                        product_matches = narrowed_matches

            if len(product_matches) == 0:
                skipped_products += 1
                unmatched_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：未找到产品"
                )
                continue
            if len(product_matches) > 1:
                skipped_products += 1
                ambiguous_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：产品匹配不唯一"
                )
                continue

            product = product_matches[0]
            matched_products += 1

            if product.category.name != product_key.category_name:
                category_warnings.append(
                    f"{product_key.category_name} / {product_key.product_name}："
                    f"Excel 分类={product_key.category_name!r}，数据库分类={product.category.name!r}"
                )

            workbook_model_code = next(
                (record.model_code for record in use_case_records if record.model_code),
                "",
            )
            if workbook_model_code and workbook_model_code != (product.model_code or ""):
                model_code_warnings.append(
                    f"{product.category.name} / {product.name}：Excel Model Code={workbook_model_code!r}，"
                    f"数据库 Model Code={(product.model_code or '')!r}"
                )

            if dry_run:
                replaced_products += 1
                continue

            with transaction.atomic():
                # 这里采用“整产品覆盖”而不是按标题增量 merge。
                # 这份工作簿已经是当前确认后的最终导入源，
                # 重跑时更重要的是让数据库与工作簿完全一致，而不是保留旧场景残留。
                product.use_cases.all().delete()
                for use_case_record in use_case_records:
                    ProductUseCase.objects.create(
                        product=product,
                        icon=use_case_record.icon,
                        title=use_case_record.title,
                        summary=use_case_record.summary,
                        sort_order=use_case_record.sort_order,
                    )
            replaced_products += 1

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Workbook products: {total_products}")
        self.stdout.write(f"  Matched products: {matched_products}")
        self.stdout.write(f"  Replaced products: {replaced_products}")
        self.stdout.write(f"  Skipped products: {skipped_products}")

        if unmatched_products:
            self.stdout.write(self.style.WARNING("\nUnmatched products"))
            for item in unmatched_products[:100]:
                self.stdout.write(f"  - {item}")
            if len(unmatched_products) > 100:
                self.stdout.write(f"  ... {len(unmatched_products) - 100} more")

        if ambiguous_products:
            self.stdout.write(self.style.WARNING("\nAmbiguous products"))
            for item in ambiguous_products[:100]:
                self.stdout.write(f"  - {item}")
            if len(ambiguous_products) > 100:
                self.stdout.write(f"  ... {len(ambiguous_products) - 100} more")

        if category_warnings:
            self.stdout.write(self.style.WARNING("\nCategory warnings"))
            for item in category_warnings[:100]:
                self.stdout.write(f"  - {item}")
            if len(category_warnings) > 100:
                self.stdout.write(f"  ... {len(category_warnings) - 100} more")

        if model_code_warnings:
            self.stdout.write(self.style.WARNING("\nModel code warnings"))
            for item in model_code_warnings[:100]:
                self.stdout.write(f"  - {item}")
            if len(model_code_warnings) > 100:
                self.stdout.write(f"  ... {len(model_code_warnings) - 100} more")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook,
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, object]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [_clean_text(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        column_indexes = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, object]] = []
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            record = {
                header: row[column_indexes[header]]
                if column_indexes[header] < len(row)
                else None
                for header in required_headers
            }
            if all(_clean_text(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_index
            records.append(record)
        return records

    def _build_use_case_records(
        self,
        raw_records: list[dict[str, object]],
    ) -> dict[ProductSheetKey, list[UseCaseImportRecord]]:
        use_case_records_by_product: dict[ProductSheetKey, list[UseCaseImportRecord]] = defaultdict(list)

        for raw_record in raw_records:
            row_number = int(raw_record["__row_number__"])
            category = _clean_text(raw_record["Category"])
            subcategory = _clean_text(raw_record["Subcategory"])
            product_name = _clean_text(raw_record["Product Name"])
            if not category:
                raise CommandError(f"ProductUseCases row {row_number}: Category 不能为空。")
            if not product_name:
                raise CommandError(f"ProductUseCases row {row_number}: Product Name 不能为空。")

            title = _clean_text(raw_record["Title"])
            if not title:
                raise CommandError(f"ProductUseCases row {row_number}: Title 不能为空。")
            if len(title) > MAX_TITLE_LENGTH:
                raise CommandError(
                    f"ProductUseCases row {row_number}: Title 超过 {MAX_TITLE_LENGTH} 字符。"
                )

            icon = _clean_text(raw_record["Icon"])
            if len(icon) > MAX_ICON_LENGTH:
                raise CommandError(
                    f"ProductUseCases row {row_number}: Icon 超过 {MAX_ICON_LENGTH} 字符。"
                )

            summary = _clean_text(raw_record["Summary"])
            if not summary:
                raise CommandError(f"ProductUseCases row {row_number}: Summary 不能为空。")

            product_key = ProductSheetKey(
                category_name=_resolve_target_category_name(category, subcategory),
                product_name=product_name,
            )
            use_case_records_by_product[product_key].append(
                UseCaseImportRecord(
                    key=product_key,
                    product_url=_clean_text(raw_record["Product URL"]),
                    model_code=_clean_text(raw_record["Model Code"]),
                    scenario_slot=self._parse_int(
                        raw_record["Scenario Slot"],
                        field_name="Scenario Slot",
                        row_number=row_number,
                    ),
                    icon=icon,
                    title=title,
                    summary=summary,
                    sort_order=self._parse_int(
                        raw_record["Sort Order"],
                        field_name="Sort Order",
                        row_number=row_number,
                    ),
                    source_buyer_fit=_clean_text(raw_record["Source Buyer Fit"]),
                    source_lead_text=_clean_text(raw_record["Source Lead Text"]),
                    source_application_summary=_clean_text(
                        raw_record["Source Application Summary"]
                    ),
                )
            )

        for product_key, records in use_case_records_by_product.items():
            if len(records) != 3:
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}："
                    f"每个产品必须恰好有 3 条场景，当前为 {len(records)}。"
                )

            slots = sorted(record.scenario_slot for record in records)
            if slots != [1, 2, 3]:
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}："
                    f"Scenario Slot 必须是 1/2/3，当前为 {slots!r}。"
                )

            sort_orders = [record.sort_order for record in records]
            if len(set(sort_orders)) != len(sort_orders):
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}："
                    "Sort Order 不能重复。"
                )

            use_case_records_by_product[product_key] = sorted(
                records,
                key=lambda item: (item.sort_order, item.scenario_slot, item.title.lower()),
            )

        return use_case_records_by_product

    def _parse_int(
        self,
        value: object,
        *,
        field_name: str,
        row_number: int,
    ) -> int:
        text = _clean_text(value)
        if not text:
            raise CommandError(f"ProductUseCases row {row_number}: {field_name} 不能为空。")
        try:
            return int(float(text))
        except (TypeError, ValueError) as exc:
            raise CommandError(
                f"ProductUseCases row {row_number}: {field_name} 无法解析为整数，收到 {value!r}。"
            ) from exc


def _resolve_target_category_name(category: str, subcategory: str) -> str:
    if subcategory and subcategory != category:
        return subcategory
    return category


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
