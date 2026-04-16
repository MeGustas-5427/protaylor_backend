"""
Import category-level Operational Fit / Buyer Review Focus seed rows from Excel.

Usage:
    python manage.py import_category_operational_items
    python manage.py import_category_operational_items --excel "F:/菱威仓管/protaylor_category.xlsx"
    python manage.py import_category_operational_items --sheet "Operational Items Seed"
    python manage.py import_category_operational_items --dry-run
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_category.xlsx")
DEFAULT_SHEET_NAME = "Operational Items Seed"
SEED_HEADERS = (
    "source_category",
    "source_subcategory",
    "target_category_name",
    "target_scope",
    "operational_fit_title",
    "buyer_review_focus_title",
    "section_code",
    "sort_order",
    "item_title",
    "item_body",
    "icon",
    "is_active",
    "source_fields_used",
    "primary_source_url",
    "evidence_note",
    "confidence",
)
ALL_MARKERS = {"── all ──", "-- all --", ""}
EXPECTED_SECTION_CODES = {"operational_fit", "buyer_review_focus"}
EXPECTED_SECTION_SORT_ORDERS = {1, 2, 3}


@dataclass(frozen=True)
class SeedGroupKey:
    source_category: str
    source_subcategory: str
    target_category_name: str
    target_scope: str


@dataclass(frozen=True)
class SeedItemRow:
    section_code: str
    sort_order: int
    item_title: str
    item_body: str
    icon: str
    is_active: bool


@dataclass(frozen=True)
class PreparedSeedGroup:
    category: "Any"
    operational_fit_title: str
    buyer_review_focus_title: str
    items: list[SeedItemRow]


class Command(BaseCommand):
    help = "Import ProductCategoryOperationalItem rows from the Operational Items Seed sheet."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to protaylor_category.xlsx. Defaults to F:/菱威仓管/protaylor_category.xlsx.",
        )
        parser.add_argument(
            "--sheet",
            default=DEFAULT_SHEET_NAME,
            help=f"Worksheet name to import. Defaults to {DEFAULT_SHEET_NAME!r}.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import ProductCategory, ProductCategoryOperationalItem

        excel_path = self._resolve_excel_path(options.get("excel"))
        sheet_name = self._clean(options.get("sheet")) or DEFAULT_SHEET_NAME
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        raw_rows = self._load_sheet_records(workbook, sheet_name, SEED_HEADERS)
        groups = self._build_seed_groups(raw_rows)

        self.stdout.write(f"Workbook groups: {len(groups)}")
        self.stdout.write(f"Workbook rows: {sum(len(items) for _, _, items in groups)}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))

        categories_by_name: dict[str, list[ProductCategory]] = defaultdict(list)
        for category in ProductCategory.objects.select_related("parent"):
            categories_by_name[category.name].append(category)

        prepared_groups: list[PreparedSeedGroup] = []
        unmatched_groups: list[str] = []
        ambiguous_groups: list[str] = []

        for group_key, titles, item_rows in groups:
            try:
                category = self._resolve_target_category(
                    group_key=group_key,
                    categories_by_name=categories_by_name,
                )
            except CommandError as exc:
                message = str(exc)
                if message.startswith("UNMATCHED:"):
                    unmatched_groups.append(message.removeprefix("UNMATCHED:").strip())
                    continue
                if message.startswith("AMBIGUOUS:"):
                    ambiguous_groups.append(message.removeprefix("AMBIGUOUS:").strip())
                    continue
                raise

            prepared_groups.append(
                PreparedSeedGroup(
                    category=category,
                    operational_fit_title=titles["operational_fit_title"],
                    buyer_review_focus_title=titles["buyer_review_focus_title"],
                    items=item_rows,
                )
            )

        if unmatched_groups:
            raise CommandError(
                "Import aborted because some seed groups did not match a ProductCategory:\n- "
                + "\n- ".join(unmatched_groups)
            )
        if ambiguous_groups:
            raise CommandError(
                "Import aborted because some seed groups matched multiple ProductCategory rows:\n- "
                + "\n- ".join(ambiguous_groups)
            )

        if not dry_run:
            with transaction.atomic():
                for prepared_group in prepared_groups:
                    prepared_group.category.operational_fit_title = prepared_group.operational_fit_title
                    prepared_group.category.buyer_review_focus_title = prepared_group.buyer_review_focus_title
                    prepared_group.category.save(
                        update_fields=("operational_fit_title", "buyer_review_focus_title")
                    )
                    ProductCategoryOperationalItem.objects.filter(
                        category=prepared_group.category
                    ).delete()
                    ProductCategoryOperationalItem.objects.bulk_create(
                        [
                            ProductCategoryOperationalItem(
                                category=prepared_group.category,
                                section=self._section_value_for_code(item.section_code),
                                title=item.item_title,
                                body=item.item_body,
                                icon=item.icon,
                                is_active=item.is_active,
                                sort_order=item.sort_order,
                            )
                            for item in prepared_group.items
                        ]
                    )

        imported_rows = sum(len(group.items) for group in prepared_groups)
        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Imported groups: {len(prepared_groups)}")
        self.stdout.write(f"  Imported rows: {imported_rows}")
        self.stdout.write(f"  Dry run: {dry_run}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: "Any",
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [self._clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        index = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {
                header: row[index[header]] if index[header] < len(row) else None
                for header in required_headers
            }
            if all(self._clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_number
            records.append(record)
        workbook.close()
        return records

    def _build_seed_groups(
        self,
        raw_rows: list[dict[str, Any]],
    ) -> list[tuple[SeedGroupKey, dict[str, str], list[SeedItemRow]]]:
        grouped_rows: dict[SeedGroupKey, list[dict[str, Any]]] = defaultdict(list)
        for raw_row in raw_rows:
            row_number = int(raw_row["__row_number__"])
            source_category = self._clean(raw_row["source_category"])
            source_subcategory = self._clean(raw_row["source_subcategory"])
            target_category_name = self._clean(raw_row["target_category_name"])
            target_scope = self._clean(raw_row["target_scope"])
            if not source_category:
                raise CommandError(f"Operational Items Seed row {row_number}: source_category is required.")
            if target_scope not in {"parent_all", "subcategory"}:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: target_scope must be parent_all or subcategory."
                )
            if not target_category_name:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: target_category_name is required."
                )

            normalized_subcategory = source_subcategory.lower()
            if target_scope == "parent_all" and normalized_subcategory not in ALL_MARKERS:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: parent_all rows must use an All-style source_subcategory."
                )
            if target_scope == "subcategory" and normalized_subcategory in ALL_MARKERS:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: subcategory rows must name a concrete source_subcategory."
                )

            section_code = self._clean(raw_row["section_code"])
            if section_code not in EXPECTED_SECTION_CODES:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: invalid section_code {section_code!r}."
                )

            sort_order = self._parse_int(raw_row["sort_order"], "sort_order", row_number)
            item_title = self._clean(raw_row["item_title"])
            item_body = self._clean(raw_row["item_body"])
            if not item_title:
                raise CommandError(f"Operational Items Seed row {row_number}: item_title is required.")
            if not item_body:
                raise CommandError(f"Operational Items Seed row {row_number}: item_body is required.")
            if len(item_title) > 32:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: item_title exceeds 32 characters."
                )
            if len(item_body) > 120:
                raise CommandError(
                    f"Operational Items Seed row {row_number}: item_body exceeds 120 characters."
                )

            group_key = SeedGroupKey(
                source_category=source_category,
                source_subcategory=source_subcategory,
                target_category_name=target_category_name,
                target_scope=target_scope,
            )
            grouped_rows[group_key].append(raw_row | {
                "section_code": section_code,
                "sort_order": sort_order,
                "item_title": item_title,
                "item_body": item_body,
                "icon": self._clean(raw_row["icon"]),
                "is_active": self._parse_bool(raw_row["is_active"], "is_active", row_number),
            })

        grouped_payload: list[tuple[SeedGroupKey, dict[str, str], list[SeedItemRow]]] = []
        for group_key, group_rows in sorted(
            grouped_rows.items(),
            key=lambda item: (
                item[0].source_category.lower(),
                item[0].target_scope,
                item[0].target_category_name.lower(),
            ),
        ):
            operational_titles = {self._clean(row["operational_fit_title"]) for row in group_rows}
            buyer_titles = {self._clean(row["buyer_review_focus_title"]) for row in group_rows}
            if len(operational_titles) != 1:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: inconsistent operational_fit_title values."
                )
            if len(buyer_titles) != 1:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: inconsistent buyer_review_focus_title values."
                )

            if len(group_rows) != 6:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: expected exactly 6 seed rows, found {len(group_rows)}."
                )

            rows_by_section: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in group_rows:
                rows_by_section[row["section_code"]].append(row)
            if set(rows_by_section.keys()) != EXPECTED_SECTION_CODES:
                raise CommandError(
                    f"{group_key.source_category} / {group_key.source_subcategory}: both section_code groups are required."
                )
            for section_code, section_rows in rows_by_section.items():
                if len(section_rows) != 3:
                    raise CommandError(
                        f"{group_key.source_category} / {group_key.source_subcategory}: section {section_code} must contain exactly 3 rows."
                    )
                section_sort_orders = {row["sort_order"] for row in section_rows}
                if section_sort_orders != EXPECTED_SECTION_SORT_ORDERS:
                    raise CommandError(
                        f"{group_key.source_category} / {group_key.source_subcategory}: section {section_code} must use sort_order 1, 2, and 3."
                    )

            items = [
                SeedItemRow(
                    section_code=row["section_code"],
                    sort_order=row["sort_order"],
                    item_title=row["item_title"],
                    item_body=row["item_body"],
                    icon=row["icon"],
                    is_active=row["is_active"],
                )
                for row in sorted(group_rows, key=lambda row: (row["section_code"], row["sort_order"]))
            ]
            grouped_payload.append(
                (
                    group_key,
                    {
                        "operational_fit_title": operational_titles.pop(),
                        "buyer_review_focus_title": buyer_titles.pop(),
                    },
                    items,
                )
            )
        return grouped_payload

    def _resolve_target_category(
        self,
        *,
        group_key: SeedGroupKey,
        categories_by_name: dict[str, list["Any"]],
    ):
        candidates: list[Any] = []
        target_name_candidates = categories_by_name.get(group_key.target_category_name, [])

        if group_key.target_scope == "parent_all":
            candidates = [category for category in target_name_candidates if category.parent_id is None]
        else:
            candidates = [
                category
                for category in target_name_candidates
                if category.parent_id is not None and category.parent.name == group_key.source_category
            ]
            if not candidates and len(target_name_candidates) == 1:
                candidates = target_name_candidates

        if not candidates:
            raise CommandError(
                "UNMATCHED: "
                f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_name} ({group_key.target_scope})"
            )
        if len(candidates) > 1:
            raise CommandError(
                "AMBIGUOUS: "
                f"{group_key.source_category} / {group_key.source_subcategory} -> {group_key.target_category_name} ({group_key.target_scope})"
            )
        return candidates[0]

    def _section_value_for_code(self, section_code: str) -> int:
        from apps.catalog.models import ProductCategoryOperationalItem

        section_lookup = {
            code: value
            for value, code in ProductCategoryOperationalItem.Section.codes().items()
        }
        return section_lookup[section_code]

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_int(self, value: Any, field_name: str, row_number: int) -> int:
        cleaned = self._clean(value)
        if not cleaned:
            raise CommandError(f"Operational Items Seed row {row_number}: {field_name} is required.")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise CommandError(
                f"Operational Items Seed row {row_number}: invalid {field_name} value {value!r}."
            ) from exc

    def _parse_bool(self, value: Any, field_name: str, row_number: int) -> bool:
        cleaned = self._clean(value).lower()
        if cleaned in {"true", "1", "yes", "y"}:
            return True
        if cleaned in {"false", "0", "no", "n"}:
            return False
        raise CommandError(
            f"Operational Items Seed row {row_number}: {field_name} must be a boolean-like value."
        )
