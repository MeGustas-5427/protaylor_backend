"""
Django 管理命令：从产品相关推荐工作簿导入 ProductRelation。

使用方式（在 backend/ 目录下）：
    python manage.py import_product_relations
    python manage.py import_product_relations --excel "F:/菱威仓管/product_relations_import.xlsx"
    python manage.py import_product_relations --dry-run

行为：
    - 默认导入源固定为 product_relations_import.xlsx
    - 只读取 ProductRelations 这一个 sheet
    - 默认采用“整产品覆盖”策略：命中产品后先删除旧 related product/resource，再按 Excel 重建
    - 产品优先用 source_url 匹配，匹配不到时回退到产品名
    - ResourceArticle 使用 slug 匹配
    - 工作簿结构错误或字段非法时直接失败
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\product_relations_import.xlsx")
RELATION_SHEET = "ProductRelations"
RELATION_HEADERS = (
    "source_product_name",
    "source_product_url",
    "relation_type",
    "sort_order",
    "related_product_name",
    "related_product_url",
    "related_resource_slug",
    "related_resource_title",
    "resource_topic_id",
    "mapping_reason",
    "confidence",
    "source_relation_qa_status",
    "import_status",
)


@dataclass(frozen=True)
class PreparedRelation:
    product: "Any"
    relation_type: int
    sort_order: int
    related_product: "Any | None"
    related_resource: "Any | None"


class Command(BaseCommand):
    help = "Import product related-product and related-resource rows from product_relations_import.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to the product relations import workbook. Defaults to product_relations_import.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Resolve and validate rows without writing to the database.",
        )
        parser.add_argument(
            "--append",
            action="store_true",
            help="Append imported relations instead of replacing existing related product/resource relations for imported products.",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required: pip install openpyxl") from exc

        from apps.catalog.models import Product, ProductRelation, ProductRelationType
        from apps.content.models import ResourceArticle

        excel_path = self._resolve_excel_path(options.get("excel"))

        dry_run: bool = options["dry_run"]
        append: bool = options["append"]
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))
        if append:
            self.stdout.write(self.style.WARNING("APPEND mode - existing relations will not be replaced."))
        self.stdout.write(f"Reading: {excel_path}")

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            relation_rows = self._load_sheet_records(workbook, RELATION_SHEET, RELATION_HEADERS)
        finally:
            workbook.close()

        if not relation_rows:
            raise CommandError("No ProductRelations rows found.")
        self.stdout.write(f"Loaded {len(relation_rows)} ProductRelations rows.")

        product_names: set[str] = set()
        product_source_urls: set[str] = set()
        resource_slugs: set[str] = set()
        for row in relation_rows:
            source_name = _clean(row.get("source_product_name"))
            related_name = _clean(row.get("related_product_name"))
            source_url = _clean(row.get("source_product_url"))
            related_url = _clean(row.get("related_product_url"))
            resource_slug = _clean(row.get("related_resource_slug"))
            if source_name:
                product_names.add(source_name)
            if related_name:
                product_names.add(related_name)
            if source_url:
                product_source_urls.add(source_url)
            if related_url:
                product_source_urls.add(related_url)
            if resource_slug:
                resource_slugs.add(resource_slug)

        product_query = Q()
        if product_names:
            product_query |= Q(name__in=product_names)
        if product_source_urls:
            product_query |= Q(source_url__in=product_source_urls)

        products_by_name: dict[str, list[Product]] = defaultdict(list)
        products_by_source_url: dict[str, list[Product]] = defaultdict(list)
        if product_query:
            for product in Product.objects.filter(product_query).select_related("category"):
                products_by_name[product.name.lower()].append(product)
                if product.source_url:
                    products_by_source_url[product.source_url].append(product)

        resources_by_slug = {
            resource.slug: resource
            for resource in ResourceArticle.objects.filter(slug__in=resource_slugs)
        }

        prepared: list[PreparedRelation] = []
        errors: list[str] = []
        imported_products: set[int] = set()

        for row_number, row in enumerate(relation_rows, start=2):
            try:
                relation_type = _relation_type_value(ProductRelationType, row.get("relation_type"))
                product = _resolve_product(
                    products_by_name=products_by_name,
                    products_by_source_url=products_by_source_url,
                    name=_required(row, "source_product_name"),
                    source_url=_clean(row.get("source_product_url")),
                    role="source product",
                )
                related_product = None
                related_resource = None

                if relation_type == ProductRelationType.RELATED_PRODUCT:
                    related_product = _resolve_product(
                        products_by_name=products_by_name,
                        products_by_source_url=products_by_source_url,
                        name=_required(row, "related_product_name"),
                        source_url=_clean(row.get("related_product_url")),
                        role="related product",
                    )
                    if related_product.pk == product.pk:
                        raise CommandError("Related product cannot be the source product.")
                else:
                    slug = _required(row, "related_resource_slug")
                    related_resource = resources_by_slug.get(slug)
                    if related_resource is None:
                        raise CommandError(f"Related ResourceArticle not found by slug: {slug}")

                prepared.append(
                    PreparedRelation(
                        product=product,
                        relation_type=int(relation_type),
                        sort_order=_int_value(row.get("sort_order")),
                        related_product=related_product,
                        related_resource=related_resource,
                    )
                )
                imported_products.add(product.pk)
            except CommandError as exc:
                errors.append(f"row {row_number}: {exc}")

        if errors:
            preview = "\n".join(errors[:20])
            suffix = f"\n... {len(errors) - 20} more errors" if len(errors) > 20 else ""
            raise CommandError(f"Product relation import validation failed:\n{preview}{suffix}")

        counts = {
            "related_product": sum(1 for item in prepared if item.relation_type == ProductRelationType.RELATED_PRODUCT),
            "related_resource": sum(1 for item in prepared if item.relation_type == ProductRelationType.RELATED_RESOURCE),
        }

        if dry_run:
            self.stdout.write(self.style.SUCCESS("Import summary"))
            self.stdout.write(f"  Workbook relations: {len(relation_rows)}")
            self.stdout.write(f"  Matched products: {len(imported_products)}")
            self.stdout.write(f"  Related product rows: {counts['related_product']}")
            self.stdout.write(f"  Related resource rows: {counts['related_resource']}")
            self.stdout.write("  Database writes: 0")
            return

        created = replaced = 0
        with transaction.atomic():
            if not append:
                deleted, _ = ProductRelation.objects.filter(
                    product_id__in=imported_products,
                    relation_type__in=[
                        ProductRelationType.RELATED_PRODUCT,
                        ProductRelationType.RELATED_RESOURCE,
                    ],
                ).delete()
                replaced = deleted

            for item in prepared:
                ProductRelation.objects.create(
                    product=item.product,
                    relation_type=item.relation_type,
                    sort_order=item.sort_order,
                    related_product=item.related_product,
                    related_resource=item.related_resource,
                )
                created += 1

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Workbook relations: {len(relation_rows)}")
        self.stdout.write(f"  Matched products: {len(imported_products)}")
        self.stdout.write(f"  Replaced existing relations: {replaced}")
        self.stdout.write(f"  Created relations: {created}")
        self.stdout.write(f"  Related product rows: {counts['related_product']}")
        self.stdout.write(f"  Related resource rows: {counts['related_resource']}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: "Any",
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [_clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        column_indexes = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, Any]] = []
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            record = {
                header: row[column_indexes[header]]
                if column_indexes[header] < len(row)
                else None
                for header in required_headers
            }
            if all(_clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_index
            records.append(record)
        return records


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _required(row: dict[str, Any], key: str) -> str:
    value = _clean(row.get(key))
    if not value:
        raise CommandError(f"Missing required field '{key}'.")
    return value


def _relation_type_value(enum_cls: "Any", raw_value: Any) -> int:
    value = _clean(raw_value)
    if not value:
        raise CommandError("Missing required field 'relation_type'.")
    normalized = value.lower().replace("-", "_").replace(" ", "_")
    for db_value, code in enum_cls.codes().items():
        if normalized in {code.lower(), str(db_value), enum_cls(db_value).name.lower()}:
            return int(db_value)
    raise CommandError(f"Invalid relation_type value: {value}")


def _resolve_product(
    *,
    products_by_name: dict[str, list["Any"]],
    products_by_source_url: dict[str, list["Any"]],
    name: str,
    source_url: str,
    role: str,
) -> "Any":
    if source_url:
        url_matches = products_by_source_url.get(source_url, [])
        if len(url_matches) == 1:
            return url_matches[0]
        if len(url_matches) > 1:
            raise CommandError(f"Ambiguous {role} source_url: {source_url}")

    matches = products_by_name.get(name.lower(), [])
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise CommandError(f"Ambiguous {role} name: {name}")

    if source_url:
        raise CommandError(f"Could not resolve {role}: {name} ({source_url})")
    raise CommandError(f"Could not resolve {role}: {name}")


def _int_value(value: Any, *, default: int = 0) -> int:
    cleaned = _clean(value)
    if not cleaned:
        return default
    try:
        return int(float(cleaned))
    except ValueError as exc:
        raise CommandError(f"Invalid integer value: {value}") from exc
