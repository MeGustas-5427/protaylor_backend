"""
Django 管理命令：从资源文章导入工作簿导入 ResourceArticle 与 FAQItem。

使用方式（在 backend/ 目录下）：
    python manage.py import_resource_articles
    python manage.py import_resource_articles --excel "F:/菱威仓管/resource_articles_import.xlsx"
    python manage.py import_resource_articles --dry-run

行为：
    - 默认导入源固定为 resource_articles_import.xlsx
    - 读取 ResourceArticles sheet，按 slug 幂等创建/更新 ResourceArticle
    - 读取 FAQItems sheet，命中文章后先删除旧 FAQ，再按 Excel 重建
    - 工作簿结构错误或字段非法时直接失败
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\resource_articles_import.xlsx")
ARTICLE_SHEET = "ResourceArticles"
FAQ_SHEET = "FAQItems"

ARTICLE_HEADERS = (
    "topic_id",
    "resource_type",
    "title",
    "slug",
    "url_path",
    "h1",
    "seo_title",
    "meta_description",
    "summary",
    "lead_text",
    "body",
    "primary_query",
    "secondary_queries",
    "status",
    "index_mode",
    "published_at",
    "qa_status",
    "is_product_card_mapped",
    "direct_product_count",
    "topic_plan_status",
    "notes",
)
FAQ_HEADERS = (
    "topic_id",
    "resource_slug",
    "sort_order",
    "question",
    "answer",
    "is_featured",
)


class Command(BaseCommand):
    help = "Import ResourceArticle rows and FAQItems from resource_articles_import.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to the resource article import workbook. Defaults to resource_articles_import.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )
        parser.add_argument(
            "--skip-faq",
            action="store_true",
            help="Import ResourceArticle rows only; leave FAQItem rows unchanged.",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required: pip install openpyxl") from exc

        from apps.content.models import FAQItem, ResourceArticle, ResourceType
        from common.enums import IndexMode, PublishStatus

        excel_path = self._resolve_excel_path(options.get("excel"))

        dry_run: bool = options["dry_run"]
        skip_faq: bool = options["skip_faq"]
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))
        self.stdout.write(f"Reading: {excel_path}")

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            article_rows = self._load_sheet_records(workbook, ARTICLE_SHEET, ARTICLE_HEADERS)
            faq_rows = [] if skip_faq else self._load_sheet_records(workbook, FAQ_SHEET, FAQ_HEADERS)
        finally:
            workbook.close()

        if not article_rows:
            raise CommandError("No ResourceArticle rows found.")
        self.stdout.write(f"Loaded {len(article_rows)} ResourceArticle rows and {len(faq_rows)} FAQItem rows.")

        article_slugs = {_clean(row.get("slug")) for row in article_rows if _clean(row.get("slug"))}
        if len(article_slugs) != len(article_rows):
            raise CommandError("Every ResourceArticle row must have a unique non-empty slug.")

        faq_by_slug: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in faq_rows:
            slug = _clean(row.get("resource_slug"))
            if not slug:
                raise CommandError("Every FAQItems row must include resource_slug.")
            if slug not in article_slugs:
                raise CommandError(f"FAQItems row points at unknown resource_slug: {slug}")
            faq_by_slug[slug].append(row)

        created = updated = faq_created = 0
        if dry_run:
            self.stdout.write(self.style.SUCCESS("Import summary"))
            self.stdout.write(f"  ResourceArticle rows: {len(article_rows)}")
            self.stdout.write(f"  FAQItem rows: {len(faq_rows)}")
            self.stdout.write("  Database writes: 0")
            return

        with transaction.atomic():
            for row in article_rows:
                slug = _required(row, "slug")
                status = _enum_value(PublishStatus, row.get("status"), default=PublishStatus.PUBLISHED)
                published_at = _parse_datetime(row.get("published_at"))
                if status == PublishStatus.PUBLISHED and published_at is None:
                    published_at = timezone.now()

                defaults = {
                    "resource_type": _enum_value(ResourceType, row.get("resource_type"), default=ResourceType.GUIDE),
                    "title": _required(row, "title"),
                    "summary": _clean(row.get("summary")),
                    "url_path": _required(row, "url_path"),
                    "h1": _required(row, "h1"),
                    "seo_title": _required(row, "seo_title"),
                    "meta_description": _required(row, "meta_description"),
                    "canonical_url": _clean(row.get("canonical_url")),
                    "index_mode": _enum_value(IndexMode, row.get("index_mode"), default=IndexMode.INDEX),
                    "lead_text": _clean(row.get("lead_text")),
                    "primary_query": _clean(row.get("primary_query")),
                    "secondary_queries": _clean(row.get("secondary_queries")),
                    "status": status,
                    "published_at": published_at,
                    "body": _required(row, "body"),
                }

                article, was_created = ResourceArticle.objects.update_or_create(
                    slug=slug,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

                if not skip_faq:
                    content_type = ContentType.objects.get_for_model(article, for_concrete_model=False)
                    FAQItem.objects.filter(content_type=content_type, object_id=article.pk).delete()
                    for faq_row in sorted(faq_by_slug.get(slug, []), key=lambda item: _int_value(item.get("sort_order"))):
                        FAQItem.objects.create(
                            content_type=content_type,
                            object_id=article.pk,
                            question=_required(faq_row, "question"),
                            answer=_required(faq_row, "answer"),
                            is_featured=_bool_value(faq_row.get("is_featured"), default=True),
                            sort_order=_int_value(faq_row.get("sort_order")),
                        )
                        faq_created += 1

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  ResourceArticle rows: {len(article_rows)}")
        self.stdout.write(f"  ResourceArticle created: {created}")
        self.stdout.write(f"  ResourceArticle updated: {updated}")
        self.stdout.write(f"  FAQItem recreated: {faq_created}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: "Any",
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [_clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        column_indexes = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, Any]] = []
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            record = {
                header: row[column_indexes[header]]
                if column_indexes[header] < len(row)
                else None
                for header in required_headers
            }
            if all(_clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_index
            records.append(record)
        return records


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _required(row: dict[str, Any], key: str) -> str:
    value = _clean(row.get(key))
    if not value:
        raise CommandError(f"Missing required field '{key}'.")
    return value


def _enum_value(enum_cls: "Any", raw_value: Any, *, default: "Any") -> int:
    value = _clean(raw_value)
    if not value:
        return int(default)
    normalized = value.lower().replace("-", "_").replace(" ", "_")
    for db_value, code in enum_cls.codes().items():
        if normalized in {code.lower(), str(db_value), enum_cls(db_value).name.lower()}:
            return int(db_value)
    raise CommandError(f"Invalid {enum_cls.__name__} value: {value}")


def _parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return timezone.make_aware(value) if timezone.is_naive(value) else value
    parsed = parse_datetime(str(value).strip())
    if parsed is None:
        raise CommandError(f"Invalid datetime value: {value}")
    return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed


def _int_value(value: Any, *, default: int = 0) -> int:
    cleaned = _clean(value)
    if not cleaned:
        return default
    try:
        return int(float(cleaned))
    except ValueError as exc:
        raise CommandError(f"Invalid integer value: {value}") from exc


def _bool_value(value: Any, *, default: bool = False) -> bool:
    cleaned = _clean(value).lower()
    if not cleaned:
        return default
    if cleaned in {"1", "true", "yes", "y"}:
        return True
    if cleaned in {"0", "false", "no", "n"}:
        return False
    raise CommandError(f"Invalid boolean value: {value}")
