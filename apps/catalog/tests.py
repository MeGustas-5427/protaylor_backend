from __future__ import annotations

import tempfile
from pathlib import Path

from django.contrib.contenttypes.models import ContentType
from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from openpyxl import Workbook

from apps.catalog.models import (
    Product,
    ProductCategory,
    ProductCategoryComparisonOverview,
    ProductCategoryComparisonRow,
    ProductCategoryFaqItem,
    ProductCategoryGuide,
    ProductCategoryGuideItem,
    ProductCategoryOperationalItem,
    ProductMedia,
    ProductRelation,
    ProductSpecGroup,
    ProductSpecRow,
)
from apps.catalog.services import get_product_detail
from apps.content.models import FAQItem, ResourceArticle
from apps.core.models import MediaAsset
from apps.core.models import PageSEO
from tests_support import build_api_client


class CatalogApiTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        call_command("seed_sample_site")

        cls.primary_product = Product.objects.select_related("category").get(
            slug="icm-t838-twin-twist-soft-serve-machine"
        )
        category = Product.objects.select_related("category").get(
            slug="icm-t838-twin-twist-soft-serve-machine"
        ).category
        product_image = MediaAsset.objects.get(title="ICM-T838 product hero")
        cls.related_resource = ResourceArticle.objects.get(slug="soft-serve-buying-guide")

        # Add a second canonical product so pagination overflow can be verified
        # against a multi-page result set without depending on production data.
        cls.extra_product = Product.objects.create(
            category=category,
            slug="icm-c220-compact-counter-soft-serve-machine",
            url_path="/products/soft-ice-cream-machine/icm-c220-compact-counter-soft-serve-machine/",
            name="ICM-C220 Compact Counter Soft Serve Machine",
            model_code="ICM-C220",
            h1="ICM-C220 Compact Counter Soft Serve Machine",
            summary="Compact soft serve model for cafes and low-footprint counters.",
            lead_text="A compact soft serve option for operators balancing footprint, menu range, and daily cleaning cadence.",
            seo_title="ICM-C220 Compact Counter Soft Serve Machine | PRO-TAYLOR",
            meta_description="Review the ICM-C220 compact counter soft serve machine for cafes and low-footprint dessert counters.",
            primary_query="compact counter soft serve machine",
            status=Product.Status.PUBLISHED,
            index_mode=Product.IndexMode.INDEX,
            is_canonical=True,
        )
        ProductMedia.objects.create(
            product=cls.extra_product,
            asset=product_image,
            media_kind=ProductMedia.MediaKind.HERO,
            is_primary=True,
            sort_order=10,
        )
        ProductRelation.objects.update_or_create(
            product=cls.primary_product,
            relation_type=ProductRelation.RelationType.RELATED_PRODUCT,
            related_product=cls.extra_product,
            defaults={"sort_order": 20},
        )

        def create_category(
            *,
            name: str,
            slug: str,
            url_path: str,
            parent: ProductCategory | None = None,
            operational_fit_title: str = "",
            buyer_review_focus_title: str = "",
            sourcing_faq_title: str = "",
        ) -> ProductCategory:
            return ProductCategory.objects.create(
                name=name,
                slug=slug,
                url_path=url_path,
                h1=name,
                summary=f"{name} summary",
                lead_text=f"{name} lead text",
                buyer_fit=f"{name} buyer fit",
                selection_guide=f"{name} selection guide",
                operational_fit_title=operational_fit_title,
                buyer_review_focus_title=buyer_review_focus_title,
                sourcing_faq_title=sourcing_faq_title,
                seo_title=f"{name} | PRO-TAYLOR",
                meta_description=f"{name} meta description",
                primary_query=name.lower(),
                status=ProductCategory.Status.PUBLISHED,
                index_mode=ProductCategory.IndexMode.INDEX,
                parent=parent,
            )

        def create_listing_ready_product(
            *,
            category: ProductCategory,
            slug: str,
            name: str,
            url_path: str,
        ) -> Product:
            product = Product.objects.create(
                category=category,
                slug=slug,
                url_path=url_path,
                name=name,
                model_code=slug.upper(),
                h1=name,
                summary=f"{name} summary",
                lead_text=f"{name} lead text",
                seo_title=f"{name} | PRO-TAYLOR",
                meta_description=f"{name} meta description",
                primary_query=name.lower(),
                status=Product.Status.PUBLISHED,
                index_mode=Product.IndexMode.INDEX,
                is_canonical=True,
            )
            ProductMedia.objects.create(
                product=product,
                asset=product_image,
                media_kind=ProductMedia.MediaKind.HERO,
                is_primary=True,
                sort_order=10,
            )
            quick_group = ProductSpecGroup.objects.create(
                product=product,
                title="Quick Facts",
                group_kind=ProductSpecGroup.GroupKind.QUICK_FACTS,
                sort_order=10,
            )
            ProductSpecRow.objects.create(
                group=quick_group,
                label="Output",
                value="30",
                unit="L/h",
                is_highlight=True,
                sort_order=10,
            )
            return product

        cls.parent_category = create_category(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            operational_fit_title="Operational Fit",
            buyer_review_focus_title="Buyer Review Focus",
            sourcing_faq_title="Sourcing FAQ",
        )
        cls.gelato_child = create_category(
            name="Gelato Batch Freezer",
            slug="gelato-batch-freezer",
            url_path="/products/gelato-batch-freezer/",
            parent=cls.parent_category,
            operational_fit_title="Gelato Operational Fit",
            buyer_review_focus_title="Gelato Buyer Review Focus",
            sourcing_faq_title="Gelato Sourcing FAQ",
        )
        cls.roll_child = create_category(
            name="Roll Ice Cream Machine",
            slug="roll-ice-cream-machine",
            url_path="/products/roll-ice-cream-machine/",
            parent=cls.parent_category,
        )
        cls.parent_direct_product = create_listing_ready_product(
            category=cls.parent_category,
            slug="ice-cream-machine-direct-model",
            name="Ice Cream Machine Direct Model",
            url_path="/products/ice-cream-machine/ice-cream-machine-direct-model/",
        )
        cls.gelato_product = create_listing_ready_product(
            category=cls.gelato_child,
            slug="gelato-test-model",
            name="Gelato Test Model",
            url_path="/products/gelato-batch-freezer/gelato-test-model/",
        )
        cls.roll_product = create_listing_ready_product(
            category=cls.roll_child,
            slug="roll-test-model",
            name="Roll Test Model",
            url_path="/products/roll-ice-cream-machine/roll-test-model/",
        )
        ProductCategoryOperationalItem.objects.bulk_create(
            [
                ProductCategoryOperationalItem(
                    category=cls.parent_category,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Application Matching",
                    body="Review equipment families against service format and expected daily throughput.",
                    icon="storefront",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.parent_category,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Utility Planning",
                    body="Confirm voltage, cooling method, and placement constraints before final shortlist review.",
                    icon="bolt",
                    sort_order=20,
                ),
                ProductCategoryOperationalItem(
                    category=cls.parent_category,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Capacity vs. Actual Demand",
                    body="Match peak-hour demand, recovery expectations, and product mix before sizing up.",
                    icon="query_stats",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.parent_category,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Utilities and Footprint",
                    body="Check electrical supply, cooling environment, drainage, and working space before order lock.",
                    icon="home_work",
                    sort_order=20,
                ),
                ProductCategoryOperationalItem(
                    category=cls.gelato_child,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Batch Capacity Fit",
                    body="Review batch size, extraction timing, and prep rhythm before comparing freezer classes.",
                    icon="speed",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.gelato_child,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Kitchen Workflow",
                    body="Check charging, rinse, and pan transfer space before locking in a batch freezer footprint.",
                    icon="storefront",
                    sort_order=20,
                ),
                ProductCategoryOperationalItem(
                    category=cls.gelato_child,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Freeze Curve Review",
                    body="Compare draw time, overrun control, and compressor recovery against daily production targets.",
                    icon="fact_check",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.gelato_child,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Utilities by Batch",
                    body="Confirm power, clearance, and condenser conditions before selecting a higher-output batch platform.",
                    icon="bolt",
                    sort_order=20,
                ),
            ]
        )
        ProductCategoryFaqItem.objects.bulk_create(
            [
                ProductCategoryFaqItem(
                    category=cls.parent_category,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="How should buyers compare ice cream machine models?",
                    answer="Start with service format, output target, utilities, and working space before asking suppliers to confirm final fit.",
                    sort_order=10,
                ),
                ProductCategoryFaqItem(
                    category=cls.parent_category,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="What details should be prepared before requesting a quote?",
                    answer="Prepare target models, intended use case, preferred voltage, destination market, and any OEM or certification needs.",
                    sort_order=20,
                ),
                ProductCategoryFaqItem(
                    category=cls.gelato_child,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="What should buyers check before choosing a gelato batch freezer?",
                    answer="Review batch size, extraction timing, blast-freeze handoff, and operator rhythm before committing to a freezer platform.",
                    sort_order=10,
                ),
                ProductCategoryFaqItem(
                    category=cls.gelato_child,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="Why does kitchen workflow matter for gelato batch freezers?",
                    answer="Charging, rinsing, and pan-transfer space can limit throughput even when the freezer headline capacity looks sufficient.",
                    sort_order=20,
                ),
            ]
        )
        cls.parent_comparison_overview = ProductCategoryComparisonOverview.objects.create(
            category=cls.parent_category,
            title="Ice Cream Machine Route Comparison",
            intro=(
                "Use this table to decide whether soft serve, batch freezing, or roll ice cream "
                "is the right production route before comparing individual models."
            ),
            dimension_heading="Decision Dimension",
            subjects_json=[
                {
                    "subject_key": "soft_ice_cream_machine",
                    "label": "Soft Ice Cream Machine",
                    "route_category_slug": "soft-ice-cream-machine",
                    "sort_order": 1,
                },
                {
                    "subject_key": "gelato_batch_freezer",
                    "label": "Gelato Batch Freezer",
                    "route_category_slug": "gelato-batch-freezer",
                    "sort_order": 2,
                },
                {
                    "subject_key": "roll_ice_cream_machine",
                    "label": "Roll Ice Cream Machine",
                    "route_category_slug": "roll-ice-cream-machine",
                    "sort_order": 3,
                },
            ],
        )
        ProductCategoryComparisonRow.objects.bulk_create(
            [
                ProductCategoryComparisonRow(
                    overview=cls.parent_comparison_overview,
                    row_key="best_fit_service_format",
                    label="Best-Fit Service Format",
                    sort_order=10,
                    cells_json={
                        "soft_ice_cream_machine": "Direct-draw service from the machine during service.",
                        "gelato_batch_freezer": (
                            "Batch-freeze product, then extract it for serving, holding, or later hardening."
                        ),
                        "roll_ice_cream_machine": (
                            "Freeze the mix on a pan, scrape it, and serve each order directly."
                        ),
                    },
                ),
                ProductCategoryComparisonRow(
                    overview=cls.parent_comparison_overview,
                    row_key="production_rhythm",
                    label="Production Rhythm",
                    sort_order=20,
                    cells_json={
                        "soft_ice_cream_machine": "Continuous dispensing during service.",
                        "gelato_batch_freezer": "Planned batch cycles with refill and restart between runs.",
                        "roll_ice_cream_machine": "One-pan or two-pan made-to-order preparation.",
                    },
                ),
            ]
        )

        cls.single_child_parent = create_category(
            name="Home Use Slush Machine",
            slug="home-use-slush-machine",
            url_path="/products/home-use-slush-machine/",
            operational_fit_title="Home Slush Operational Fit",
            buyer_review_focus_title="Home Slush Buyer Review Focus",
            sourcing_faq_title="Home Slush Sourcing FAQ",
        )
        cls.single_child = create_category(
            name="2L Slush Machine",
            slug="2l-slush-machine",
            url_path="/products/2l-slush-machine/",
            parent=cls.single_child_parent,
            operational_fit_title="2L Slush Operational Fit",
            buyer_review_focus_title="2L Slush Buyer Review Focus",
            sourcing_faq_title="2L Slush Sourcing FAQ",
        )
        cls.single_child_product = create_listing_ready_product(
            category=cls.single_child,
            slug="2l-test-model",
            name="2L Test Model",
            url_path="/products/2l-slush-machine/2l-test-model/",
        )
        ProductCategoryOperationalItem.objects.bulk_create(
            [
                ProductCategoryOperationalItem(
                    category=cls.single_child_parent,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Parent Footprint Fit",
                    body="Use the parent listing view to frame compact frozen drink options before deciding whether a narrower bowl size matters.",
                    icon="storefront",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.single_child_parent,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Parent Throughput Framing",
                    body="Validate how much rush pressure the category must absorb before leaning on a compact slush machine answer.",
                    icon="query_stats",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.single_child,
                    section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
                    title="Compact Counter Fit",
                    body="Match true rush length, refill rhythm, and counter width before defaulting to a 2L tabletop machine.",
                    icon="storefront",
                    sort_order=10,
                ),
                ProductCategoryOperationalItem(
                    category=cls.single_child,
                    section=ProductCategoryOperationalItem.Section.BUYER_REVIEW_FOCUS,
                    title="Cooling and Ventilation Limits",
                    body="Check ambient heat, side clearance, and recovery expectations before treating a small bowl machine as all-day equipment.",
                    icon="bolt",
                    sort_order=10,
                ),
            ]
        )
        ProductCategoryFaqItem.objects.bulk_create(
            [
                ProductCategoryFaqItem(
                    category=cls.single_child_parent,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="How should buyers compare home use slush machine models?",
                    answer="Start with drink type, service rhythm, counter footprint, and operating environment before narrowing to a specific compact bowl size.",
                    sort_order=10,
                ),
                ProductCategoryFaqItem(
                    category=cls.single_child_parent,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="What should buyers confirm before requesting a home use slush machine quote?",
                    answer="Prepare peak servings, flavor plan, voltage, placement conditions, and refill expectations before asking for a compact slush recommendation.",
                    sort_order=20,
                ),
                ProductCategoryFaqItem(
                    category=cls.single_child,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="When is a 2L slush machine the right choice?",
                    answer="Use a 2L tabletop slush machine when menu scope is narrow, service rhythm is moderate, and the counter footprint cannot support a larger frozen drink platform.",
                    sort_order=10,
                ),
                ProductCategoryFaqItem(
                    category=cls.single_child,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="What should buyers confirm before requesting a 2L slush machine quote?",
                    answer="Confirm drink type, peak servings, flavor count, voltage, usable counter width, refill rhythm, and recipe sugar range before comparing models.",
                    sort_order=20,
                ),
            ]
        )

        resource_image = MediaAsset.objects.create(
            title="Soft serve buying guide og",
            asset_kind=MediaAsset.AssetKind.IMAGE,
            file_url="https://cdn.example.com/protaylor/soft-serve-buying-guide-og.jpg",
            alt_text="Soft serve buying guide resource card",
            mime_type="image/jpeg",
        )
        PageSEO.objects.update_or_create(
            content_type=ContentType.objects.get_for_model(cls.related_resource, for_concrete_model=False),
            object_id=cls.related_resource.id,
            defaults={"og_image": resource_image},
        )

        FAQItem.objects.create(
            content_type=ContentType.objects.get_for_model(cls.primary_product, for_concrete_model=False),
            object_id=cls.primary_product.id,
            question="Does the ICM-T838 support OEM branding?",
            answer="Yes. Branding, control panel language, and export preparation can be aligned before order confirmation.",
            is_featured=False,
            sort_order=30,
        )

    def setUp(self) -> None:
        self.client = build_api_client()

    def test_category_endpoint_returns_products(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/soft-ice-cream-machine")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["slug"], "soft-ice-cream-machine")
        self.assertGreaterEqual(len(payload["products"]), 1)

    def test_category_products_endpoint_returns_paginated_listing_contract(self) -> None:
        # 这里固定 page_size=1，是为了让分页窗口更稳定。
        # 即使将来这个分类下再补更多样例产品，这组断言也不会轻易漂移。
        response = self.client.get("/api/v1/catalog/categories/soft-ice-cream-machine/products?page=1&page_size=1")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["slug"], "soft-ice-cream-machine")
        self.assertEqual(payload["pagination"]["requested_page"], 1)
        self.assertEqual(payload["pagination"]["current_page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 1)
        self.assertEqual(payload["pagination"]["total_items"], 2)
        self.assertEqual(payload["pagination"]["total_pages"], 2)
        self.assertEqual(len(payload["items"]), 1)
        self.assertIn("card_image_url", payload["items"][0])
        self.assertIn("metrics", payload["items"][0])
        self.assertEqual(payload["operational_fit_title"], "Operational Fit")
        self.assertEqual(payload["buyer_review_focus_title"], "Buyer Review Focus")
        self.assertEqual(payload["sourcing_faq_title"], "Sourcing FAQ")
        self.assertEqual(payload["operational_fit_items"], [])
        self.assertEqual(payload["buyer_review_focus_items"], [])
        self.assertEqual(payload["sourcing_faq_items"], [])

    def test_category_products_endpoint_overflow_page_returns_last_page(self) -> None:
        # 对这个公开 B2B 分类目录来说，超界页码应该回最后一页，
        # 不能把“分类存在但页码过期”误判成“资源不存在”。
        response = self.client.get(
            "/api/v1/catalog/categories/soft-ice-cream-machine/products?page=999&page_size=1"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["pagination"]["requested_page"], 999)
        self.assertEqual(payload["pagination"]["current_page"], 2)
        self.assertFalse(payload["pagination"]["has_next"])
        self.assertTrue(payload["pagination"]["has_previous"])
        self.assertEqual(payload["items"][0]["slug"], "icm-t838-twin-twist-soft-serve-machine")

    def test_product_paths_endpoint_returns_published_canonical_pairs(self) -> None:
        # 这个接口只服务前端 `generateStaticParams()`，
        # 所以测试只验证 category/product slug 对，不去耦合更重的详情字段。
        response = self.client.get("/api/v1/catalog/products/paths")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(
            {
                "category_slug": "soft-ice-cream-machine",
                "product_slug": "icm-t838-twin-twist-soft-serve-machine",
            },
            payload,
        )
        self.assertIn(
            {
                "category_slug": "soft-ice-cream-machine",
                "product_slug": self.extra_product.slug,
            },
            payload,
        )

    def test_category_paths_endpoint_returns_published_category_pairs(self) -> None:
        # 这个接口只服务前端分类静态路径发现，因此断言只钉住最小字段契约。
        response = self.client.get("/api/v1/catalog/categories/paths")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(
            {
                "slug": "soft-ice-cream-machine",
                "url_path": "/products/soft-ice-cream-machine/",
            },
            payload,
        )
        self.assertIn(
            {
                "slug": "ice-cream-machine",
                "url_path": "/products/ice-cream-machine/",
            },
            payload,
        )
        self.assertNotIn(
            {
                "slug": "gelato-batch-freezer",
                "url_path": "/products/gelato-batch-freezer/",
            },
            payload,
        )

    def test_categories_endpoint_returns_top_level_navigation_cards(self) -> None:
        response = self.client.get("/api/v1/catalog/categories")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ice_cream_category = next(item for item in payload if item["slug"] == "ice-cream-machine")
        self.assertEqual(ice_cream_category["product_count"], 3)
        self.assertTrue(ice_cream_category["has_children"])
        self.assertNotIn("gelato-batch-freezer", [item["slug"] for item in payload])

    def test_parent_category_products_endpoint_aggregates_children_and_returns_tabs(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/ice-cream-machine/products")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["slug"], "ice-cream-machine")
        self.assertIsNone(payload["active_subcategory_slug"])
        self.assertEqual(len(payload["subcategory_tabs"]), 2)
        self.assertEqual({tab["slug"] for tab in payload["subcategory_tabs"]}, {"gelato-batch-freezer", "roll-ice-cream-machine"})
        self.assertEqual(payload["pagination"]["total_items"], 3)
        self.assertEqual(
            {item["subcategory_slug"] for item in payload["items"]},
            {"ice-cream-machine", "gelato-batch-freezer", "roll-ice-cream-machine"},
        )
        self.assertIn("subcategory_name", payload["items"][0])
        self.assertEqual(payload["operational_fit_title"], "Operational Fit")
        self.assertEqual(payload["buyer_review_focus_title"], "Buyer Review Focus")
        self.assertEqual(payload["sourcing_faq_title"], "Sourcing FAQ")
        self.assertEqual(
            [item["title"] for item in payload["operational_fit_items"]],
            ["Application Matching", "Utility Planning"],
        )
        self.assertEqual(
            [item["section_code"] for item in payload["operational_fit_items"]],
            ["operational_fit", "operational_fit"],
        )
        self.assertEqual(
            [item["title"] for item in payload["buyer_review_focus_items"]],
            ["Capacity vs. Actual Demand", "Utilities and Footprint"],
        )
        self.assertEqual(
            [item["sort_order"] for item in payload["buyer_review_focus_items"]],
            [10, 20],
        )
        self.assertEqual(
            [item["question"] for item in payload["sourcing_faq_items"]],
            [
                "How should buyers compare ice cream machine models?",
                "What details should be prepared before requesting a quote?",
            ],
        )
        self.assertIsNotNone(payload["comparison_overview"])
        self.assertEqual(
            payload["comparison_overview"]["title"],
            "Ice Cream Machine Route Comparison",
        )
        self.assertEqual(
            [item["subject_key"] for item in payload["comparison_overview"]["subjects"]],
            [
                "soft_ice_cream_machine",
                "gelato_batch_freezer",
                "roll_ice_cream_machine",
            ],
        )
        self.assertEqual(
            [row["row_key"] for row in payload["comparison_overview"]["rows"]],
            ["best_fit_service_format", "production_rhythm"],
        )

    def test_parent_category_products_endpoint_filters_by_subcategory(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/categories/ice-cream-machine/products?subcategory_slug=gelato-batch-freezer"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["active_subcategory_slug"], "gelato-batch-freezer")
        self.assertEqual(payload["pagination"]["total_items"], 1)
        self.assertEqual(payload["items"][0]["slug"], self.gelato_product.slug)
        self.assertEqual(payload["items"][0]["subcategory_slug"], "gelato-batch-freezer")
        self.assertEqual(payload["operational_fit_title"], "Gelato Operational Fit")
        self.assertEqual(payload["buyer_review_focus_title"], "Gelato Buyer Review Focus")
        self.assertEqual(payload["sourcing_faq_title"], "Gelato Sourcing FAQ")
        self.assertEqual(
            [item["title"] for item in payload["operational_fit_items"]],
            ["Batch Capacity Fit", "Kitchen Workflow"],
        )
        self.assertEqual(
            [item["title"] for item in payload["buyer_review_focus_items"]],
            ["Freeze Curve Review", "Utilities by Batch"],
        )
        self.assertEqual(
            [item["question"] for item in payload["sourcing_faq_items"]],
            [
                "What should buyers check before choosing a gelato batch freezer?",
                "Why does kitchen workflow matter for gelato batch freezers?",
            ],
        )
        self.assertIsNone(payload["comparison_overview"])

    def test_parent_category_products_endpoint_falls_back_to_parent_operational_content_when_child_has_none(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/categories/ice-cream-machine/products?subcategory_slug=roll-ice-cream-machine"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["active_subcategory_slug"], "roll-ice-cream-machine")
        self.assertEqual(payload["operational_fit_title"], "Operational Fit")
        self.assertEqual(payload["buyer_review_focus_title"], "Buyer Review Focus")
        self.assertEqual(payload["sourcing_faq_title"], "Sourcing FAQ")
        self.assertEqual(
            [item["title"] for item in payload["operational_fit_items"]],
            ["Application Matching", "Utility Planning"],
        )
        self.assertEqual(
            [item["title"] for item in payload["buyer_review_focus_items"]],
            ["Capacity vs. Actual Demand", "Utilities and Footprint"],
        )
        self.assertEqual(
            [item["question"] for item in payload["sourcing_faq_items"]],
            [
                "How should buyers compare ice cream machine models?",
                "What details should be prepared before requesting a quote?",
            ],
        )
        self.assertIsNone(payload["comparison_overview"])

    def test_single_child_parent_products_endpoint_hides_tabs_but_aggregates_child(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/home-use-slush-machine/products")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["slug"], "home-use-slush-machine")
        self.assertEqual(payload["active_subcategory_slug"], "2l-slush-machine")
        self.assertEqual(payload["subcategory_tabs"], [])
        self.assertEqual(payload["pagination"]["total_items"], 1)
        self.assertEqual(payload["items"][0]["slug"], self.single_child_product.slug)
        self.assertEqual(payload["operational_fit_title"], "Home Slush Operational Fit")
        self.assertEqual(payload["buyer_review_focus_title"], "Home Slush Buyer Review Focus")
        self.assertEqual(payload["sourcing_faq_title"], "Home Slush Sourcing FAQ")
        self.assertEqual(
            [item["title"] for item in payload["operational_fit_items"]],
            ["Parent Footprint Fit"],
        )
        self.assertEqual(
            [item["title"] for item in payload["buyer_review_focus_items"]],
            ["Parent Throughput Framing"],
        )
        self.assertEqual(
            [item["question"] for item in payload["sourcing_faq_items"]],
            [
                "How should buyers compare home use slush machine models?",
                "What should buyers confirm before requesting a home use slush machine quote?",
            ],
        )
        self.assertIsNone(payload["comparison_overview"])

    def test_leaf_category_products_endpoint_rejects_subcategory_filter(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/categories/soft-ice-cream-machine/products?subcategory_slug=gelato-batch-freezer"
        )

        self.assertEqual(response.status_code, 400)

    def test_parent_category_products_endpoint_rejects_unknown_subcategory_filter(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/categories/ice-cream-machine/products?subcategory_slug=not-a-real-child"
        )

        self.assertEqual(response.status_code, 400)

    def test_product_endpoint_returns_related_and_specs(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/products/soft-ice-cream-machine/icm-t838-twin-twist-soft-serve-machine"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["model_code"], "ICM-T838")
        self.assertEqual(payload["hero_eyebrow"], "Soft Ice Cream Machines")
        self.assertGreaterEqual(len(payload["quick_facts"]), 1)
        self.assertGreaterEqual(len(payload["related_resources"]), 1)
        self.assertGreaterEqual(len(payload["related_products"]), 1)
        self.assertIn("group_kind_code", payload["spec_groups"][0])
        self.assertIn("image_url", payload["related_products"][0])
        self.assertIn("eyebrow", payload["related_products"][0])
        self.assertEqual(payload["related_resources"][0]["eyebrow"], "RESOURCE CENTER")
        self.assertTrue(payload["related_resources"][0]["image_url"])
        self.assertEqual(len(payload["faq_items"]), 3)
        self.assertTrue(
            any(faq["question"] == "Does the ICM-T838 support OEM branding?" for faq in payload["faq_items"])
        )

    def test_product_detail_service_only_exposes_first_three_product_faqs(self) -> None:
        FAQItem.objects.create(
            content_type=ContentType.objects.get_for_model(self.primary_product, for_concrete_model=False),
            object_id=self.primary_product.id,
            question="Should I over-size this model for every site?",
            answer="No. Oversizing usually creates a worse fit if daily demand, utilities, and floor plan do not require it.",
            is_featured=True,
            sort_order=40,
        )

        payload = get_product_detail(
            "soft-ice-cream-machine",
            "icm-t838-twin-twist-soft-serve-machine",
        )

        self.assertEqual(len(payload.faq_items), 3)
        self.assertNotIn(
            "Should I over-size this model for every site?",
            [faq.question for faq in payload.faq_items],
        )

    def test_import_product_faqs_command_replaces_product_bound_faqs(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ProductTechnicalFaqs"
        sheet.append(
            [
                "Category",
                "Subcategory",
                "Product Name",
                "Product URL",
                "Source Product URL",
                "Model Code",
                "Series Label",
                "Primary Query",
                "Resource Topic ID",
                "Resource Title",
                "FAQ Family",
                "FAQ Slot",
                "Question",
                "Answer",
                "Sort Order",
                "Question Word Count",
                "Answer Word Count",
                "Key Signals",
                "QA Status",
                "Notes",
            ]
        )
        rows = [
            (1, 10, "Is this model suitable for a compact dessert counter?"),
            (2, 20, "How should I size this model for peak-hour demand?"),
            (3, 30, "What should I confirm before requesting a quote for this model?"),
        ]
        for slot, sort_order, question in rows:
            sheet.append(
                [
                    self.primary_product.category.name,
                    self.primary_product.category.name,
                    self.primary_product.name,
                    self.primary_product.url_path,
                    self.primary_product.source_url,
                    self.primary_product.model_code,
                    self.primary_product.series_label,
                    self.primary_product.primary_query,
                    "RA-001",
                    "How to Choose a Commercial Soft Serve Ice Cream Machine",
                    "Soft Serve / Frozen Yogurt",
                    slot,
                    question,
                    "Use the published configuration as a first filter, then confirm output, voltage, and available space.",
                    sort_order,
                    10,
                    16,
                    "output: 36-40 L/H",
                    "generated",
                    "",
                ]
            )

        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        handle.close()
        workbook.save(handle.name)
        workbook.close()

        try:
            call_command("import_product_faqs", excel=handle.name)
        finally:
            Path(handle.name).unlink(missing_ok=True)

        product_ct = ContentType.objects.get_for_model(self.primary_product, for_concrete_model=False)
        imported_faqs = list(
            FAQItem.objects.filter(content_type=product_ct, object_id=self.primary_product.id).order_by("sort_order", "id")
        )
        self.assertEqual(len(imported_faqs), 3)
        self.assertEqual(
            [faq.question for faq in imported_faqs],
            [
                "Is this model suitable for a compact dessert counter?",
                "How should I size this model for peak-hour demand?",
                "What should I confirm before requesting a quote for this model?",
            ],
        )

    def test_product_detail_service_stays_within_expected_query_budget(self) -> None:
        ContentType.objects.get_for_model(Product, for_concrete_model=False)

        with CaptureQueriesContext(connection) as ctx:
            payload = get_product_detail(
                "soft-ice-cream-machine",
                "icm-t838-twin-twist-soft-serve-machine",
            )

        self.assertEqual(payload.slug, "icm-t838-twin-twist-soft-serve-machine")
        self.assertLessEqual(len(ctx.captured_queries), 12)


class ImportCategoryOperationalItemsCommandTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.parent_category = ProductCategory.objects.create(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            h1="Ice Cream Machine",
            seo_title="Ice Cream Machine | PRO-TAYLOR",
            meta_description="Ice Cream Machine meta description",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        cls.child_category = ProductCategory.objects.create(
            name="Gelato Batch Freezer",
            slug="gelato-batch-freezer",
            url_path="/products/gelato-batch-freezer/",
            h1="Gelato Batch Freezer",
            seo_title="Gelato Batch Freezer | PRO-TAYLOR",
            meta_description="Gelato Batch Freezer meta description",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.parent_category,
        )

    def _build_operational_seed_workbook(
        self,
        *,
        groups: list[dict[str, object]],
    ) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Operational Items Seed"
        sheet.append(
            [
                "source_category",
                "source_subcategory",
                "target_category_name",
                "target_scope",
                "operational_fit_title",
                "buyer_review_focus_title",
                "section_code",
                "sort_order",
                "item_title",
                "item_body",
                "icon",
                "is_active",
                "source_fields_used",
                "primary_source_url",
                "evidence_note",
                "confidence",
            ]
        )
        for group in groups:
            for row in group["rows"]:
                sheet.append(
                    [
                        group["source_category"],
                        group["source_subcategory"],
                        group["target_category_name"],
                        group["target_scope"],
                        group["operational_fit_title"],
                        group["buyer_review_focus_title"],
                        row["section_code"],
                        row["sort_order"],
                        row["item_title"],
                        row["item_body"],
                        row["icon"],
                        row["is_active"],
                        "summary;selection_guide",
                        "https://example.com/source",
                        "Evidence note",
                        0.95,
                    ]
                )

        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        return handle.name

    def test_import_category_operational_items_imports_parent_all_seed(self) -> None:
        workbook_path = self._build_operational_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "── All ──",
                    "target_category_name": "Ice Cream Machine",
                    "target_scope": "parent_all",
                    "operational_fit_title": "Operational Fit",
                    "buyer_review_focus_title": "Buyer Review Focus",
                    "rows": [
                        {
                            "section_code": "operational_fit",
                            "sort_order": 1,
                            "item_title": "Application Matching",
                            "item_body": "Check service style and expected daily throughput before narrowing machine families.",
                            "icon": "storefront",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 2,
                            "item_title": "Utility Planning",
                            "item_body": "Confirm voltage, cooling method, and placement limits before approving the shortlist.",
                            "icon": "bolt",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 3,
                            "item_title": "Quote Readiness",
                            "item_body": "Bring target output and site constraints into the inquiry so model confirmation moves faster.",
                            "icon": "fact_check",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 1,
                            "item_title": "Capacity vs Demand",
                            "item_body": "Match peak demand, recovery expectations, and menu mix before sizing up.",
                            "icon": "speed",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 2,
                            "item_title": "Utilities and Space",
                            "item_body": "Check electrical supply, airflow, and working space before locking the model.",
                            "icon": "home_work",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 3,
                            "item_title": "Procurement Inputs",
                            "item_body": "Align quantity, target market, and compliance requirements before asking for final pricing.",
                            "icon": "fact_check",
                            "is_active": True,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_operational_items", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        self.parent_category.refresh_from_db()
        self.assertEqual(self.parent_category.operational_fit_title, "Operational Fit")
        self.assertEqual(self.parent_category.buyer_review_focus_title, "Buyer Review Focus")

        parent_items = list(
            ProductCategoryOperationalItem.objects.filter(category=self.parent_category).order_by("section", "sort_order")
        )
        self.assertEqual(len(parent_items), 6)
        self.assertEqual(
            [(item.section_code, item.sort_order, item.title) for item in parent_items],
            [
                ("operational_fit", 1, "Application Matching"),
                ("operational_fit", 2, "Utility Planning"),
                ("operational_fit", 3, "Quote Readiness"),
                ("buyer_review_focus", 1, "Capacity vs Demand"),
                ("buyer_review_focus", 2, "Utilities and Space"),
                ("buyer_review_focus", 3, "Procurement Inputs"),
            ],
        )

    def test_import_category_operational_items_replaces_existing_rows_on_rerun(self) -> None:
        ProductCategoryOperationalItem.objects.create(
            category=self.parent_category,
            section=ProductCategoryOperationalItem.Section.OPERATIONAL_FIT,
            title="Legacy Row",
            body="Legacy body",
            icon="storefront",
            sort_order=1,
        )
        self.parent_category.operational_fit_title = "Legacy Operational"
        self.parent_category.buyer_review_focus_title = "Legacy Buyer Focus"
        self.parent_category.save(update_fields=("operational_fit_title", "buyer_review_focus_title"))

        workbook_path = self._build_operational_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "── All ──",
                    "target_category_name": "Ice Cream Machine",
                    "target_scope": "parent_all",
                    "operational_fit_title": "New Operational Fit",
                    "buyer_review_focus_title": "New Buyer Review Focus",
                    "rows": [
                        {
                            "section_code": "operational_fit",
                            "sort_order": 1,
                            "item_title": "Menu Match",
                            "item_body": "Choose machine families around product format and serving rhythm.",
                            "icon": "storefront",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 2,
                            "item_title": "Site Utilities",
                            "item_body": "Check voltage, cooling, and placement before approving a final shortlist.",
                            "icon": "bolt",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 3,
                            "item_title": "Inquiry Inputs",
                            "item_body": "Prepare output target and site notes before asking suppliers to confirm fit.",
                            "icon": "fact_check",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 1,
                            "item_title": "Peak Throughput",
                            "item_body": "Size for peak demand instead of only average daily volume.",
                            "icon": "speed",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 2,
                            "item_title": "Footprint Review",
                            "item_body": "Confirm working room and ventilation before model lock.",
                            "icon": "home_work",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 3,
                            "item_title": "Procurement Ready",
                            "item_body": "Align market, quantity, and compliance details before final quotation.",
                            "icon": "fact_check",
                            "is_active": False,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_operational_items", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        self.parent_category.refresh_from_db()
        self.assertEqual(self.parent_category.operational_fit_title, "New Operational Fit")
        self.assertEqual(self.parent_category.buyer_review_focus_title, "New Buyer Review Focus")
        parent_items = list(
            ProductCategoryOperationalItem.objects.filter(category=self.parent_category).order_by("section", "sort_order")
        )
        self.assertEqual(len(parent_items), 6)
        self.assertNotIn("Legacy Row", [item.title for item in parent_items])
        self.assertFalse(parent_items[-1].is_active)

    def test_import_category_operational_items_imports_subcategory_seed_to_child_category(self) -> None:
        workbook_path = self._build_operational_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "Gelato Batch Freezer",
                    "target_category_name": "Gelato Batch Freezer",
                    "target_scope": "subcategory",
                    "operational_fit_title": "Operational Fit",
                    "buyer_review_focus_title": "Buyer Review Focus",
                    "rows": [
                        {
                            "section_code": "operational_fit",
                            "sort_order": 1,
                            "item_title": "Batch Size Match",
                            "item_body": "Choose bowl size around real batch cadence, not only headline output.",
                            "icon": "storefront",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 2,
                            "item_title": "Cooling Setup",
                            "item_body": "Confirm power and condenser style before shortlisting gelato freezers.",
                            "icon": "bolt",
                            "is_active": True,
                        },
                        {
                            "section_code": "operational_fit",
                            "sort_order": 3,
                            "item_title": "Operator Skill",
                            "item_body": "Match controls and workflow to the team that will run daily batches.",
                            "icon": "fact_check",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 1,
                            "item_title": "Output Reality",
                            "item_body": "Check production per batch and pull-down time before promising menu volume.",
                            "icon": "speed",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 2,
                            "item_title": "Kitchen Footprint",
                            "item_body": "Review working clearance and cleaning access around the freezer body.",
                            "icon": "home_work",
                            "is_active": True,
                        },
                        {
                            "section_code": "buyer_review_focus",
                            "sort_order": 3,
                            "item_title": "Service Readiness",
                            "item_body": "Confirm spare parts and support response before committing a production line.",
                            "icon": "fact_check",
                            "is_active": True,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_operational_items", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        child_items = list(
            ProductCategoryOperationalItem.objects.filter(category=self.child_category).order_by("section", "sort_order")
        )
        self.assertEqual(len(child_items), 6)
        self.assertEqual(child_items[0].title, "Batch Size Match")
        self.assertEqual(child_items[-1].title, "Service Readiness")
        self.parent_category.refresh_from_db()
        self.child_category.refresh_from_db()
        self.assertEqual(self.child_category.operational_fit_title, "Operational Fit")
        self.assertEqual(self.parent_category.operational_fit_title, "")


class ImportCategoryFaqsCommandTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.parent_category = ProductCategory.objects.create(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            h1="Ice Cream Machine",
            seo_title="Ice Cream Machine | PRO-TAYLOR",
            meta_description="Ice Cream Machine meta description",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        cls.child_category = ProductCategory.objects.create(
            name="Gelato Batch Freezer",
            slug="gelato-batch-freezer",
            url_path="/products/gelato-batch-freezer/",
            h1="Gelato Batch Freezer",
            seo_title="Gelato Batch Freezer | PRO-TAYLOR",
            meta_description="Gelato Batch Freezer meta description",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.parent_category,
        )

    def _build_category_faq_seed_workbook(
        self,
        *,
        groups: list[dict[str, object]],
    ) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Category Sourcing FAQ Seed"
        sheet.append(
            [
                "source_category",
                "source_subcategory",
                "target_scope",
                "target_category_slug",
                "target_category_name",
                "placement_code",
                "sourcing_faq_title",
                "sort_order",
                "question",
                "answer",
                "is_active",
                "source_fields_used",
                "primary_source_url",
                "secondary_source_url",
                "evidence_note",
                "confidence",
            ]
        )
        for group in groups:
            for row in group["rows"]:
                sheet.append(
                    [
                        group["source_category"],
                        group["source_subcategory"],
                        group["target_scope"],
                        group["target_category_slug"],
                        group["target_category_name"],
                        group["placement_code"],
                        group["sourcing_faq_title"],
                        row["sort_order"],
                        row["question"],
                        row["answer"],
                        row["is_active"],
                        "summary;selection_guide",
                        "https://example.com/primary",
                        "https://example.com/secondary",
                        "Evidence note",
                        0.95,
                    ]
                )

        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        return handle.name

    def test_import_category_faqs_imports_parent_all_seed(self) -> None:
        workbook_path = self._build_category_faq_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "── All ──",
                    "target_scope": "parent_all",
                    "target_category_slug": "ice-cream-machine",
                    "target_category_name": "Ice Cream Machine",
                    "placement_code": "plp_sourcing",
                    "sourcing_faq_title": "Sourcing FAQ",
                    "rows": [
                        {
                            "sort_order": 1,
                            "question": "How should buyers split soft serve, gelato, and rolled ice cream paths?",
                            "answer": "Use product style, service rhythm, and operator workflow to separate soft serve, batch freezing, and roll-plate equipment before shortlisting.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 2,
                            "question": "What details should be prepared before asking for an ice cream machine quote?",
                            "answer": "Prepare target format, peak output, voltage, cooling style, and counter or kitchen limits so suppliers can confirm the correct machine family.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 3,
                            "question": "Which site constraints usually change the recommended ice cream machine family?",
                            "answer": "Rush pattern, working space, ventilation, and cleaning discipline can move a project toward soft serve, gelato batch, or rolled ice cream equipment.",
                            "is_active": True,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_faqs", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        self.parent_category.refresh_from_db()
        self.assertEqual(self.parent_category.sourcing_faq_title, "Sourcing FAQ")
        parent_items = list(
            ProductCategoryFaqItem.objects.filter(category=self.parent_category).order_by("sort_order")
        )
        self.assertEqual(len(parent_items), 3)
        self.assertEqual(
            [(item.placement_code, item.sort_order) for item in parent_items],
            [("plp_sourcing", 1), ("plp_sourcing", 2), ("plp_sourcing", 3)],
        )
        self.assertEqual(
            [item.question for item in parent_items],
            [
                "How should buyers split soft serve, gelato, and rolled ice cream paths?",
                "What details should be prepared before asking for an ice cream machine quote?",
                "Which site constraints usually change the recommended ice cream machine family?",
            ],
        )

    def test_import_category_faqs_replaces_existing_rows_on_rerun(self) -> None:
        ProductCategoryFaqItem.objects.create(
            category=self.parent_category,
            placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
            question="Legacy question",
            answer="Legacy answer",
            sort_order=1,
        )
        self.parent_category.sourcing_faq_title = "Legacy FAQ"
        self.parent_category.save(update_fields=("sourcing_faq_title",))

        workbook_path = self._build_category_faq_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "── All ──",
                    "target_scope": "parent_all",
                    "target_category_slug": "ice-cream-machine",
                    "target_category_name": "Ice Cream Machine",
                    "placement_code": "plp_sourcing",
                    "sourcing_faq_title": "New Sourcing FAQ",
                    "rows": [
                        {
                            "sort_order": 1,
                            "question": "How do buyers know when a family-level comparison is enough?",
                            "answer": "Start with service model and product format, then confirm output and utility limits before asking suppliers to validate a final shortlist.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 2,
                            "question": "What quote inputs matter most for a category-level ice cream machine inquiry?",
                            "answer": "Peak servings, preferred format, voltage, cooling style, and working-space limits usually determine which machine family survives the first review.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 3,
                            "question": "What usually pushes a buyer into the wrong ice cream machine family?",
                            "answer": "Ignoring rush length, ventilation, cleaning windows, or operator workflow often sends the inquiry toward the wrong machine class.",
                            "is_active": False,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_faqs", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        self.parent_category.refresh_from_db()
        self.assertEqual(self.parent_category.sourcing_faq_title, "New Sourcing FAQ")
        parent_items = list(
            ProductCategoryFaqItem.objects.filter(category=self.parent_category).order_by("sort_order")
        )
        self.assertEqual(len(parent_items), 3)
        self.assertNotIn("Legacy question", [item.question for item in parent_items])
        self.assertFalse(parent_items[-1].is_active)

    def test_import_category_faqs_imports_subcategory_seed_to_child_category(self) -> None:
        workbook_path = self._build_category_faq_seed_workbook(
            groups=[
                {
                    "source_category": "Ice Cream Machine",
                    "source_subcategory": "Gelato Batch Freezer",
                    "target_scope": "subcategory",
                    "target_category_slug": "gelato-batch-freezer",
                    "target_category_name": "Gelato Batch Freezer",
                    "placement_code": "plp_sourcing",
                    "sourcing_faq_title": "Gelato Sourcing FAQ",
                    "rows": [
                        {
                            "sort_order": 1,
                            "question": "When is a gelato batch freezer the right subcategory instead of soft serve equipment?",
                            "answer": "Choose a gelato batch freezer when the operator needs discrete batch production, texture control, and pan handoff instead of continuous draw dispensing.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 2,
                            "question": "What should buyers prepare before asking for a gelato batch freezer quote?",
                            "answer": "Prepare batch size, cycle target, kitchen power, condenser limits, and pan-transfer workflow so suppliers can confirm the correct freezer size.",
                            "is_active": True,
                        },
                        {
                            "sort_order": 3,
                            "question": "Which constraints most often break fit for a gelato batch freezer project?",
                            "answer": "Tight transfer space, weak ventilation, unrealistic cycle expectations, and limited operator discipline usually change the final freezer recommendation.",
                            "is_active": True,
                        },
                    ],
                }
            ]
        )

        try:
            call_command("import_category_faqs", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        child_items = list(
            ProductCategoryFaqItem.objects.filter(category=self.child_category).order_by("sort_order")
        )
        self.assertEqual(len(child_items), 3)
        self.assertEqual(child_items[0].question, "When is a gelato batch freezer the right subcategory instead of soft serve equipment?")
        self.assertEqual(child_items[-1].sort_order, 3)

        self.parent_category.refresh_from_db()
        self.child_category.refresh_from_db()
        self.assertEqual(self.child_category.sourcing_faq_title, "Gelato Sourcing FAQ")
        self.assertEqual(self.parent_category.sourcing_faq_title, "")


class CategoryGuideApiTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.category = ProductCategory.objects.create(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            h1="Ice Cream Machine",
            summary="Ice Cream Machine summary",
            lead_text="Ice Cream Machine lead text",
            seo_title="Ice Cream Machine | PRO-TAYLOR",
            meta_description="Ice Cream Machine meta description",
            primary_query="ice cream machine",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        cls.category_without_guide = ProductCategory.objects.create(
            name="Soft Ice Cream Machine",
            slug="soft-ice-cream-machine",
            url_path="/products/soft-ice-cream-machine/",
            h1="Soft Ice Cream Machine",
            seo_title="Soft Ice Cream Machine | PRO-TAYLOR",
            meta_description="Soft Ice Cream Machine meta description",
            primary_query="soft ice cream machine",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        cls.child_category_with_guide = ProductCategory.objects.create(
            name="Gelato Batch Freezer",
            slug="gelato-batch-freezer",
            url_path="/products/gelato-batch-freezer/",
            h1="Gelato Batch Freezer",
            seo_title="Gelato Batch Freezer | PRO-TAYLOR",
            meta_description="Gelato Batch Freezer meta description",
            primary_query="gelato batch freezer",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.category,
        )
        cls.inactive_guide_category = ProductCategory.objects.create(
            name="Blast Freezer",
            slug="blast-freezer",
            url_path="/products/blast-freezer/",
            h1="Blast Freezer",
            seo_title="Blast Freezer | PRO-TAYLOR",
            meta_description="Blast Freezer meta description",
            primary_query="blast freezer",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        image = MediaAsset.objects.create(
            title="Ice Cream Machine Guide Hero",
            asset_kind=MediaAsset.AssetKind.IMAGE,
            file_url="https://cdn.example.com/protaylor/ice-cream-machine-guide.jpg",
            alt_text="Ice cream machine guide hero",
            mime_type="image/jpeg",
        )
        guide = ProductCategoryGuide.objects.create(
            category=cls.category,
            hero_eyebrow="Editorial Buying Guide",
            hero_title="A Professional Guide to Ice Cream Machine",
            answer_summary="Use this guide to split production route decisions before comparing individual machines.",
            hero_primary_cta_label="View Models in This Category",
            hero_primary_cta_href="/products/ice-cream-machine/",
            hero_secondary_cta_label="Request Quote",
            hero_secondary_cta_href="/contact/",
            hero_image=image,
            hero_image_alt="Ice cream machine guide hero override",
            hero_note_title="Category Review",
            hero_note_copy="Start with operating fit before comparing individual models.",
            hero_note_quote='"Start with route fit before model fit."',
            hero_note_attribution="PRO-TAYLOR Category Review",
            definition_title="Defining Your Review Scope",
            definition_copy="First paragraph for route framing.\n\nSecond paragraph for shortlist logic.",
            contexts_title="Operational Contexts",
            matrix_title="Buyer Review Focus",
            matrix_eyebrow="Factors That Determine Shortlist Quality",
            paths_title="Recommended Paths",
            paths_eyebrow="Tailored for your range",
            paths_mode=ProductCategoryGuide.PathMode.ROUTE_GUIDANCE,
            trust_title="Global Engineering Standards",
            trust_copy="Confirm voltage, compliance, packing, and service support before order lock.",
            trust_mode=ProductCategoryGuide.TrustMode.CATEGORY_CURATED,
            faq_title="Category FAQ",
            resources_title="Related Resources",
            resources_mode=ProductCategoryGuide.ResourcesMode.CATEGORY_CURATED,
            cta_title="Ready to Configure Your System?",
            cta_copy="Browse models or request technical confirmation with your target route.",
            cta_mode=ProductCategoryGuide.CtaMode.LISTING_FIRST,
            cta_primary_label="Browse Current Models",
            cta_primary_href="/products/ice-cream-machine/",
            cta_secondary_label="Request Quote",
            cta_secondary_href="/contact/",
        )
        ProductCategoryGuide.objects.create(
            category=cls.child_category_with_guide,
            hero_title="Child guides should not be listed for sitemap.",
        )
        ProductCategoryGuide.objects.create(
            category=cls.inactive_guide_category,
            is_active=False,
            hero_title="Inactive guides should not be listed for sitemap.",
        )
        ProductCategoryGuideItem.objects.bulk_create(
            [
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.DEFINITION_CARD,
                    item_key="best_for",
                    eyebrow="best_for",
                    title="Route-first buyers",
                    body="Best for teams comparing soft serve, batch, and roll ice cream operating routes.",
                    icon="storefront",
                    sort_order=10,
                ),
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.OPERATIONAL_CONTEXT,
                    item_key="dessert_counter",
                    title="Dessert Counter Service",
                    body="Use when counter rhythm and visible service format shape the machine choice.",
                    asset=image,
                    asset_alt="Dessert counter guide context",
                    sort_order=10,
                ),
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.DECISION_FACTOR,
                    item_key="service_format",
                    title="Service Format",
                    body="Decide direct-draw, batch production, or made-to-order pan service first.",
                    icon="tune",
                    sort_order=10,
                ),
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.PATH,
                    item_key="soft_serve_route",
                    eyebrow="01",
                    title="Start with soft serve",
                    body="Choose this path when continuous direct-draw service is the main need.",
                    supporting_points="Open the soft serve route\nCompare model cards\nConfirm quote requirements",
                    href="/products/soft-ice-cream-machine/",
                    sort_order=10,
                ),
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.TRUST_METRIC,
                    item_key="ce_compliance",
                    eyebrow="CE",
                    title="Compliance-ready exports",
                    sort_order=10,
                ),
                ProductCategoryGuideItem(
                    guide=guide,
                    section=ProductCategoryGuideItem.Section.RELATED_RESOURCE,
                    item_key="soft_serve_buying_guide",
                    eyebrow="Technical Guide",
                    title="Soft Serve Buying Guide",
                    href="/resources/soft-serve-buying-guide/",
                    sort_order=10,
                ),
            ]
        )
        ProductCategoryFaqItem.objects.bulk_create(
            [
                ProductCategoryFaqItem(
                    category=cls.category,
                    placement=ProductCategoryFaqItem.Placement.PLP_SOURCING,
                    question="PLP FAQ should not appear in guide.",
                    answer="This item belongs to the product listing page.",
                    sort_order=10,
                ),
                ProductCategoryFaqItem(
                    category=cls.category,
                    placement=ProductCategoryFaqItem.Placement.GUIDE_FAQ,
                    question="Why open the guide before comparing models?",
                    answer="The guide separates route fit before model-level output and configuration review.",
                    sort_order=10,
                ),
            ]
        )

    def setUp(self) -> None:
        self.client = build_api_client()

    def test_category_guide_endpoint_returns_structured_guide_contract(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/ice-cream-machine/guide")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        guide = payload["guide"]
        self.assertEqual(payload["slug"], "ice-cream-machine")
        self.assertEqual(guide["hero_title"], "A Professional Guide to Ice Cream Machine")
        self.assertEqual(guide["hero_image_url"], "https://cdn.example.com/protaylor/ice-cream-machine-guide.jpg")
        self.assertEqual(
            guide["definition_paragraphs"],
            ["First paragraph for route framing.", "Second paragraph for shortlist logic."],
        )
        self.assertEqual(guide["definition_cards"][0]["role_code"], "best_for")
        self.assertEqual(guide["contexts"][0]["image_alt"], "Dessert counter guide context")
        self.assertEqual(guide["decision_factors"][0]["item_key"], "service_format")
        self.assertEqual(guide["paths_mode_code"], "route_guidance")
        self.assertEqual(
            guide["paths"][0]["bullets"],
            ["Open the soft serve route", "Compare model cards", "Confirm quote requirements"],
        )
        self.assertEqual(guide["standards_mode_code"], "category_curated")
        self.assertEqual(guide["standards_stats"][0]["value"], "CE")
        self.assertEqual(
            [item["question"] for item in guide["faqs"]],
            ["Why open the guide before comparing models?"],
        )
        self.assertEqual(guide["resources"][0]["label"], "Technical Guide")
        self.assertEqual(guide["cta_mode_code"], "listing_first")

    def test_category_guide_endpoint_returns_404_without_active_top_level_guide(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/soft-ice-cream-machine/guide")

        self.assertEqual(response.status_code, 404)

    def test_category_guide_paths_endpoint_returns_only_active_top_level_guides(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/guide/paths")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item["slug"] for item in payload], ["ice-cream-machine"])
        self.assertEqual(payload[0]["url_path"], "/products/ice-cream-machine/")
        self.assertIn("last_modified", payload[0])


class CategoryComparisonOverviewApiTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.parent_category = ProductCategory.objects.create(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            h1="Ice Cream Machine",
            summary="Ice Cream Machine summary",
            lead_text="Ice Cream Machine lead text",
            seo_title="Ice Cream Machine | PRO-TAYLOR",
            meta_description="Ice Cream Machine meta description",
            primary_query="ice cream machine",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        cls.soft_child = ProductCategory.objects.create(
            name="Soft Ice Cream Machine",
            slug="soft-ice-cream-machine",
            url_path="/products/soft-ice-cream-machine/",
            h1="Soft Ice Cream Machine",
            summary="Soft Ice Cream Machine summary",
            lead_text="Soft Ice Cream Machine lead text",
            seo_title="Soft Ice Cream Machine | PRO-TAYLOR",
            meta_description="Soft Ice Cream Machine meta description",
            primary_query="soft ice cream machine",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.parent_category,
        )
        cls.gelato_child = ProductCategory.objects.create(
            name="Gelato Batch Freezer",
            slug="gelato-batch-freezer",
            url_path="/products/gelato-batch-freezer/",
            h1="Gelato Batch Freezer",
            summary="Gelato Batch Freezer summary",
            lead_text="Gelato Batch Freezer lead text",
            seo_title="Gelato Batch Freezer | PRO-TAYLOR",
            meta_description="Gelato Batch Freezer meta description",
            primary_query="gelato batch freezer",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.parent_category,
        )
        cls.roll_child = ProductCategory.objects.create(
            name="Roll Ice Cream Machine",
            slug="roll-ice-cream-machine",
            url_path="/products/roll-ice-cream-machine/",
            h1="Roll Ice Cream Machine",
            summary="Roll Ice Cream Machine summary",
            lead_text="Roll Ice Cream Machine lead text",
            seo_title="Roll Ice Cream Machine | PRO-TAYLOR",
            meta_description="Roll Ice Cream Machine meta description",
            primary_query="roll ice cream machine",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
            parent=cls.parent_category,
        )

        for category, slug, name in [
            (cls.parent_category, "ice-cream-machine-direct-model", "Ice Cream Machine Direct Model"),
            (cls.gelato_child, "gelato-test-model", "Gelato Test Model"),
        ]:
            Product.objects.create(
                category=category,
                slug=slug,
                url_path=f"/products/{category.slug}/{slug}/",
                name=name,
                model_code=slug.upper(),
                h1=name,
                summary=f"{name} summary",
                lead_text=f"{name} lead text",
                seo_title=f"{name} | PRO-TAYLOR",
                meta_description=f"{name} meta description",
                primary_query=name.lower(),
                status=Product.Status.PUBLISHED,
                index_mode=Product.IndexMode.INDEX,
                is_canonical=True,
            )

        overview = ProductCategoryComparisonOverview.objects.create(
            category=cls.parent_category,
            title="Ice Cream Machine Route Comparison",
            intro="Use this table to split routes before comparing specific models.",
            dimension_heading="Decision Dimension",
            subjects_json=[
                {
                    "subject_key": "soft_ice_cream_machine",
                    "label": "Soft Ice Cream Machine",
                    "route_category_slug": "soft-ice-cream-machine",
                    "sort_order": 1,
                },
                {
                    "subject_key": "gelato_batch_freezer",
                    "label": "Gelato Batch Freezer",
                    "route_category_slug": "gelato-batch-freezer",
                    "sort_order": 2,
                },
                {
                    "subject_key": "roll_ice_cream_machine",
                    "label": "Roll Ice Cream Machine",
                    "route_category_slug": "roll-ice-cream-machine",
                    "sort_order": 3,
                },
            ],
        )
        ProductCategoryComparisonRow.objects.bulk_create(
            [
                ProductCategoryComparisonRow(
                    overview=overview,
                    row_key="best_fit_service_format",
                    label="Best-Fit Service Format",
                    sort_order=10,
                    cells_json={
                        "soft_ice_cream_machine": "Direct-draw service from the machine during service.",
                        "gelato_batch_freezer": "Batch-freeze product, then extract it for serving.",
                        "roll_ice_cream_machine": "Freeze the mix on a pan, scrape it, and serve each order directly.",
                    },
                ),
                ProductCategoryComparisonRow(
                    overview=overview,
                    row_key="production_rhythm",
                    label="Production Rhythm",
                    sort_order=20,
                    cells_json={
                        "soft_ice_cream_machine": "Continuous dispensing during service.",
                        "gelato_batch_freezer": "Planned batch cycles with refill and restart between runs.",
                        "roll_ice_cream_machine": "One-pan or two-pan made-to-order preparation.",
                    },
                ),
            ]
        )

    def test_parent_category_products_endpoint_returns_comparison_overview(self) -> None:
        response = self.client.get("/api/v1/catalog/categories/ice-cream-machine/products")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIsNotNone(payload["comparison_overview"])
        self.assertEqual(payload["comparison_overview"]["title"], "Ice Cream Machine Route Comparison")
        self.assertEqual(
            [subject["subject_key"] for subject in payload["comparison_overview"]["subjects"]],
            [
                "soft_ice_cream_machine",
                "gelato_batch_freezer",
                "roll_ice_cream_machine",
            ],
        )
        self.assertEqual(
            [row["row_key"] for row in payload["comparison_overview"]["rows"]],
            ["best_fit_service_format", "production_rhythm"],
        )

    def test_subcategory_filtered_products_endpoint_hides_comparison_overview(self) -> None:
        response = self.client.get(
            "/api/v1/catalog/categories/ice-cream-machine/products?subcategory_slug=gelato-batch-freezer"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["active_subcategory_slug"], "gelato-batch-freezer")
        self.assertEqual(payload["pagination"]["total_items"], 1)
        self.assertIsNone(payload["comparison_overview"])


class ImportCategoryComparisonOverviewCommandTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.parent_category = ProductCategory.objects.create(
            name="Ice Cream Machine",
            slug="ice-cream-machine",
            url_path="/products/ice-cream-machine/",
            h1="Ice Cream Machine",
            seo_title="Ice Cream Machine | PRO-TAYLOR",
            meta_description="Ice Cream Machine meta description",
            status=ProductCategory.Status.PUBLISHED,
            index_mode=ProductCategory.IndexMode.INDEX,
        )
        for name, slug in [
            ("Soft Ice Cream Machine", "soft-ice-cream-machine"),
            ("Gelato Batch Freezer", "gelato-batch-freezer"),
            ("Roll Ice Cream Machine", "roll-ice-cream-machine"),
        ]:
            ProductCategory.objects.create(
                name=name,
                slug=slug,
                url_path=f"/products/{slug}/",
                h1=name,
                seo_title=f"{name} | PRO-TAYLOR",
                meta_description=f"{name} meta description",
                status=ProductCategory.Status.PUBLISHED,
                index_mode=ProductCategory.IndexMode.INDEX,
                parent=cls.parent_category,
            )

    def _build_category_comparison_workbook(
        self,
        *,
        subjects: list[dict[str, object]],
        rows: list[dict[str, object]],
    ) -> str:
        workbook = Workbook()
        workbook.active.title = "Workbook Meta"
        ordered_subjects = sorted(
            subjects,
            key=lambda subject: (int(subject["sort_order"]), str(subject["subject_key"])),
        )
        subject_slots = max(4, len(ordered_subjects))

        subject_sheet = workbook.create_sheet("Comparison Overview Subjects")
        subject_sheet.append(
            [
                "category_name",
                "category_slug",
                "subject_key",
                "sort_order",
                "route_category_name",
                "route_category_slug",
                "label_override",
                "status",
                "source_doc",
                "notes",
            ]
        )
        for subject in subjects:
            subject_sheet.append(
                [
                    subject["category_name"],
                    subject["category_slug"],
                    subject["subject_key"],
                    subject["sort_order"],
                    subject["route_category_name"],
                    subject["route_category_slug"],
                    subject.get("label_override", ""),
                    subject.get("status", "active"),
                    "docs/15-Comparison-Overview-通用后端变更方案-2表版.md",
                    "Test seed",
                ]
            )

        draft_sheet = workbook.create_sheet("Comparison Overview Draft")
        draft_headers = [
            "category_name",
            "category_slug",
            "module_type",
            "module_title",
            "module_intro",
            "dimension_heading",
        ]
        for slot_index in range(1, subject_slots + 1):
            draft_headers.extend(
                [
                    f"subject_{slot_index}_key",
                    f"subject_{slot_index}_label",
                ]
            )
        draft_headers.extend(
            [
                "row_key",
                "sort_order",
                "decision_dimension",
            ]
        )
        for slot_index in range(1, subject_slots + 1):
            draft_headers.extend(
                [
                    f"subject_{slot_index}_text",
                ]
            )
        for slot_index in range(1, subject_slots + 1):
            draft_headers.extend(
                [
                    f"subject_{slot_index}_evidence_strength",
                ]
            )
        draft_headers.extend(
            [
                "synthesis_flag",
                "primary_source_summary",
                "evidence_note",
                "status",
                "source_doc",
            ]
        )
        draft_sheet.append(draft_headers)
        for row in rows:
            row_values: list[object] = [
                row["category_name"],
                row["category_slug"],
                "comparison_overview",
                row["module_title"],
                row["module_intro"],
                "Decision Dimension",
            ]
            for slot_index in range(subject_slots):
                subject = ordered_subjects[slot_index] if slot_index < len(ordered_subjects) else None
                row_values.extend(
                    [
                        subject["subject_key"] if subject else "",
                        (subject.get("label_override") or subject["route_category_name"]) if subject else "",
                    ]
                )

            row_values.extend(
                [
                    row["row_key"],
                    row["sort_order"],
                    row["decision_dimension"],
                ]
            )

            for slot_index in range(subject_slots):
                subject = ordered_subjects[slot_index] if slot_index < len(ordered_subjects) else None
                row_values.append(row[subject["subject_key"]] if subject else "")

            default_evidence_by_slot = {1: "strong", 2: "strong", 3: "medium", 4: ""}
            for slot_index in range(subject_slots):
                subject = ordered_subjects[slot_index] if slot_index < len(ordered_subjects) else None
                row_values.append(
                    row.get(
                        f"subject_{slot_index + 1}_evidence_strength",
                        default_evidence_by_slot.get(slot_index + 1, "medium") if subject else "",
                    )
                )

            row_values.extend(
                [
                    "no",
                    "Primary source summary",
                    "Evidence note",
                    row.get("status", "finalized_for_staging"),
                    "docs/13-Ice-Cream-Machine-Draft-Revision.md",
                ]
            )
            draft_sheet.append(row_values)

        handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        return handle.name

    def test_import_category_comparison_overview_imports_parent_matrix(self) -> None:
        workbook_path = self._build_category_comparison_workbook(
            subjects=[
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "soft_ice_cream_machine",
                    "sort_order": 1,
                    "route_category_name": "Soft Ice Cream Machine",
                    "route_category_slug": "soft-ice-cream-machine",
                },
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "gelato_batch_freezer",
                    "sort_order": 2,
                    "route_category_name": "Gelato Batch Freezer",
                    "route_category_slug": "gelato-batch-freezer",
                },
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "roll_ice_cream_machine",
                    "sort_order": 3,
                    "route_category_name": "Roll Ice Cream Machine",
                    "route_category_slug": "roll-ice-cream-machine",
                },
            ],
            rows=[
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "module_title": "Ice Cream Machine Route Comparison",
                    "module_intro": "Use this table to split routes before comparing specific models.",
                    "row_key": "best_fit_service_format",
                    "sort_order": 10,
                    "decision_dimension": "Best-Fit Service Format",
                    "soft_ice_cream_machine": "Direct-draw service from the machine during service.",
                    "gelato_batch_freezer": "Batch-freeze product, then extract it for serving.",
                    "roll_ice_cream_machine": "Freeze the mix on a pan, scrape it, and serve each order directly.",
                },
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "module_title": "Ice Cream Machine Route Comparison",
                    "module_intro": "Use this table to split routes before comparing specific models.",
                    "row_key": "production_rhythm",
                    "sort_order": 20,
                    "decision_dimension": "Production Rhythm",
                    "soft_ice_cream_machine": "Continuous dispensing during service.",
                    "gelato_batch_freezer": "Planned batch cycles with refill and restart between runs.",
                    "roll_ice_cream_machine": "One-pan or two-pan made-to-order preparation.",
                },
            ],
        )

        try:
            call_command("import_category_comparison_overview", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        overview = ProductCategoryComparisonOverview.objects.get(category=self.parent_category)
        rows = list(overview.rows.order_by("sort_order"))
        self.assertEqual(overview.title, "Ice Cream Machine Route Comparison")
        self.assertEqual(overview.dimension_heading, "Decision Dimension")
        self.assertEqual(len(overview.subjects_json), 3)
        self.assertEqual(
            [subject["route_category_slug"] for subject in overview.subjects_json],
            ["soft-ice-cream-machine", "gelato-batch-freezer", "roll-ice-cream-machine"],
        )
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].row_key, "best_fit_service_format")
        self.assertEqual(
            rows[1].cells_json["gelato_batch_freezer"],
            "Planned batch cycles with refill and restart between runs.",
        )

    def test_import_category_comparison_overview_replaces_existing_rows_on_rerun(self) -> None:
        overview = ProductCategoryComparisonOverview.objects.create(
            category=self.parent_category,
            title="Legacy comparison",
            intro="Legacy intro",
            dimension_heading="Legacy heading",
            subjects_json=[
                {
                    "subject_key": "soft_ice_cream_machine",
                    "label": "Soft Ice Cream Machine",
                    "route_category_slug": "soft-ice-cream-machine",
                    "sort_order": 1,
                },
                {
                    "subject_key": "gelato_batch_freezer",
                    "label": "Gelato Batch Freezer",
                    "route_category_slug": "gelato-batch-freezer",
                    "sort_order": 2,
                },
            ],
        )
        ProductCategoryComparisonRow.objects.create(
            overview=overview,
            row_key="legacy_row",
            label="Legacy Row",
            sort_order=10,
            cells_json={
                "soft_ice_cream_machine": "Legacy soft",
                "gelato_batch_freezer": "Legacy gelato",
            },
        )

        workbook_path = self._build_category_comparison_workbook(
            subjects=[
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "soft_ice_cream_machine",
                    "sort_order": 1,
                    "route_category_name": "Soft Ice Cream Machine",
                    "route_category_slug": "soft-ice-cream-machine",
                },
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "gelato_batch_freezer",
                    "sort_order": 2,
                    "route_category_name": "Gelato Batch Freezer",
                    "route_category_slug": "gelato-batch-freezer",
                },
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "subject_key": "roll_ice_cream_machine",
                    "sort_order": 3,
                    "route_category_name": "Roll Ice Cream Machine",
                    "route_category_slug": "roll-ice-cream-machine",
                },
            ],
            rows=[
                {
                    "category_name": "Ice Cream Machine",
                    "category_slug": "ice-cream-machine",
                    "module_title": "Updated Route Comparison",
                    "module_intro": "Updated intro",
                    "row_key": "operator_interaction",
                    "sort_order": 10,
                    "decision_dimension": "Operator Interaction",
                    "soft_ice_cream_machine": "Monitor mix, dispense product, and keep service moving.",
                    "gelato_batch_freezer": "Load mix, run the cycle, extract product, and manage the next batch.",
                    "roll_ice_cream_machine": "Spread, chop, scrape, and plate each order by hand.",
                },
            ],
        )

        try:
            call_command("import_category_comparison_overview", excel=workbook_path)
        finally:
            Path(workbook_path).unlink(missing_ok=True)

        overview.refresh_from_db()
        rows = list(overview.rows.order_by("sort_order"))
        self.assertEqual(overview.title, "Updated Route Comparison")
        self.assertEqual(overview.intro, "Updated intro")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].row_key, "operator_interaction")
        self.assertNotEqual(rows[0].label, "Legacy Row")
