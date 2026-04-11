"""
Django 管理命令：从产品媒体工作簿导入 MediaAsset / ProductMedia。

使用方式（在 backend/ 目录下）：
    python manage.py import_product_media
    python manage.py import_product_media --excel "F:/菱威仓管/protaylor_product_media_import.xlsx"
    python manage.py import_product_media --dry-run

行为：
    - 只读取 MediaAssets / ProductMedia 两个 sheet
    - 默认导入源固定为 protaylor_product_media_import.xlsx
    - MediaAsset 以 file_url 为幂等键，命中后更新元数据
    - Excel 里的空 Alt Text 不会清空数据库已有 alt_text
    - ProductMedia 采用“整产品覆盖”，只覆盖工作簿里出现过的产品
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_product_media_import.xlsx")
SHEET_MEDIA_ASSETS = "MediaAssets"
SHEET_PRODUCT_MEDIA = "ProductMedia"

MEDIA_ASSET_HEADERS = (
    "Asset File URL",
    "Asset Kind Code",
    "Title",
    "Alt Text",
    "Mime Type",
    "Width",
    "Height",
)
PRODUCT_MEDIA_HEADERS = (
    "Category",
    "Subcategory",
    "Product Name",
    "Model Code",
    "Source Product URL",
    "Asset File URL",
    "Media Kind Code",
    "Is Primary",
    "Alt Override",
    "Sort Order",
)

MAX_URL_LENGTH = 200
MAX_TITLE_LENGTH = 255
MAX_ALT_LENGTH = 255
MAX_MIME_TYPE_LENGTH = 120


@dataclass(frozen=True)
class AssetImportRecord:
    file_url: str
    asset_kind_code: str
    title: str
    alt_text: str
    mime_type: str
    width: int
    height: int


@dataclass(frozen=True)
class ProductSheetKey:
    category_name: str
    product_name: str


@dataclass(frozen=True)
class ProductMediaImportRecord:
    key: ProductSheetKey
    model_code: str
    source_product_url: str
    asset_file_url: str
    media_kind_code: str
    is_primary: bool
    alt_override: str
    sort_order: int


class Command(BaseCommand):
    help = "Import MediaAsset/ProductMedia from protaylor_product_media_import.xlsx."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help="Path to the media import workbook. Defaults to protaylor_product_media_import.xlsx.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import Product, ProductCategory, ProductMedia
        from apps.core.models import MediaAsset

        excel_path = self._resolve_excel_path(options.get("excel"))
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)

        asset_rows = self._load_sheet_records(workbook, SHEET_MEDIA_ASSETS, MEDIA_ASSET_HEADERS)
        product_media_rows = self._load_sheet_records(
            workbook,
            SHEET_PRODUCT_MEDIA,
            PRODUCT_MEDIA_HEADERS,
        )

        asset_records = self._build_asset_records(asset_rows)
        product_media_by_product = self._build_product_media_records(
            product_media_rows,
            asset_records,
        )

        total_asset_rows = len(asset_records)
        total_product_media_rows = sum(len(rows) for rows in product_media_by_product.values())

        self.stdout.write(f"MediaAssets rows: {total_asset_rows}")
        self.stdout.write(f"ProductMedia rows: {total_product_media_rows}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no database writes."))

        existing_assets = {
            asset.file_url: asset
            for asset in MediaAsset.objects.filter(file_url__in=list(asset_records.keys()))
        }
        asset_stats = self._plan_asset_changes(existing_assets, asset_records)

        category_names = {product_key.category_name for product_key in product_media_by_product}
        categories_by_name: dict[str, list[ProductCategory]] = defaultdict(list)
        for category in ProductCategory.objects.filter(name__in=category_names):
            categories_by_name[category.name].append(category)

        unique_category_ids = [
            matches[0].id for matches in categories_by_name.values() if len(matches) == 1
        ]
        product_names = {product_key.product_name for product_key in product_media_by_product}
        products_by_key: dict[tuple[int, str], list[Product]] = defaultdict(list)
        for product in Product.objects.filter(
            category_id__in=unique_category_ids,
            name__in=product_names,
        ).select_related("category"):
            products_by_key[(product.category_id, product.name)].append(product)

        matched_products = 0
        replaced_products = 0
        skipped_products = 0
        unmatched_products: list[str] = []
        ambiguous_categories: list[str] = []
        ambiguous_products: list[str] = []
        model_code_warnings: list[str] = []

        matched_product_payloads: list[tuple[Product, list[ProductMediaImportRecord]]] = []

        for product_key in sorted(
            product_media_by_product.keys(),
            key=lambda item: (item.category_name.lower(), item.product_name.lower()),
        ):
            category_matches = categories_by_name.get(product_key.category_name, [])
            if len(category_matches) == 0:
                skipped_products += 1
                unmatched_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：未找到分类"
                )
                continue
            if len(category_matches) > 1:
                skipped_products += 1
                ambiguous_categories.append(
                    f"{product_key.category_name} / {product_key.product_name}：分类名重复，无法唯一匹配"
                )
                continue

            category = category_matches[0]
            product_matches = products_by_key.get((category.id, product_key.product_name), [])
            if len(product_matches) == 0:
                skipped_products += 1
                unmatched_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：未找到产品"
                )
                continue
            if len(product_matches) > 1:
                skipped_products += 1
                ambiguous_products.append(
                    f"{product_key.category_name} / {product_key.product_name}：产品名重复，无法唯一匹配"
                )
                continue

            product = product_matches[0]
            matched_products += 1
            media_rows = product_media_by_product[product_key]
            workbook_model_code = next((row.model_code for row in media_rows if row.model_code), "")
            if workbook_model_code and workbook_model_code != (product.model_code or ""):
                model_code_warnings.append(
                    f"{product.category.name} / {product.name}：Excel Model Code={workbook_model_code!r}，"
                    f"数据库 Model Code={(product.model_code or '')!r}"
                )

            matched_product_payloads.append((product, media_rows))

        if not dry_run:
            # 这里用一个总事务包住“资产更新 + 产品媒体覆盖”。
            # 原因很直接：这两张 sheet 是同一份导入契约，任何一步中途失败都不应该留下半套数据。
            with transaction.atomic():
                resolved_assets = self._apply_asset_changes(existing_assets, asset_records)
                for product, media_rows in matched_product_payloads:
                    # ProductMedia 采用整产品覆盖，而不是增量 merge。
                    # 这能保证数据库与工作簿的产品图库关系完全一致，不残留历史主图/旧图库脏数据。
                    product.media_items.all().delete()
                    for media_row in media_rows:
                        ProductMedia.objects.create(
                            product=product,
                            asset=resolved_assets[media_row.asset_file_url],
                            media_kind=self._map_media_kind_code(media_row.media_kind_code),
                            is_primary=media_row.is_primary,
                            alt_override=media_row.alt_override,
                            sort_order=media_row.sort_order,
                        )
                    replaced_products += 1
        else:
            replaced_products = matched_products

        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(
            f"  Assets: create={asset_stats['created']} update={asset_stats['updated']} skip={asset_stats['skipped']}"
        )
        self.stdout.write(
            f"  Products: matched={matched_products} replaced={replaced_products} skipped={skipped_products}"
        )

        if unmatched_products:
            self.stdout.write(self.style.WARNING("\nUnmatched products"))
            for item in unmatched_products[:100]:
                self.stdout.write(f"  - {item}")
            if len(unmatched_products) > 100:
                self.stdout.write(f"  ... {len(unmatched_products) - 100} more")

        if ambiguous_categories:
            self.stdout.write(self.style.WARNING("\nAmbiguous categories"))
            for item in ambiguous_categories[:100]:
                self.stdout.write(f"  - {item}")
            if len(ambiguous_categories) > 100:
                self.stdout.write(f"  ... {len(ambiguous_categories) - 100} more")

        if ambiguous_products:
            self.stdout.write(self.style.WARNING("\nAmbiguous products"))
            for item in ambiguous_products[:100]:
                self.stdout.write(f"  - {item}")
            if len(ambiguous_products) > 100:
                self.stdout.write(f"  ... {len(ambiguous_products) - 100} more")

        if model_code_warnings:
            self.stdout.write(self.style.WARNING("\nModel code warnings"))
            for item in model_code_warnings[:100]:
                self.stdout.write(f"  - {item}")
            if len(model_code_warnings) > 100:
                self.stdout.write(f"  ... {len(model_code_warnings) - 100} more")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook,
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, object]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [_clean_text(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        column_indexes = {header: headers.index(header) for header in required_headers}
        records: list[dict[str, object]] = []
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            record = {
                header: row[column_indexes[header]]
                if column_indexes[header] < len(row)
                else None
                for header in required_headers
            }
            if all(_clean_text(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_index
            records.append(record)

        return records

    def _build_asset_records(
        self,
        raw_records: list[dict[str, object]],
    ) -> dict[str, AssetImportRecord]:
        asset_records: dict[str, AssetImportRecord] = {}
        duplicate_urls: list[str] = []

        for raw_record in raw_records:
            row_number = int(raw_record["__row_number__"])
            file_url = _clean_text(raw_record["Asset File URL"])
            if not file_url:
                raise CommandError(f"MediaAssets row {row_number}: Asset File URL 不能为空。")
            if len(file_url) > MAX_URL_LENGTH:
                raise CommandError(
                    f"MediaAssets row {row_number}: Asset File URL 超过 {MAX_URL_LENGTH} 字符。"
                )

            if file_url in asset_records:
                duplicate_urls.append(file_url)
                continue

            asset_kind_code = _clean_text(raw_record["Asset Kind Code"]).lower()
            if asset_kind_code != "image":
                raise CommandError(
                    f"MediaAssets row {row_number}: Asset Kind Code 只允许 'image'，收到 {asset_kind_code!r}。"
                )

            title = _clean_text(raw_record["Title"])
            if not title:
                raise CommandError(f"MediaAssets row {row_number}: Title 不能为空。")
            if len(title) > MAX_TITLE_LENGTH:
                raise CommandError(
                    f"MediaAssets row {row_number}: Title 超过 {MAX_TITLE_LENGTH} 字符。"
                )

            alt_text = _clean_text(raw_record["Alt Text"])
            if len(alt_text) > MAX_ALT_LENGTH:
                raise CommandError(
                    f"MediaAssets row {row_number}: Alt Text 超过 {MAX_ALT_LENGTH} 字符。"
                )

            mime_type = _clean_text(raw_record["Mime Type"])
            if not mime_type:
                raise CommandError(f"MediaAssets row {row_number}: Mime Type 不能为空。")
            if len(mime_type) > MAX_MIME_TYPE_LENGTH:
                raise CommandError(
                    f"MediaAssets row {row_number}: Mime Type 超过 {MAX_MIME_TYPE_LENGTH} 字符。"
                )

            width = self._parse_positive_int(
                raw_record["Width"],
                field_name="Width",
                sheet_name=SHEET_MEDIA_ASSETS,
                row_number=row_number,
            )
            height = self._parse_positive_int(
                raw_record["Height"],
                field_name="Height",
                sheet_name=SHEET_MEDIA_ASSETS,
                row_number=row_number,
            )

            asset_records[file_url] = AssetImportRecord(
                file_url=file_url,
                asset_kind_code=asset_kind_code,
                title=title,
                alt_text=alt_text,
                mime_type=mime_type,
                width=width,
                height=height,
            )

        if duplicate_urls:
            preview = ", ".join(duplicate_urls[:10])
            more = "" if len(duplicate_urls) <= 10 else f" ... and {len(duplicate_urls) - 10} more"
            raise CommandError(f"MediaAssets contains duplicate Asset File URL values: {preview}{more}")

        return asset_records

    def _build_product_media_records(
        self,
        raw_records: list[dict[str, object]],
        asset_records: dict[str, AssetImportRecord],
    ) -> dict[ProductSheetKey, list[ProductMediaImportRecord]]:
        product_media_by_product: dict[ProductSheetKey, list[ProductMediaImportRecord]] = defaultdict(list)
        invalid_asset_refs: list[str] = []
        duplicate_product_asset_refs: list[str] = []
        product_asset_sets: dict[ProductSheetKey, set[str]] = defaultdict(set)

        for raw_record in raw_records:
            row_number = int(raw_record["__row_number__"])
            category = _clean_text(raw_record["Category"])
            subcategory = _clean_text(raw_record["Subcategory"])
            product_name = _clean_text(raw_record["Product Name"])
            if not category:
                raise CommandError(f"ProductMedia row {row_number}: Category 不能为空。")
            if not product_name:
                raise CommandError(f"ProductMedia row {row_number}: Product Name 不能为空。")

            product_key = ProductSheetKey(
                category_name=_resolve_target_category_name(category, subcategory),
                product_name=product_name,
            )

            asset_file_url = _clean_text(raw_record["Asset File URL"])
            if asset_file_url not in asset_records:
                invalid_asset_refs.append(
                    f"row {row_number} -> {product_key.category_name} / {product_key.product_name} / {asset_file_url}"
                )
                continue

            if asset_file_url in product_asset_sets[product_key]:
                duplicate_product_asset_refs.append(
                    f"row {row_number} -> {product_key.category_name} / {product_key.product_name} / {asset_file_url}"
                )
                continue
            product_asset_sets[product_key].add(asset_file_url)

            media_kind_code = _clean_text(raw_record["Media Kind Code"]).lower()
            if media_kind_code not in {"hero", "gallery"}:
                raise CommandError(
                    f"ProductMedia row {row_number}: Media Kind Code 只允许 'hero' 或 'gallery'，"
                    f"收到 {media_kind_code!r}。"
                )

            alt_override = _clean_text(raw_record["Alt Override"])
            if len(alt_override) > MAX_ALT_LENGTH:
                raise CommandError(
                    f"ProductMedia row {row_number}: Alt Override 超过 {MAX_ALT_LENGTH} 字符。"
                )

            product_media_by_product[product_key].append(
                ProductMediaImportRecord(
                    key=product_key,
                    model_code=_clean_text(raw_record["Model Code"]),
                    source_product_url=_clean_text(raw_record["Source Product URL"]),
                    asset_file_url=asset_file_url,
                    media_kind_code=media_kind_code,
                    is_primary=self._parse_bool(
                        raw_record["Is Primary"],
                        field_name="Is Primary",
                        sheet_name=SHEET_PRODUCT_MEDIA,
                        row_number=row_number,
                    ),
                    alt_override=alt_override,
                    sort_order=self._parse_int(
                        raw_record["Sort Order"],
                        field_name="Sort Order",
                        sheet_name=SHEET_PRODUCT_MEDIA,
                        row_number=row_number,
                    ),
                )
            )

        if invalid_asset_refs:
            preview = "; ".join(invalid_asset_refs[:10])
            more = "" if len(invalid_asset_refs) <= 10 else f" ... and {len(invalid_asset_refs) - 10} more"
            raise CommandError(f"ProductMedia contains Asset File URL values not found in MediaAssets: {preview}{more}")

        if duplicate_product_asset_refs:
            preview = "; ".join(duplicate_product_asset_refs[:10])
            more = (
                ""
                if len(duplicate_product_asset_refs) <= 10
                else f" ... and {len(duplicate_product_asset_refs) - 10} more"
            )
            raise CommandError(f"ProductMedia contains duplicate product/asset rows: {preview}{more}")

        for product_key, rows in product_media_by_product.items():
            primary_rows = [row for row in rows if row.is_primary]
            hero_rows = [row for row in rows if row.media_kind_code == "hero"]

            if len(primary_rows) != 1:
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}: "
                    f"必须恰好有 1 条 is_primary=True，当前为 {len(primary_rows)}。"
                )
            if len(hero_rows) != 1:
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}: "
                    f"必须恰好有 1 条 hero 媒体，当前为 {len(hero_rows)}。"
                )
            if primary_rows[0].media_kind_code != "hero":
                raise CommandError(
                    f"{product_key.category_name} / {product_key.product_name}: "
                    "主图行必须同时标记为 hero。"
                )

            rows.sort(key=lambda item: (item.sort_order, item.asset_file_url))

        return product_media_by_product

    def _plan_asset_changes(self, existing_assets, asset_records) -> dict[str, int]:
        stats = {"created": 0, "updated": 0, "skipped": 0}
        for file_url, record in asset_records.items():
            existing_asset = existing_assets.get(file_url)
            if existing_asset is None:
                stats["created"] += 1
                continue

            update_fields = self._compute_asset_update_fields(existing_asset, record)
            if update_fields:
                stats["updated"] += 1
            else:
                stats["skipped"] += 1

        return stats

    def _apply_asset_changes(self, existing_assets, asset_records):
        from apps.core.models import MediaAsset

        resolved_assets = dict(existing_assets)
        for file_url, record in asset_records.items():
            existing_asset = resolved_assets.get(file_url)
            if existing_asset is None:
                asset = MediaAsset.objects.create(
                    title=record.title,
                    asset_kind=MediaAsset.AssetKind.IMAGE,
                    file_url=record.file_url,
                    alt_text=record.alt_text,
                    mime_type=record.mime_type,
                    width=record.width,
                    height=record.height,
                )
                resolved_assets[file_url] = asset
                continue

            update_fields = self._compute_asset_update_fields(existing_asset, record)
            if update_fields:
                existing_asset.save(update_fields=update_fields)

        return resolved_assets

    def _compute_asset_update_fields(self, asset, record: AssetImportRecord) -> list[str]:
        from apps.core.models import MediaAsset

        update_fields: list[str] = []
        if asset.asset_kind != MediaAsset.AssetKind.IMAGE:
            asset.asset_kind = MediaAsset.AssetKind.IMAGE
            update_fields.append("asset_kind")
        if asset.title != record.title:
            asset.title = record.title
            update_fields.append("title")
        if asset.mime_type != record.mime_type:
            asset.mime_type = record.mime_type
            update_fields.append("mime_type")
        if asset.width != record.width:
            asset.width = record.width
            update_fields.append("width")
        if asset.height != record.height:
            asset.height = record.height
            update_fields.append("height")

        # 这里故意不把空 Alt Text 回写成空字符串。
        # 这份工作簿的资产层 alt 目前是留空设计，若未来后台人工补录了更好的 alt，
        # 重跑导入不应该把那部分人工资产语义冲掉。
        if record.alt_text and asset.alt_text != record.alt_text:
            asset.alt_text = record.alt_text
            update_fields.append("alt_text")

        return update_fields

    def _parse_bool(
        self,
        value: object,
        *,
        field_name: str,
        sheet_name: str,
        row_number: int,
    ) -> bool:
        text = _clean_text(value).lower()
        if text in {"true", "1", "yes", "y"}:
            return True
        if text in {"false", "0", "no", "n"}:
            return False
        raise CommandError(
            f"{sheet_name} row {row_number}: {field_name} 无法解析为布尔值，收到 {value!r}。"
        )

    def _parse_int(
        self,
        value: object,
        *,
        field_name: str,
        sheet_name: str,
        row_number: int,
    ) -> int:
        text = _clean_text(value)
        if not text:
            raise CommandError(f"{sheet_name} row {row_number}: {field_name} 不能为空。")
        try:
            return int(float(text))
        except (TypeError, ValueError) as exc:
            raise CommandError(
                f"{sheet_name} row {row_number}: {field_name} 无法解析为整数，收到 {value!r}。"
            ) from exc

    def _parse_positive_int(
        self,
        value: object,
        *,
        field_name: str,
        sheet_name: str,
        row_number: int,
    ) -> int:
        parsed = self._parse_int(
            value,
            field_name=field_name,
            sheet_name=sheet_name,
            row_number=row_number,
        )
        if parsed <= 0:
            raise CommandError(
                f"{sheet_name} row {row_number}: {field_name} 必须大于 0，收到 {parsed!r}。"
            )
        return parsed

    def _map_media_kind_code(self, media_kind_code: str) -> int:
        from apps.catalog.models import ProductMedia

        mapping = {
            "hero": ProductMedia.MediaKind.HERO,
            "gallery": ProductMedia.MediaKind.GALLERY,
        }
        return mapping[media_kind_code]


def _resolve_target_category_name(category: str, subcategory: str) -> str:
    if subcategory and subcategory != category:
        return subcategory
    return category


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
