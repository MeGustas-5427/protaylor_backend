"""
Import category-level Comparison Overview seed rows from Excel.

Usage:
    python manage.py import_category_comparison_overview
    python manage.py import_category_comparison_overview --excel "F:/菱威仓管/protaylor_comparison_overview_staging.xlsx"
    python manage.py import_category_comparison_overview --dry-run
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


DEFAULT_EXCEL_PATH = Path(r"F:\菱威仓管\protaylor_comparison_overview_staging.xlsx")
DEFAULT_SUBJECTS_SHEET_NAME = "Comparison Overview Subjects"
DEFAULT_DRAFT_SHEET_NAME = "Comparison Overview Draft"
SUBJECT_HEADERS = (
    "category_name",
    "category_slug",
    "subject_key",
    "sort_order",
    "route_category_name",
    "route_category_slug",
    "label_override",
    "status",
    "source_doc",
    "notes",
)
DRAFT_FIXED_HEADERS = (
    "category_name",
    "category_slug",
    "module_type",
    "module_title",
    "module_intro",
    "row_key",
    "sort_order",
    "decision_dimension",
    "status",
    "source_doc",
)
FINALIZED_ROW_STATUSES = {"finalized_for_staging", "approved_for_staging", "approved"}
NORMALIZED_DIMENSION_HEADING_HEADER = "dimension_heading"
LEGACY_DIMENSION_HEADING_HEADER = "column_1_label"
SUBJECT_SLOT_KEY_PATTERN = re.compile(r"^subject_(\d+)_key$")


@dataclass(frozen=True)
class SubjectSeed:
    category_name: str
    category_slug: str
    subject_key: str
    sort_order: int
    route_category_name: str
    route_category_slug: str
    label_override: str

    @property
    def label(self) -> str:
        return self.label_override or self.route_category_name


@dataclass(frozen=True)
class DraftSeedRow:
    category_name: str
    category_slug: str
    module_title: str
    module_intro: str
    dimension_heading: str
    row_key: str
    sort_order: int
    label: str
    cells_json: dict[str, str]


@dataclass(frozen=True)
class PreparedOverviewSeed:
    category: Any
    title: str
    intro: str
    dimension_heading: str
    subjects_json: list[dict[str, Any]]
    rows: list[DraftSeedRow]


class Command(BaseCommand):
    help = "Import ProductCategoryComparisonOverview rows from workbook staging sheets."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            default=None,
            help=(
                "Path to protaylor_comparison_overview_staging.xlsx. "
                "Defaults to F:/菱威仓管/protaylor_comparison_overview_staging.xlsx."
            ),
        )
        parser.add_argument(
            "--subjects-sheet",
            default=DEFAULT_SUBJECTS_SHEET_NAME,
            help=f"Worksheet name for subjects. Defaults to {DEFAULT_SUBJECTS_SHEET_NAME!r}.",
        )
        parser.add_argument(
            "--draft-sheet",
            default=DEFAULT_DRAFT_SHEET_NAME,
            help=f"Worksheet name for draft rows. Defaults to {DEFAULT_DRAFT_SHEET_NAME!r}.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database.",
        )

    def handle(self, *args, **options) -> None:
        from openpyxl import load_workbook

        from apps.catalog.models import (
            ProductCategory,
            ProductCategoryComparisonOverview,
            ProductCategoryComparisonRow,
        )

        excel_path = self._resolve_excel_path(options.get("excel"))
        subjects_sheet_name = self._clean(options.get("subjects_sheet")) or DEFAULT_SUBJECTS_SHEET_NAME
        draft_sheet_name = self._clean(options.get("draft_sheet")) or DEFAULT_DRAFT_SHEET_NAME
        dry_run = bool(options["dry_run"])

        self.stdout.write(f"Reading: {excel_path}")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            subject_records = self._load_sheet_records(workbook, subjects_sheet_name, SUBJECT_HEADERS)
            draft_records = self._load_sheet_records(workbook, draft_sheet_name, DRAFT_FIXED_HEADERS)
        finally:
            workbook.close()

        subject_groups = self._build_subject_groups(subject_records)
        draft_groups = self._build_draft_groups(draft_records, subject_groups)

        self.stdout.write(f"Workbook groups: {len(draft_groups)}")
        self.stdout.write(f"Workbook rows: {sum(len(rows) for rows in draft_groups.values())}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes."))

        categories_by_slug = {
            category.slug: category
            for category in ProductCategory.objects.select_related("parent")
        }

        prepared_groups: list[PreparedOverviewSeed] = []
        unmatched_groups: list[str] = []

        for category_slug, rows in draft_groups.items():
            subjects = subject_groups.get(category_slug)
            if not subjects:
                raise CommandError(
                    f"{category_slug}: draft rows exist but Comparison Overview Subjects rows are missing."
                )

            category = categories_by_slug.get(category_slug)
            if category is None:
                unmatched_groups.append(f"{category_slug}: target category slug not found")
                continue

            if rows[0].category_name != category.name:
                raise CommandError(
                    f"{category_slug}: draft category_name {rows[0].category_name!r} does not match "
                    f"database category name {category.name!r}."
                )

            group_errors: list[str] = []
            subjects_json: list[dict[str, Any]] = []
            for subject in subjects:
                route_category = categories_by_slug.get(subject.route_category_slug)
                if route_category is None:
                    group_errors.append(
                        f"{category_slug}: route_category_slug {subject.route_category_slug!r} not found"
                    )
                    continue
                if route_category.parent_id != category.id:
                    group_errors.append(
                        f"{category_slug}: route_category_slug {subject.route_category_slug!r} "
                        f"is not a direct child of {category.slug!r}"
                    )
                    continue
                if route_category.name != subject.route_category_name:
                    raise CommandError(
                        f"{category_slug}: route_category_name {subject.route_category_name!r} does not match "
                        f"database category name {route_category.name!r} for slug {subject.route_category_slug!r}."
                    )

                subjects_json.append(
                    {
                        "subject_key": subject.subject_key,
                        "label": subject.label,
                        "route_category_slug": subject.route_category_slug,
                        "sort_order": subject.sort_order,
                    }
                )

            if group_errors:
                unmatched_groups.extend(group_errors)
                continue

            overview = ProductCategoryComparisonOverview(
                category=category,
                title=rows[0].module_title,
                intro=rows[0].module_intro,
                dimension_heading=rows[0].dimension_heading,
                subjects_json=subjects_json,
                is_active=True,
            )
            overview.full_clean(validate_unique=False)

            prepared_groups.append(
                PreparedOverviewSeed(
                    category=category,
                    title=overview.title,
                    intro=overview.intro,
                    dimension_heading=overview.dimension_heading,
                    subjects_json=overview.subjects_json,
                    rows=rows,
                )
            )

        if unmatched_groups:
            raise CommandError(
                "Import aborted because some comparison seed groups did not match ProductCategory rows:\n- "
                + "\n- ".join(unmatched_groups)
            )

        if not dry_run:
            with transaction.atomic():
                for group in prepared_groups:
                    overview, _ = ProductCategoryComparisonOverview.objects.update_or_create(
                        category=group.category,
                        defaults={
                            "title": group.title,
                            "intro": group.intro,
                            "dimension_heading": group.dimension_heading,
                            "subjects_json": group.subjects_json,
                            "is_active": True,
                        },
                    )
                    ProductCategoryComparisonRow.objects.filter(overview=overview).delete()

                    row_models: list[ProductCategoryComparisonRow] = []
                    for row in group.rows:
                        row_model = ProductCategoryComparisonRow(
                            overview=overview,
                            row_key=row.row_key,
                            label=row.label,
                            sort_order=row.sort_order,
                            cells_json=row.cells_json,
                            is_active=True,
                        )
                        row_model.full_clean()
                        row_models.append(row_model)
                    ProductCategoryComparisonRow.objects.bulk_create(row_models)

        imported_rows = sum(len(group.rows) for group in prepared_groups)
        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"  Imported groups: {len(prepared_groups)}")
        self.stdout.write(f"  Imported rows: {imported_rows}")
        self.stdout.write(f"  Dry run: {dry_run}")

    def _resolve_excel_path(self, excel_arg: str | None) -> Path:
        excel_path = Path(excel_arg) if excel_arg else DEFAULT_EXCEL_PATH
        if not excel_path.exists():
            raise CommandError(f"Workbook not found: {excel_path}")
        return excel_path

    def _load_sheet_records(
        self,
        workbook: Any,
        sheet_name: str,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Workbook must contain sheet {sheet_name!r}; available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration as exc:
            raise CommandError(f"Sheet {sheet_name!r} is empty.") from exc

        headers = [self._clean(value) for value in header_row]
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            raise CommandError(
                f"Sheet {sheet_name!r} is missing required headers: {', '.join(missing_headers)}"
            )

        records: list[dict[str, Any]] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
            if all(self._clean(value) == "" for value in record.values()):
                continue
            record["__row_number__"] = row_number
            records.append(record)
        return records

    def _build_subject_groups(self, subject_records: list[dict[str, Any]]) -> dict[str, list[SubjectSeed]]:
        subject_groups: dict[str, list[SubjectSeed]] = defaultdict(list)
        for record in subject_records:
            row_number = int(record["__row_number__"])
            status = self._clean(record["status"]).lower()
            if status not in {"active", "1", "true", "yes"}:
                continue

            category_name = self._clean(record["category_name"])
            category_slug = self._clean(record["category_slug"])
            subject_key = self._clean(record["subject_key"])
            route_category_name = self._clean(record["route_category_name"])
            route_category_slug = self._clean(record["route_category_slug"])
            label_override = self._clean(record["label_override"])
            sort_order = self._parse_int(record["sort_order"], "sort_order", row_number)

            if not category_name or not category_slug:
                raise CommandError(
                    f"Comparison Overview Subjects row {row_number}: category_name and category_slug are required."
                )
            if not subject_key:
                raise CommandError(
                    f"Comparison Overview Subjects row {row_number}: subject_key is required."
                )
            if not route_category_name or not route_category_slug:
                raise CommandError(
                    f"Comparison Overview Subjects row {row_number}: route category name and slug are required."
                )

            subject_groups[category_slug].append(
                SubjectSeed(
                    category_name=category_name,
                    category_slug=category_slug,
                    subject_key=subject_key,
                    sort_order=sort_order,
                    route_category_name=route_category_name,
                    route_category_slug=route_category_slug,
                    label_override=label_override,
                )
            )

        for category_slug, subjects in subject_groups.items():
            if len({subject.category_name for subject in subjects}) != 1:
                raise CommandError(f"{category_slug}: inconsistent category_name values in subject sheet.")
            if len({subject.subject_key for subject in subjects}) != len(subjects):
                raise CommandError(f"{category_slug}: duplicate subject_key values in subject sheet.")
            if len({subject.sort_order for subject in subjects}) != len(subjects):
                raise CommandError(f"{category_slug}: duplicate sort_order values in subject sheet.")
            subject_groups[category_slug] = sorted(subjects, key=lambda subject: (subject.sort_order, subject.subject_key))
        return subject_groups

    def _build_draft_groups(
        self,
        draft_records: list[dict[str, Any]],
        subject_groups: dict[str, list[SubjectSeed]],
    ) -> dict[str, list[DraftSeedRow]]:
        grouped_rows: dict[str, list[DraftSeedRow]] = defaultdict(list)

        for record in draft_records:
            row_number = int(record["__row_number__"])
            category_name = self._clean(record["category_name"])
            category_slug = self._clean(record["category_slug"])
            module_type = self._clean(record["module_type"])
            status = self._clean(record["status"]).lower()

            if module_type != "comparison_overview":
                continue
            if status not in FINALIZED_ROW_STATUSES:
                continue

            subjects = subject_groups.get(category_slug)
            if not subjects:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: no subject mapping found for category {category_slug!r}."
                )

            module_title = self._clean(record["module_title"])
            module_intro = self._clean(record["module_intro"])
            row_key = self._clean(record["row_key"])
            decision_dimension = self._clean(record["decision_dimension"])
            sort_order = self._parse_int(record["sort_order"], "sort_order", row_number)

            if not category_name or not category_slug:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: category_name and category_slug are required."
                )
            if not module_title:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: module_title is required."
                )
            if not row_key or not decision_dimension:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: row_key and decision_dimension are required."
                )

            if NORMALIZED_DIMENSION_HEADING_HEADER in record:
                dimension_heading, cells_json = self._build_normalized_cells_payload(
                    record=record,
                    row_number=row_number,
                    category_slug=category_slug,
                    subjects=subjects,
                )
            else:
                dimension_heading, cells_json = self._build_legacy_cells_payload(
                    record=record,
                    row_number=row_number,
                    category_slug=category_slug,
                    subjects=subjects,
                )

            grouped_rows[category_slug].append(
                DraftSeedRow(
                    category_name=category_name,
                    category_slug=category_slug,
                    module_title=module_title,
                    module_intro=module_intro,
                    dimension_heading=dimension_heading,
                    row_key=row_key,
                    sort_order=sort_order,
                    label=decision_dimension,
                    cells_json=cells_json,
                )
            )

        for category_slug, rows in grouped_rows.items():
            if len({row.category_name for row in rows}) != 1:
                raise CommandError(f"{category_slug}: inconsistent category_name values in draft sheet.")
            if len({row.module_title for row in rows}) != 1:
                raise CommandError(f"{category_slug}: inconsistent module_title values in draft sheet.")
            if len({row.module_intro for row in rows}) != 1:
                raise CommandError(f"{category_slug}: inconsistent module_intro values in draft sheet.")
            if len({row.dimension_heading for row in rows}) != 1:
                raise CommandError(f"{category_slug}: inconsistent dimension heading values in draft sheet.")
            if len({row.row_key for row in rows}) != len(rows):
                raise CommandError(f"{category_slug}: duplicate row_key values in draft sheet.")
            if len({row.sort_order for row in rows}) != len(rows):
                raise CommandError(f"{category_slug}: duplicate sort_order values in draft sheet.")
            grouped_rows[category_slug] = sorted(rows, key=lambda row: (row.sort_order, row.row_key))

        return grouped_rows

    def _build_normalized_cells_payload(
        self,
        *,
        record: dict[str, Any],
        row_number: int,
        category_slug: str,
        subjects: list[SubjectSeed],
    ) -> tuple[str, dict[str, str]]:
        dimension_heading = self._clean(record.get(NORMALIZED_DIMENSION_HEADING_HEADER))
        if not dimension_heading:
            raise CommandError(
                f"Comparison Overview Draft row {row_number}: {NORMALIZED_DIMENSION_HEADING_HEADER} is required."
            )

        cells_json: dict[str, str] = {}
        for slot_index, subject in enumerate(subjects, start=1):
            key_header = f"subject_{slot_index}_key"
            label_header = f"subject_{slot_index}_label"
            text_header = f"subject_{slot_index}_text"

            for header in (key_header, label_header, text_header):
                if header not in record:
                    raise CommandError(
                        f"Comparison Overview Draft row {row_number}: missing header {header!r}."
                    )

            slot_subject_key = self._clean(record[key_header])
            if slot_subject_key != subject.subject_key:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: {key_header} "
                    f"must equal {subject.subject_key!r} for category {category_slug!r}."
                )

            label_value = self._clean(record[label_header])
            if label_value != subject.label:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: {label_header} "
                    f"must equal {subject.label!r} for category {category_slug!r}."
                )

            body = self._clean(record[text_header])
            if not body:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: {text_header} is required."
                )
            cells_json[subject.subject_key] = body

        expected_slot_count = len(subjects)
        for slot_index in self._list_subject_slot_indexes(record):
            if slot_index <= expected_slot_count:
                continue

            extra_values = [
                self._clean(record.get(f"subject_{slot_index}_key")),
                self._clean(record.get(f"subject_{slot_index}_label")),
                self._clean(record.get(f"subject_{slot_index}_text")),
            ]
            if any(extra_values):
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: subject slot {slot_index} "
                    "contains values but no active subject mapping exists for it."
                )

        return dimension_heading, cells_json

    def _build_legacy_cells_payload(
        self,
        *,
        record: dict[str, Any],
        row_number: int,
        category_slug: str,
        subjects: list[SubjectSeed],
    ) -> tuple[str, dict[str, str]]:
        dimension_heading = self._clean(record.get(LEGACY_DIMENSION_HEADING_HEADER))
        if not dimension_heading:
            raise CommandError(
                f"Comparison Overview Draft row {row_number}: {LEGACY_DIMENSION_HEADING_HEADER} is required."
            )

        cells_json: dict[str, str] = {}
        for subject in subjects:
            label_column = f"column_{subject.sort_order + 1}_label"
            if label_column not in record:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: missing header {label_column!r}."
                )
            label_value = self._clean(record[label_column])
            if label_value != subject.label:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: {label_column} "
                    f"must equal {subject.label!r} for category {category_slug!r}."
                )
            if subject.subject_key not in record:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: missing subject column {subject.subject_key!r}."
                )
            body = self._clean(record[subject.subject_key])
            if not body:
                raise CommandError(
                    f"Comparison Overview Draft row {row_number}: subject column {subject.subject_key!r} is required."
                )
            cells_json[subject.subject_key] = body

        return dimension_heading, cells_json

    @staticmethod
    def _list_subject_slot_indexes(record: dict[str, Any]) -> list[int]:
        indexes: set[int] = set()
        for header in record:
            match = SUBJECT_SLOT_KEY_PATTERN.match(header)
            if match:
                indexes.add(int(match.group(1)))
        return sorted(indexes)

    @staticmethod
    def _clean(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_int(self, value: Any, field_name: str, row_number: int) -> int:
        cleaned = self._clean(value)
        if not cleaned:
            raise CommandError(f"Comparison Overview row {row_number}: {field_name} is required.")
        try:
            return int(float(cleaned))
        except ValueError as exc:
            raise CommandError(
                f"Comparison Overview row {row_number}: invalid {field_name} value {value!r}."
            ) from exc
